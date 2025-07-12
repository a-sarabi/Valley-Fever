import seaborn as sns
import os
GPU_ID="0"
os.environ["CUDA_VISIBLE_DEVICES"] = GPU_ID
import gc
from tensorflow.keras.backend import clear_session

print(os.environ['PATH'])
#
# from dlib import cuda
# #cuda.set_device(1)
from tensorflow.python.client import device_lib
print(device_lib.list_local_devices())
import os
import openpyxl
import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.model_selection import train_test_split
import keras.backend as K
import tensorflow as tf
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_percentage_error, mean_squared_error, mean_absolute_error
from spektral.layers import GCNConv, SAGPool, MinCutPool, GlobalAttentionPool, JustBalancePool, DiffPool
from spektral.utils import normalized_adjacency
from sklearn.metrics.pairwise import cosine_similarity
from gatv2_conv import GATv2Conv  # Import the GATv2Conv layer


import random
import datetime
import math
#from tensorflow.keras.layers import LSTM, Attention, GlobalAveragePooling1D, Bidirectional
from joblib import Parallel, delayed
from keras.layers import Dense, Dropout, Embedding, LSTM, Bidirectional
from keras.layers import BatchNormalization
from tensorflow.keras.layers import Dropout, Input
from spektral.layers import GATConv
from keras.layers import Add, Concatenate
from tensorflow.keras.layers import Lambda, MultiHeadAttention, LayerNormalization
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras.regularizers import l2

# Add Positional Embedding Layer
class PositionalEmbedding(layers.Layer):
    """
    Adds positional embeddings to the input embeddings.
    """
    def __init__(self, max_sequence_length, d_model, **kwargs):
        super().__init__(**kwargs)
        self.token_embeddings = layers.Dense(d_model)  # Linear projection to d_model dimension
        self.pos_embeddings = layers.Embedding(input_dim=max_sequence_length, output_dim=d_model)
        self.max_sequence_length = max_sequence_length
        self.d_model = d_model

    def call(self, x):
        """
        x shape: (batch_size, sequence_length, input_dim)
        """
        seq_len = tf.shape(x)[1]
        # Ensure positions don't exceed max_sequence_length
        positions = tf.range(start=0, limit=seq_len, delta=1)
        positions = tf.minimum(positions, self.max_sequence_length - 1)
        
        embedded_tokens = self.token_embeddings(x)  # (batch_size, seq_len, d_model)
        embedded_positions = self.pos_embeddings(positions)  # (seq_len, d_model)
        
        # Add positional embeddings to token embeddings
        return embedded_tokens + embedded_positions

# Create a look-ahead mask for decoder self-attention
def create_look_ahead_mask(size):
    """
    Creates a causal/look-ahead mask to prevent decoder from attending to future positions.
    
    Args:
        size: The sequence length for both query and key dimensions
    
    Returns:
        A mask of shape (1, 1, size, size) for use with MultiHeadAttention
    """
    # Create a 2D upper triangular matrix with ones (will be masked positions)
    mask_2d = 1 - tf.linalg.band_part(tf.ones((size, size)), -1, 0)
    
    # Expand dimensions to (1, 1, size, size) for compatibility with MultiHeadAttention
    # This shape represents (batch_size, num_heads, query_seq_len, key_seq_len)
    mask_4d = mask_2d[tf.newaxis, tf.newaxis, :, :]
    
    return mask_4d  # Shape: (1, 1, size, size)

# Add Transformer Encoder Block
class TransformerEncoder(layers.Layer):
    """
    Transformer Encoder block with multi-head attention and feed-forward network.
    """
    def __init__(self, d_model, num_heads, ff_dim, dropout=0.1, **kwargs):
        super().__init__(**kwargs)
        # Self-attention
        self.mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        self.ffn = keras.Sequential([
            layers.Dense(ff_dim, activation='relu'),
            layers.Dense(d_model),
        ])

        self.layernorm1 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm2 = layers.LayerNormalization(epsilon=1e-6)
        self.dropout1 = layers.Dropout(dropout)
        self.dropout2 = layers.Dropout(dropout)

    def call(self, x, training=False, mask=None):
        """
        Process input through the transformer encoder block.
        
        Args:
            x: Input tensor of shape (batch_size, seq_len, d_model)
            training: Whether in training mode
            mask: Optional attention mask of shape (batch_size, 1, 1, seq_len)
        
        Returns:
            Processed tensor of shape (batch_size, seq_len, d_model)
        """
        # Self Attention
        attn_output, _ = self.mha(
            query=x, value=x, key=x, 
            attention_mask=mask,
            return_attention_scores=True
        )
        attn_output = self.dropout1(attn_output, training=training)
        out1 = self.layernorm1(x + attn_output)

        # Feed-Forward
        ffn_output = self.ffn(out1)
        ffn_output = self.dropout2(ffn_output, training=training)
        out2 = self.layernorm2(out1 + ffn_output)

        return out2

# Add Transformer Decoder Block
class TransformerDecoder(layers.Layer):
    """
    Transformer Decoder block with masked self-attention, cross-attention and feed-forward network.
    """
    def __init__(self, d_model, num_heads, ff_dim, dropout=0.1, **kwargs):
        super().__init__(**kwargs)
        # Self-attention (decoder-side)
        self.self_mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        # Cross-attention (encoder-to-decoder)
        self.cross_mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        self.ffn = keras.Sequential([
            layers.Dense(ff_dim, activation='relu'),
            layers.Dense(d_model),
        ])

        self.layernorm1 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm2 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm3 = layers.LayerNormalization(epsilon=1e-6)
        self.dropout1 = layers.Dropout(dropout)
        self.dropout2 = layers.Dropout(dropout)
        self.dropout3 = layers.Dropout(dropout)

    def call(self, x, enc_output, training=False,
             look_ahead_mask=None, padding_mask=None):
        """
        x shape:          (batch_size, target_seq_len, d_model)  # decoder input
        enc_output shape: (batch_size, input_seq_len, d_model)
        look_ahead_mask:  (1, 1, target_seq_len, target_seq_len) or None
        padding_mask:     (batch_size, 1, 1, input_seq_len) or None
        """
        # Apply the look_ahead_mask if not provided but needed
        target_seq_len = tf.shape(x)[1]
        if look_ahead_mask is None:
            look_ahead_mask = create_look_ahead_mask(target_seq_len)
            
        # 1) Masked self-attention
        # The mask shape should be (1, 1, target_seq_len, target_seq_len)
        attn1, attn_weights_1 = self.self_mha(
            query=x, value=x, key=x, 
            attention_mask=look_ahead_mask,
            return_attention_scores=True
        )
        attn1 = self.dropout1(attn1, training=training)
        out1 = self.layernorm1(x + attn1)

        # 2) Cross-attention with encoder output
        # For cross-attention, padding_mask shape should be (batch_size, 1, target_seq_len, input_seq_len)
        attn2, attn_weights_2 = self.cross_mha(
            query=out1, value=enc_output, key=enc_output,
            attention_mask=padding_mask,
            return_attention_scores=True
        )
        attn2 = self.dropout2(attn2, training=training)
        out2 = self.layernorm2(out1 + attn2)

        # 3) Feed-forward
        ffn_output = self.ffn(out2)
        ffn_output = self.dropout3(ffn_output, training=training)
        out3 = self.layernorm3(out2 + ffn_output)

        return out3


def data_preprocess(df, diff_order, start_index, end_index,moving_average,MA_window_size):
    '''
    This function removes columns with more than 10% missing values and fills in the NaNs in the remaining columns.
    '''
    # start_date = pd.to_datetime(start_date)
    # end_date = pd.to_datetime(end_date)

    df = pd.DataFrame(df)
    df = df.loc[start_index:end_index]



    # Apply differencing according to the specified diff_order
    df_diff = df.copy()
    for _ in range(diff_order):


        if moving_average==True:
            # Assuming df_diff is your DataFrame


            # Function to pad the beginning of the DataFrame
            def pad_head(df, pad_width):
                # head = pd.DataFrame([df.iloc[0]] * pad_width)
                head = pd.DataFrame([df.iloc[0].values] * pad_width, columns=df.columns)
                return pd.concat([head, df]).reset_index(drop=True)

            # Apply padding only to the head
            pad_width = MA_window_size - 1
            df_padded = pad_head(df_diff, pad_width)

            # Compute the moving average on the padded DataFrame
            df_ma_padded = df_padded.rolling(window=MA_window_size, center=False).mean()

            # Now, remove the padding by trimming the extra rows from the start of the moving average DataFrame
            df_ma = df_ma_padded.iloc[pad_width:].reset_index(drop=True)

            # Shift the moving average DataFrame by one to align it for differencing
            df_ma_shifted = df_ma.shift(1)

            # Ensure df_diff is aligned in size with df_ma_shifted
            df_diff_aligned = df_diff.iloc[:df_ma_shifted.shape[0]]

            # Calculate the difference between the actual value at t+1 and the moving average at t
            df_diff = df_diff_aligned - df_ma_shifted
        else:
            df_diff = df_diff.diff(periods=1)


    # # Select the related range of data
    # df2 = df2.loc[start_index:end_index]
    # Adjust the index offset for the differencing order
    df_diff = df_diff.loc[start_index + diff_order:end_index]


    return df, df_diff


def preprocessing(data_, data_diff_, diff_order, sequence_length, horizon, stride, use_graph_layer,
                  save_instance,ignore_first_instance_stride, save_x_seq, save_y_seq, save_correlation_seq,
                  save_decoder_y_seq, save_last_known_values, target_name,
                  target_as_feature, target_comp_updated_list, differenced_target, differenced_X, moving_average,MA_window_size):
    '''

    This function takes in a time series data, normalizes it, generates sequences, and splits the data into
    training and testing sets.

    data: Time series data
    sequence_length: Length of the sequence used in the model
    horizon: The number of steps to predict in the future
    stride: The step size used to generate sequences
    test_size: The proportion of the data used for testing
    scaler: Scaler used to normalize the data
    '''
    if differenced_X == True:
        X = data_diff_
    else:
        # first row of X is dropped because of differencing
        X = data_[diff_order:]

    X_corr = data_diff_

    if differenced_target == True:
        Y = data_diff_[target_name].values
    else:
        Y = data_[diff_order:][target_name].values

    Y_original = data_[diff_order:][target_name].values


    if target_as_feature == False:
        # Drop columns from X based on target list
        X = X.drop(target_comp_updated_list, axis=1)
        X_corr = X_corr.drop(target_comp_updated_list, axis=1)
    X = X.values

    # Sequence Generation Part
    x_seq = []
    y_seq = []
    correlation_seq = []
    decoder_y_seq = []
    n = np.shape(X)[1]
    # Create ones matrix of size (n, n)
    corr = np.ones((n, n))

    # def calculate_corr(j, k, df_window):
    #     return j, k, chatterjee(df_window[:, j], df_window[:, k])

    if use_graph_layer == True:
        if cal_all_adj == False:
            # restart save_correlation
            save_correlation_seq = []

            # calculate starting correlation based on all available data
            # results = Parallel(n_jobs=-1)(delayed(calculate_corr)(j, k, X[:len(X) - horizon]) for j in range(n) for k in range(n))

            # # Populate the correlation matrix
            # for j, k, result in results:
            #     corr[j][k] = result

            # we put absolute value because GCN does not accept negative adjacency
            corr = np.corrcoef(X_corr[:len(X_corr) - horizon], rowvar=False)
            # Apply the filter to keep only correlations with absolute values greater than 0.05
            corr = np.where(np.abs(corr) > 0.05, corr, 0)

            # Normalize the corr to be in the range [0, 1]
            # normalized_corr = (1 + corr) / 2
            correlation_seq.append(corr)


    for instance in range(save_instance, len(X) - sequence_length - horizon +1, stride):
        if ignore_first_instance_stride:
            ignore_first_instance_stride = False
            continue
        x_seq.append(X[instance: instance + sequence_length])

        if Y_sequence == True:
            y_seq.append(Y_original[instance + sequence_length:instance + sequence_length + horizon])
        else:
            y_seq.append(np.expand_dims(np.sum(Y_original[instance + sequence_length:instance + sequence_length + horizon], axis=0), axis=0))

        decoder_y_seq.append(Y[instance + sequence_length - horizon:instance + sequence_length])
        df_window = X[instance: instance + sequence_length]

        if use_graph_layer == True:
            if cal_all_adj == True:
                # Parallelize the calculation of correlations
                results = Parallel(n_jobs=-1)(
                    delayed(calculate_corr)(j, k, df_window) for j in range(n) for k in range(n))

                # Populate the correlation matrix
                for j, k, result in results:
                    corr[j][k] = result

                correlation_seq.append(corr)

        # # Immediately after sequence generation, capture the last known value
        # if differenced_target == True:

        if moving_average==True:
            ## Apply moving average on last_value_before_sequence
            last_value_before_sequence = data_[target_name].iloc[max(0,
                                                                     instance + sequence_length - diff_order + 2 - MA_window_size):instance + sequence_length - diff_order + 2].mean()
        else:
            # Adjust the index to capture the last known value correctly
            last_value_before_sequence = data_[target_name].iloc[instance + sequence_length - diff_order + 1]

        last_value_before_sequence = np.expand_dims(last_value_before_sequence, axis=0)
        save_last_known_values.append(last_value_before_sequence)
        # else:
        #     save_last_known_values.append(1)

    del X
    save_instance = instance
    save_x_seq.extend(x_seq)
    save_y_seq.extend(y_seq)
    save_decoder_y_seq.extend(decoder_y_seq)
    save_correlation_seq.extend(correlation_seq)


    # last_known_values_input = [item[0] for item in last_known_values]  # Reshape as necessary
    return save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values


def plot_prediction_graph(original_y, forecast_list, test_boundary,
                          horizon, Test_KPI, iter, save_dict):
    plt.figure(figsize=(15, 10))
    # plt.axvline(x=test_boundary, ls=':', color='black')
    # plt.plot(test_index, actual_list[:len(forecast_list)], 'r--',marker='o')
    plt.plot(original_y.index[-horizon:], original_y[-horizon:], 'r--', marker='o')
    plt.plot(forecast_list[-1][1], forecast_list[-1][0], 'b', linewidth=0.6, marker='o')

    # Create a legend with custom labels
    actual_labels = ['actual_{}'.format(i) for i in range(len(target_name))]
    forecast_labels = ['forecast_{}'.format(i) for i in range(len(target_name))]
    legend_labels = actual_labels + forecast_labels
    plt.legend(legend_labels)

    # plt.legend(['test boundary', 'actual', 'prediction'])
    plt.xlabel('Date')
    plt.ylabel('Value')
    plt.text(0.02, 0.8, 'KPI for Test Range = {:.4f}'.format(Test_KPI), bbox=dict(facecolor='red', alpha=0.2),
             transform=plt.gca().transAxes, va="top", ha="left")
    plt.title('{}- Ahead Model_it={}'.format(horizon, iter + 1), loc='center')

    plt.savefig("{}\prediction_{}_it={}.png".format(save_dict, horizon, iter + 1), dpi=300)
    plt.close()


def plot_norm_diff_graph(Y_seq, y_pred, datetime_index, test_boundary,
                    horizon, Test_KPI, iter, save_dict):
    plt.figure(figsize=(15, 10))
    plt.axvline(x=test_boundary, ls=':', color='black')

    for i in range(-1, 0):
        plt.plot(datetime_index[i], Y_seq[i], 'r--', marker='o')
        plt.plot(datetime_index[i], y_pred[i], 'b', linewidth=0.6, marker='o')

    # Create a legend with custom labels
    actual_labels = ['actual_{}'.format(i) for i in range(len(target_name))]
    forecast_labels = ['forecast_{}'.format(i) for i in range(len(target_name))]
    legend_labels = ['test boundary'] + actual_labels + forecast_labels
    plt.legend(legend_labels)

    plt.xlabel('Date')
    plt.ylabel('Value')
    plt.text(0.02, 0.8, 'KPI for Test Range = {:.4f}'.format(Test_KPI), bbox=dict(facecolor='red', alpha=0.2),
             transform=plt.gca().transAxes, va="top", ha="left")
    plt.title('norm_diff_{}- Ahead Model_it={}'.format(horizon, iter + 1), loc='center')

    plt.savefig("{}\dnorm_diff_{}_it={}.png".format(save_dict, horizon, iter + 1), dpi=300)
    plt.close()



class ReverseDifferencingLayer(tf.keras.layers.Layer):
    def __init__(self, **kwargs):
        super(ReverseDifferencingLayer, self).__init__(**kwargs)

    def call(self, inputs):
        """
        Perform the reversal of differencing operation.

        Args:
        inputs: A list of two tensors, where
                inputs[0] is the tensor of predicted differenced values of shape (batch_size, sequence_length, len(target)),
                and inputs[1] is the tensor of the last known value of the series of shape (batch_size, 1, len(target)).

        Returns:
        A tensor of shape (batch_size, sequence_length, len(target)) containing the predictions in the original scale.
        """
        preds, last_known_value = inputs
        # Use tf.cumsum to reverse the differencing, adding the last known value as the base.
        # TensorFlow handles the broadcasting automatically.
        reversed_preds = tf.cumsum(preds, axis=1) + last_known_value
        return reversed_preds


def final_prediction(x_seq,Y_scaler, scaler2, decoder_Y_seq,last_known_values, correlation_seq, sequence_length, horizon, model, data, stride,
                     target_name, differenced_target, diff_order, save_dict):

    x_seq = np.expand_dims(x_seq[-1], axis=0)
    decoder_Y_seq = np.expand_dims(decoder_Y_seq[-1], axis=0)
    last_known_values = np.expand_dims(last_known_values[-1], axis=0)

    if len(correlation_seq) > 0:
        correlation_seq = np.expand_dims(correlation_seq[-1], axis=0)


    # Prediction
    y_pred = model.predict([x_seq, correlation_seq, decoder_Y_seq, last_known_values])

    if use_graph_layer:
        # Get attention coefficients
        gc_1_attn, gc_2_attn = model.attention_submodel.predict([x_seq, correlation_seq, decoder_Y_seq, last_known_values])
        # Visualize attention

        # Assuming gc_1_attn and gc_2_attn have shape (batch_size, N, attn_heads, N)
        # Step 1: Merge the heads by averaging
        gc_1_attn_merged = np.mean(gc_1_attn, axis=2)  # Shape: (batch_size, N, N)
        gc_2_attn_merged = np.mean(gc_2_attn, axis=2)  # Shape: (batch_size, N, N)

        # Step 2: Visualize attention coefficients for the first batch
        attn_coef_mean_1 = gc_1_attn_merged[0]  # Select the first graph in the batch
        plt.figure(figsize=(6, 4))
        sns.heatmap(attn_coef_mean_1, cmap="viridis")
        plt.title("Attention Coefficients - GAT Layer 1 (Merged Heads)")
        plt.xlabel("Node j")
        plt.ylabel("Node i")
        plt.savefig("gat_layer_1.png")
        plt.savefig(f"{save_dict}/gat_layer_1_it={iter + 1}.png", dpi=300)

        attn_coef_mean_2 = gc_2_attn_merged[0]  # Select the first graph in the batch
        plt.figure(figsize=(6, 4))
        sns.heatmap(attn_coef_mean_2, cmap="viridis")
        plt.title("Attention Coefficients - GAT Layer 2 (Merged Heads)")
        plt.xlabel("Node j")
        plt.ylabel("Node i")
        plt.savefig(f"{save_dict}/gat_layer_2_it={iter + 1}.png", dpi=300)





    y_pred_rev = y_pred



    if normalized_data==True:

        if reverse_normalization==True:
            y_pred_rev = np.expand_dims(Y_scaler.inverse_transform(np.squeeze(y_pred_rev, axis=0)),
                                     axis=0)  # reverse scaler to plot later in the original scale
        else:
            # It seems papers used normalized data for prediction results
            y_pred_rev = np.expand_dims(np.squeeze(y_pred_rev, axis=0),
                                        axis=0)
    else:
        # y_pred_rev = scaler2.transform(np.squeeze(y_pred_rev, axis=0))
        # y_pred_rev = np.expand_dims(y_pred_rev,
        #                         axis=0) # reverse scaler to plot later in the original scale

        y_pred_rev = np.expand_dims(np.squeeze(y_pred_rev, axis=0),
                                    axis=0)


    n_rows = data.shape[0] - sequence_length - horizon
    indices = np.arange(1 + sequence_length, 1 + n_rows + sequence_length, stride)


    if Y_sequence == True:

        if normalized_data == True:
            if reverse_normalization == True:
                actual_array = np.expand_dims(Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + np.arange(horizon)]), axis=0)
            else:
                # It seems papers used normalized data for prediction results
                actual_array = np.expand_dims(
                    data[target_name].values[indices[-1, None] + np.arange(horizon)], axis=0)
        else:

            actual_array = np.expand_dims(
                data[target_name].values[indices[-1, None] + np.arange(horizon)], axis=0)


        datetime_index = np.expand_dims(data.index.to_numpy()[indices[-1, None] + np.arange(horizon)], axis=0)
    else:
        if normalized_data == True:
            actual_array = Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + horizon - 1])
        else:
            actual_array = data[target_name].values[indices[-1, None] + horizon - 1]
        datetime_index = data.index.to_numpy()[indices[-1, None] + horizon - 1]


    return actual_array, y_pred_rev, datetime_index, y_pred



def data_set_generation(data, data_diff, max_lag, target_as_feature, target_name):
    # update the input_components list
    input_components = list(data.columns)

    # remove the target column before lagging the variables
    if not target_as_feature:
        for target in target_name:
            if target in input_components:
                input_components.remove(target)

    # List of updated input components.
    target_comp_updated_list = []

    # Initialize forecast dataset.
    new_data = pd.DataFrame()
    new_data_diff = pd.DataFrame()

    # Iterate through input components and their lags.
    for col in input_components:
        for lag in range(0, max_lag + 1):
            # Generate new component name.
            comp_new = f"{col} ({-lag})"
            if col in target_name:
                target_comp_updated_list.append(comp_new)
            # Get start index based on optimal time lag.
            new_data_time_series = data[col].shift(lag).rename(comp_new)
            new_data_diff_time_series = data_diff[col].shift(lag).rename(comp_new)

            # Add to dataset and forecast dataset
            new_data = pd.concat([new_data, new_data_time_series], axis=1)
            new_data_diff = pd.concat([new_data_diff, new_data_diff_time_series], axis=1)

    # Drop the first few rows to remove Nans.

    new_data = new_data.iloc[max_lag:]
    new_data_diff = new_data_diff.iloc[max_lag:]


    update_target_name = []
    if target_as_feature == True:
        for i, col in enumerate(target_name):
            update_target_name.append(f"{col} (0)")
        target_name = update_target_name
    else:
        # append the removed target column to datasets
        new_data[target_name] = data[target_name]
        new_data_diff[target_name] = data_diff[target_name]
        target_comp_updated_list.extend(target_name)

    # new_data = new_data.reset_index(drop=True)
    # new_data_diff = new_data_diff.reset_index(drop=True)

    return new_data, new_data_diff, target_comp_updated_list, target_name



def CORR(flattened_actual, flattened_forecast):
    """
    Calculate the Empirical Correlation Coefficient between two flattened arrays.

    Parameters:
    flattened_actual (array-like): Flattened array of actual values.
    flattened_forecast (array-like): Flattened array of forecasted values.

    Returns:
    float: Empirical Correlation Coefficient.
    """

    # Ensuring the inputs are numpy arrays
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    # Calculate the standard deviation and mean of the inputs
    sigma_p = flattened_forecast.std()
    sigma_g = flattened_actual.std()
    mean_p = flattened_forecast.mean()
    mean_g = flattened_actual.mean()

    # Calculate the correlation only if sigma_g and sigma_p are not equal to zero
    valid_indices = (sigma_g != 0) and (sigma_p != 0)
    if np.any(valid_indices):
        correlation = np.mean(
            ((flattened_forecast[valid_indices] - mean_p) * (flattened_actual[valid_indices] - mean_g)) / (sigma_p * sigma_g)
        )
    else:
        correlation = 0

    return correlation


def RSE(flattened_actual, flattened_forecast):
    """
    Calculate the Root Relative Squared Error between two flattened arrays or multi-dimensional arrays.

    Parameters:
    flattened_actual (array-like): Flattened array or multi-dimensional array of actual values.
    flattened_forecast (array-like): Flattened array or multi-dimensional array of forecasted values.

    Returns:
    float: Root Relative Squared Error.
    """

    # Ensuring the inputs are numpy arrays
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    # Calculate the mean of the actual values
    mean_actual = np.mean(flattened_actual)

    # Calculate the numerator (sum of squared differences between actual and forecast values)
    numerator = np.sum((flattened_actual - flattened_forecast) ** 2)

    # Calculate the denominator (sum of squared differences between actual values and mean of actual values)
    denominator = np.sum((flattened_actual - mean_actual) ** 2)

    # Calculate RSE
    if denominator != 0:
        rse = np.sqrt(numerator / denominator)
    else:
        rse = np.nan

    return rse

def mean_absolute_scaled_error(y_true, y_pred):
    naive_forecast = tf.roll(y_true, shift=1, axis=0)  # Shift the tensor by 1 position
    errors = tf.abs(y_true - y_pred)
    naive_errors = tf.abs(y_true - naive_forecast)
    epsilon = 1e-8  # Small constant to prevent division by zero
    mase = tf.reduce_mean(errors / (naive_errors + epsilon))
    return mase


def symmetric_mean_absolute_percentage_error(y_true, y_pred):
    y_true = K.maximum(K.epsilon(), y_true)  # Avoid division by zero
    return K.mean(2 * K.abs(y_pred - y_true) / (y_pred + y_true), axis=-1)


def symmetric_mean_absolute_percentage_error_loss(y_true, y_pred):
    epsilon = 1e-7  # A small constant to avoid division by zero
    y_true = K.maximum(epsilon, y_true)
    smape_loss = K.mean(2 * K.abs(y_pred - y_true) / (y_pred + y_true), axis=-1)
    return smape_loss








'''
##################################################
# Main codes
##################################################
'''

if __name__ == '__main__':

    # Define the base path and file pattern
    base_path = [
                  # 'C:/Users/sarab/Desktop/Valley Fever/Data/Aggregated Data/Maricopa/Maricopa - For Analysis.csv'
                "D:/Shared/desktop 3/Valley Fever/Weather Data MARICOPA (MMWR Weeks)/Clean Data(Maricopa)/Clean Processed Data2.csv"
                ]

    # Loop over the numbers you're interested in
    for model_index, file_path in enumerate(base_path):  # Adjust the range as needed

        # Extract the model name from the file path
        model_name = file_path.split('/')[-1].split('.')[0]

        # for ETD
        data = pd.read_csv(file_path)
        data.drop('Date', axis=1, inplace=True)  # axis=1 specifies that you are dropping a column


        # for horizon in [96, 192, 336, 720]:
        for horizon in [8]:
                model_index = model_name + '_'+ str(horizon)

                # Data preprocessing parameters

                model_weights_update_iter = int(round(0.2 * len(data)))+1 # I did not want to have dynamic retraining for simplicity
                sequence_length = 3*horizon
                stride = 1
                dynamic_test_size = math.ceil(horizon / stride)
                use_graph_layer = True
                cal_all_adj = False
                target_as_feature = True
                differenced_target = True
                differenced_X = True
                diff_order=1
                normalized_data = True
                reverse_normalization=True
                initial_training_normalization= False
                repeat_corr = True
                Y_sequence = True
                single_step = False
                moving_average=False
                MA_window_size=12
                batch_size = 10
                feature_gate_k_percent = 0.1  # Percentage of top features to keep (0.5 = 50%)
                max_lag = 6


                # Set Target Name
                # Replace with the actual name of your target column
                target_name_original = ['MARICOPA']  # Example: change to your actual column name
                target_name = target_name_original

                new_row = []
                # Initialize a list to store new rows
                buffered_rows = []
                save_to_excel_iter = 10


                head_size = 32
                num_heads = 16
                ff_dim = 256
                mlp_units = [512]
                dropout = 0.05
                mlp_dropout = 0.05

                '''#################### create an output file in the current directory ########################'''
                # Generate a string representing the current date and time
                now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                directory = "results_{}".format(now)
                if not os.path.exists(directory):
                    os.makedirs(directory)

                # Create the output file name with the current date and time
                output_filename = directory + '/dynamic_results_' + now +str(model_index) + '.xlsx'

                if os.path.exists(output_filename) == False:
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    # make Header
                    new_row = ['iteration',
                               'prediction_date',
                               'actual_value',
                               'forecast_value',
                               'Test_MAPE',
                               'Test_MAE',
                               'Test_MSE',
                               'Test_CORR',
                               'Test_RSE']

                    sheet.append(new_row)
                    wb.save(output_filename)



                start_index = 0
                train_end = 850
                # val_end = 12 * 30 * 24 + 4 * 30 * 24
                # test_end = 12 * 30 * 24 + 4 * 30 * 24 + 4 * 30 * 24
                test_start_index = 900 # val_end
                test_end_index = 987 # test_end



                # test_range = pd.date_range(start=test_start_date, end=test_end_date, freq='H')
                test_range = range(test_start_index, test_end_index)
                initial_train_data = data.loc[start_index:train_end]

                df_norm=data.copy()
                Y_scaler=[]

                if initial_training_normalization ==True and  normalized_data == True:

                    # apply normalization on initial data
                    scaler = StandardScaler()  # scaler applied to the predictors
                    # normalize data based on initial training set
                    scaler.fit(initial_train_data)
                    # keep the scaler attributes to reverse the transform later
                    scaler2 = StandardScaler()
                    Y_scaler = scaler2.fit(initial_train_data[target_name])

                    df_norm = scaler.transform(data)

                    # Wrap back in a DataFrame
                    df_norm = pd.DataFrame(
                        df_norm,
                        columns=data.columns,
                        index=data.index
                    )











                if normalized_data == False:
                    scaler2 = StandardScaler()
                    Y_scaler = scaler2.fit(initial_train_data[target_name])



                forecast_list = []

                actual_list = []

                iter = 0
                save_instance = 0
                ignore_first_instance_stride = False
                ignore_weights = True
                save_x_seq = []
                save_y_seq = []
                save_correlation_seq = []
                save_decoder_y_seq = []
                save_last_known_values = []
                run_first_time = True
                iter = horizon - 1


                if ignore_weights==False:
                    iter = 2619


                for end_index in test_range[iter:]:

                    print("##################### run # {} / end index: {} ########################### ".format(iter, end_index))
                    SEED = 33
                    tf.random.set_seed(SEED)
                    os.environ['PYTHONHASHSEED'] = str(SEED)
                    np.random.seed(SEED)
                    random.seed(SEED)

                    session_conf = tf.compat.v1.ConfigProto(
                        intra_op_parallelism_threads=1,
                        inter_op_parallelism_threads=1
                    )
                    sess = tf.compat.v1.Session(
                        graph=tf.compat.v1.get_default_graph(),
                        config=session_conf
                    )
                    tf.compat.v1.keras.backend.set_session(sess)




                    # Data preprocessing

                    if initial_training_normalization==False and  normalized_data == True:

                        df = data.loc[start_index:end_index]
                        # apply normalization on initial data
                        scaler = StandardScaler()
                        # normalize data based on dynamic training set
                        # scaler.fit(df[:-horizon])
                        # df_norm = scaler.transform(df)  # keep the normalized data to predict later
                        scaler.fit(df[:-horizon])
                        df_norm_array = scaler.transform(df)

                        # Wrap back in a DataFrame
                        df_norm = pd.DataFrame(
                            df_norm_array,
                            columns=df.columns,
                            index=df.index
                        )

                        # keep the scaler attributes to reverse the transform later
                        scaler2 = StandardScaler()
                        Y_scaler = scaler2.fit(df[target_name_original][:-horizon])




                    data_norm, data_diff= data_preprocess(df_norm,diff_order, start_index, end_index,moving_average,MA_window_size)

                    # generate lagged variables
                    data_norm, data_diff, target_comp_updated_list, target_name = data_set_generation(data_norm,
                                                                                                 data_diff,
                                                                                                 max_lag=max_lag,
                                                                                                 target_as_feature=target_as_feature,
                                                                                                 target_name=target_name_original)

                    save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values = preprocessing(
                        data_=data_norm,
                        data_diff_=data_diff,
                        diff_order=diff_order,
                        sequence_length=sequence_length,
                        horizon=horizon, stride=stride,
                        use_graph_layer=use_graph_layer,
                        save_instance=save_instance,
                        ignore_first_instance_stride=ignore_first_instance_stride,
                        save_x_seq=save_x_seq,
                        save_y_seq=save_y_seq,
                        save_correlation_seq=save_correlation_seq,
                        save_decoder_y_seq=save_decoder_y_seq,
                        save_last_known_values=save_last_known_values,
                        target_name=target_name,
                        target_as_feature=target_as_feature,
                        target_comp_updated_list=target_comp_updated_list,
                        differenced_target=differenced_target,
                        differenced_X=differenced_X,
                        moving_average=moving_average,
                        MA_window_size=MA_window_size

                    )




                    del data_diff
                    # Test data split
                    if repeat_corr == True:
                        X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test,last_known_values_train,last_known_values_test = \
                            train_test_split(np.asarray(save_x_seq),
                                             np.asarray(save_y_seq),
                                             np.asarray(save_decoder_y_seq),
                                             np.asarray(save_last_known_values),
                                             test_size=dynamic_test_size,
                                             shuffle=False,
                                             random_state=1004)  # random_state ignored because shuffle = False

                        correlation_train = np.array(save_correlation_seq)
                        correlation_test = correlation_train


                    else:
                        X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test, last_known_values_train,last_known_values_test, correlation_train, correlation_test = \
                            train_test_split(np.asarray(save_x_seq),
                                             np.asarray(save_y_seq),
                                             np.asarray(save_decoder_y_seq),
                                             np.asarray(save_last_known_values),
                                             np.asarray(save_correlation_seq),
                                             test_size=dynamic_test_size,
                                             shuffle=False,
                                             random_state=1004)  # random_state ignored because shuffle = False


                    ignore_first_instance_stride = True


                    def data_generator(X, correlation, decoder_Y, Y, last_known_values, batch_size, new_data_ratio=0.1):
                        while True:
                            # Number of new data samples to include in each batch
                            new_data_count = int(batch_size * new_data_ratio)

                            # Indices of the new data
                            new_data_indices = np.arange(len(X) - new_data_count, len(X))

                            for i in range(0, len(X), batch_size):
                                # Select random indices from the entire dataset
                                random_indices = np.random.permutation(len(X) - new_data_count)

                                # Combine new data indices with random indices and shuffle
                                combined_indices = np.concatenate([new_data_indices, random_indices[:batch_size - new_data_count]])
                                np.random.shuffle(combined_indices)

                                batch_X = X[combined_indices]
                                batch_decoder_Y = decoder_Y[combined_indices]
                                batch_last_known_values = last_known_values[combined_indices]
                                batch_Y = Y[combined_indices]

                                if  len(correlation) > 0:
                                    if repeat_corr == True:
                                        batch_correlation = np.repeat(correlation[np.newaxis, ...], len(batch_X), axis=0)
                                    else:
                                        batch_correlation = correlation[combined_indices]
                                else:
                                    batch_correlation=[]

                                yield [batch_X, batch_correlation, batch_decoder_Y, batch_last_known_values], batch_Y

                    class FeatureGate(layers.Layer):
                        def __init__(self, num_features, k_percent=0.5, **kwargs):
                            super().__init__(**kwargs)
                            self.num_features = num_features
                            self.k_percent = k_percent  # Percentage of top features to keep

                        def build(self, input_shape):
                            # Create a trainable parameter (vector) for gating each feature
                            # shape=(num_features,)
                            # Use a randomly initialized weight vector instead of zeros
                            # This will create more varied gate values
                            self.logits = self.add_weight(
                                shape=(self.num_features,),
                                initializer="random_normal",  # Use random normal initialization
                                trainable=True,
                                name="feature_gate_logits"
                            )

                        def call(self, inputs):
                            """
                            inputs shape: (batch_size, num_features, time)
                            Returns:
                                Tensor of same shape, with top-k% features preserved and others zeroed out.
                            """
                            # Convert logits -> gates in [0,1] (via sigmoid)
                            gates = tf.nn.sigmoid(self.logits)  # shape = (num_features,)
                            
                            # Determine k value (number of features to keep)
                            k = tf.cast(tf.math.ceil(self.k_percent * tf.cast(self.num_features, tf.float32)), tf.int32)
                            
                            # Get the values and indices of the top-k gates
                            _, top_k_indices = tf.nn.top_k(gates, k=k)
                            
                            # Create a mask with zeros everywhere except at top-k indices
                            mask = tf.zeros_like(gates)
                            mask = tf.tensor_scatter_nd_update(
                                mask,
                                tf.expand_dims(top_k_indices, axis=1),
                                tf.ones((k,), dtype=tf.float32)
                            )
                            
                            # Apply the mask to gates
                            masked_gates = gates * mask
                            
                            # Reshape to broadcast across batch & time
                            # gates shape => (1, num_features, 1)
                            gates_reshaped = tf.reshape(masked_gates, (1, -1, 1))
                            
                            # Multiply each feature by its gate
                            return inputs * gates_reshaped
                            
                        def get_feature_importance(self):
                            """
                            Returns the learned feature importance gates in [0,1].
                            Higher values indicate more important features.
                            
                            Returns:
                                numpy array of shape (num_features,) representing feature importance
                            """
                            return tf.nn.sigmoid(self.logits).numpy()
                            
                        def get_selected_features(self):
                            """
                            Returns the indices of the selected top-k features.
                            
                            Returns:
                                numpy array of indices of selected features
                            """
                            gates = tf.nn.sigmoid(self.logits)
                            k = tf.cast(tf.math.ceil(self.k_percent * tf.cast(self.num_features, tf.float32)), tf.int32)
                            _, top_k_indices = tf.nn.top_k(gates, k=k)
                            return top_k_indices.numpy()

                    def graph_processing_block(inputs, inp_lap, head_size, num_heads, dropout, horizon):
                        """
                        Process inputs through graph convolutional layers.
                        
                        The goal is to maintain the temporal dimension intact while enhancing feature 
                        representations using graph relationships.
                        
                        Note: The inputs shape is (batch_size, sequence_length, n_features) and 
                        we transpose to (batch_size, n_features, sequence_length) for graph processing.
                        After graph processing, we transpose back to maintain the time dimension.
                        """
                        l2_reg = 2.5e-4  # L2 regularization rate
                        
                        # Save original sequence length for reshaping later
                        batch_size, original_seq_len, n_features = tf.shape(inputs)[0], tf.shape(inputs)[1], tf.shape(inputs)[2]
                        
                        # Transpose from (batch, time, features) to (batch, features, time)
                        x = tf.transpose(inputs, perm=[0, 2, 1])
                        
                        # Replace MinCutPool with FeatureGate - use n_features as num_features
                        x = FeatureGate(num_features=int(x.shape[1]), k_percent=feature_gate_k_percent, name='feature_gate')(x)
                        
                        do_1 = Dropout(dropout)(x)
                        
                        # GAT operates on the node/feature dimension
                        gc_1, gc_1_attn = GATv2Conv(
                            int(math.ceil(x.shape[2])),
                            attn_heads=num_heads,
                            concat_heads=False,
                            dropout_rate=dropout,
                            activation="relu",
                            kernel_regularizer=l2(l2_reg),
                            attn_kernel_regularizer=l2(l2_reg),
                            bias_regularizer=l2(l2_reg),
                            return_attn_coef=True,
                        )([do_1, inp_lap])
                        
                        # Transpose back to (batch, time, features)
                        # After graph processing, we need to ensure the temporal dimension is preserved
                        graph_output = tf.transpose(gc_1, perm=[0, 2, 1])
                        
                        # For compatibility with the attention visualizations
                        gc_2_attn = gc_1_attn
                        
                        return graph_output, gc_1_attn, gc_2_attn

                    def analyze_feature_importance(model, max_lag=max_lag):
                        """
                        Extracts and analyzes feature importance from the FeatureGate layer.
                        
                        Args:
                            model: The trained Keras model containing a FeatureGate layer
                            
                        Returns:
                            numpy array of feature importance values
                        """
                        # Find the FeatureGate layer in the model
                        feature_gate_layer = None
                        for layer in model.layers:
                            if isinstance(layer, FeatureGate):
                                feature_gate_layer = layer
                                break
                                
                        # If nested, search through sub-models
                        if feature_gate_layer is None:
                            for layer in model.layers:
                                if hasattr(layer, 'layers'):
                                    for sublayer in layer.layers:
                                        if isinstance(sublayer, FeatureGate):
                                            feature_gate_layer = sublayer
                                            break
                        
                        if feature_gate_layer is None:
                            print("Could not find FeatureGate layer in the model!")
                            return None
                            
                        # Get feature importance values
                        importance = feature_gate_layer.get_feature_importance()
                        
                        # Get the indices of selected features
                        selected_features = feature_gate_layer.get_selected_features()
                        k_percent = feature_gate_layer.k_percent
                        
                        # Get feature names from X_train's transposed shape
                        # For GAT, we transpose from (batch, time, features) to (batch, features, time)
                        # So the feature dimension corresponds to the columns in the original data
                        
                        # Get column names from the original data (before lagging)
                        original_columns = list(data.columns)
                        
                        # In preprocessing, we generate lagged features with names like "feature (-lag)"
                        # We can recreate these feature names for the visualization
                        lagged_feature_names = []
                        for col in original_columns:
                            for lag in range(0, max_lag+1):  # max_lag=6 as set in data_set_generation
                                lagged_feature_names.append(f"{col} ({-lag})")
                                
                        # Print the top features
                        feature_indices = np.argsort(importance)[::-1]  # Descending order
                        
                        print("Feature Importance Analysis:")
                        print("-" * 60)
                        print(f"{'Feature Index':<10} {'Feature Name':<30} {'Importance':<10} {'Selected':<10}")
                        print("-" * 60)
                        
                        for idx in feature_indices:
                            is_selected = "Yes" if idx in selected_features else "No"
                            feature_name = lagged_feature_names[idx] if idx < len(lagged_feature_names) else f"Feature {idx}"
                            print(f"{idx:<10} {feature_name:<30} {importance[idx]:.6f} {is_selected:<10}")
                            
                        print(f"\nKeeping top {k_percent*100:.1f}% of features ({len(selected_features)}/{len(importance)})")
                        
                        # Create more informative visualizations - show both all features and selected features
                        
                        # 1. Full visualization with selected/unselected features colored differently
                        plt.figure(figsize=(14, max(8, len(feature_indices) * 0.25)))  # Adjust height based on number of features
                        
                        # Create color array - green for selected, gray for unselected
                        colors = ['lightgray'] * len(feature_indices)
                        for i, idx in enumerate(feature_indices):
                            if idx in selected_features:
                                colors[i] = 'lightgreen'
                        
                        # Create horizontal bar chart with sorted importance
                        bars = plt.barh(range(len(feature_indices)), importance[feature_indices], color=colors)
                        
                        # Add feature names as y-tick labels
                        feature_labels = []
                        for idx in feature_indices:
                            if idx < len(lagged_feature_names):
                                feature_labels.append(lagged_feature_names[idx])
                            else:
                                feature_labels.append(f"Feature {idx}")
                        
                        plt.yticks(range(len(feature_indices)), feature_labels)
                        
                        # Add value labels to the end of each bar
                        for i, v in enumerate(importance[feature_indices]):
                            plt.text(v + 0.01, i, f"{v:.4f}", va='center')
                        
                        # Create legend
                        from matplotlib.patches import Patch
                        legend_elements = [
                            Patch(facecolor='lightgreen', label='Selected'),
                            Patch(facecolor='lightgray', label='Filtered out')
                        ]
                        plt.legend(handles=legend_elements, loc='upper right')
                        
                        # Set x-axis limits based on data range with a small buffer
                        min_val = max(0, np.min(importance) - 0.05)  # Add buffer, but don't go below 0
                        max_val = np.max(importance) + 0.05  # Add buffer
                        plt.xlim(min_val, max_val)
                        
                        plt.xlabel('Importance Score')
                        plt.ylabel('Features')
                        plt.title(f'Feature Importance with Top {k_percent*100:.1f}% Selection')
                        plt.tight_layout()
                        
                        # Save the figure
                        plt.savefig(f'{directory}/feature_importance_{model_index}.png', dpi=300)
                        plt.close()
                        
                        # 2. Visualization showing only selected features for clarity
                        plt.figure(figsize=(12, max(6, len(selected_features) * 0.3)))  # Adjust height based on selected features
                        
                        # Get indices sorted by importance
                        selected_indices_sorted = [idx for idx in feature_indices if idx in selected_features]
                        
                        # Sort in ascending order for horizontal bar chart (bottom to top)
                        selected_indices_sorted = selected_indices_sorted[::-1]
                        
                        plt.barh(range(len(selected_indices_sorted)), 
                                 importance[selected_indices_sorted], 
                                 color='lightgreen')
                        
                        # Create labels for selected features
                        selected_labels = []
                        for idx in selected_indices_sorted:
                            if idx < len(lagged_feature_names):
                                selected_labels.append(lagged_feature_names[idx])
                            else:
                                selected_labels.append(f"Feature {idx}")
                        
                        plt.yticks(range(len(selected_indices_sorted)), selected_labels)
                        
                        # Add value labels
                        for i, idx in enumerate(selected_indices_sorted):
                            plt.text(importance[idx] + 0.01, i, f"{importance[idx]:.4f}", va='center')
                        
                        # Set x-axis limits for selected features plot
                        min_val = max(0, np.min(importance[selected_indices_sorted]) - 0.05)
                        max_val = np.max(importance[selected_indices_sorted]) + 0.05
                        plt.xlim(min_val, max_val)
                        
                        plt.xlabel('Importance Score')
                        plt.ylabel('Selected Features')
                        plt.title(f'Top {k_percent*100:.1f}% Selected Features')
                        plt.tight_layout()
                        
                        # Save the selected features figure
                        plt.savefig(f'{directory}/selected_features_{model_index}.png', dpi=300)
                        plt.close()
                        
                        return importance
                    

                    def transformer_encoder_decoder_block(graph_output, decoder_inputs, d_model, num_heads, ff_dim, horizon, target_name, dropout):
                        """
                        Transformer-based encoder-decoder block for time series forecasting.
                        
                        The encoder processes the input time series enhanced by GNN,
                        and the decoder generates the forecast in an auto-regressive manner.
                        
                        Args:
                            graph_output: Output from graph_processing_block, shape (batch_size, seq_length, features)
                            decoder_inputs: Initial decoder input, shape (batch_size, horizon, features)
                            d_model: Model dimension
                            num_heads: Number of attention heads
                            ff_dim: Feed-forward network dimension
                            horizon: Forecast horizon
                            target_name: Target variable name
                            dropout: Dropout rate
                        
                        Returns:
                            The forecast output
                        """
                        l2_reg = 2.5e-4  # L2 regularization rate
                        
                        # Get the actual sequence length from the input for positional encoding
                        sequence_length = tf.shape(graph_output)[1]
                        max_encoder_length = 200  # Maximum possible sequence length, should be larger than any expected sequence
                        
                        # Add positional embeddings to encoder input
                        encoder_embedding = PositionalEmbedding(max_sequence_length=max_encoder_length, d_model=d_model)
                        enc_emb = encoder_embedding(graph_output)
                        
                        # Encoder stacks (2 layers)
                        encoder_output = enc_emb
                        for i in range(2):  # Number of encoder layers
                            encoder_block = TransformerEncoder(d_model, num_heads, ff_dim, dropout=dropout)
                            encoder_output = encoder_block(encoder_output)
                        
                        # Create look-ahead mask for decoder to ensure causal attention
                        # In auto-regressive forecasting, each position can only attend to previous positions
                        look_ahead_mask = create_look_ahead_mask(horizon)  # Now returns shape (1, 1, horizon, horizon)
                        
                        # Add positional embeddings to decoder input
                        decoder_embedding = PositionalEmbedding(max_sequence_length=horizon, d_model=d_model)
                        dec_emb = decoder_embedding(decoder_inputs)
                        
                        # Decoder stacks (2 layers)
                        decoder_output = dec_emb
                        for i in range(2):  # Number of decoder layers
                            decoder_block = TransformerDecoder(d_model, num_heads, ff_dim, dropout=dropout)
                            # Pass the look_ahead_mask to ensure causality in the decoder
                            decoder_output = decoder_block(
                                decoder_output, 
                                encoder_output,
                                look_ahead_mask=look_ahead_mask
                            )
                        
                        # Final output layer
                        if Y_sequence == True:
                            # Output sequence with TimeDistributed dense layer
                            outputs = keras.layers.TimeDistributed(keras.layers.Dense(len(target_name)))(decoder_output)
                        else:
                            # Single output
                            outputs = keras.layers.Dense(len(target_name))(decoder_output[:, -1:, :])
                        
                        return outputs



                    #with strategy.scope():
                    def build_model(input_shape, correlation_shape, use_graph_layer):

                        inputs = keras.Input(shape=input_shape)
                        if use_graph_layer == True:
                            inp_lap = keras.Input(shape=correlation_shape)
                        else:
                            inp_lap = []

                        if use_graph_layer:
                            graph_output, gc_1_attn, gc_2_attn = graph_processing_block(inputs, inp_lap, head_size,
                                                                                        num_heads, dropout,
                                                                                        horizon)
                        else:
                            graph_output = inputs

                        # Project to consistent dimension for transformer if needed
                        d_model = 512  # Transformer embedding dimension
                        graph_output = layers.Dense(d_model)(graph_output)
                        
                        # Create decoder inputs
                        decoder_inputs = tf.keras.Input(
                            shape=(horizon, len(target_name)), name='decoder_inputs')

                        # Use transformer encoder-decoder instead of LSTM
                        outputs = transformer_encoder_decoder_block(
                            graph_output=graph_output,
                            decoder_inputs=decoder_inputs,
                            d_model=d_model,
                            num_heads=num_heads,
                            ff_dim=ff_dim,
                            horizon=horizon,
                            target_name=target_name,
                            dropout=dropout
                        )

                        # Assume 'last_known_value_input' is an additional input layer for the last known value before prediction starts
                        last_known_value_input = tf.keras.Input(shape=(1, len(target_name)), name='last_known_values')
                        #
                        if differenced_target == True:
                            # Integrate the ReverseDifferencingLayer
                            outputs = ReverseDifferencingLayer()([outputs, last_known_value_input])

                            # Return a model with outputs for training and a submodel for attention
                        model = keras.Model(
                            [inputs, inp_lap, decoder_inputs, last_known_value_input],
                            outputs,  # Use only `outputs` for training
                        )
                        if use_graph_layer:
                            # Store attention as a submodel
                            model.attention_submodel = keras.Model(
                                [inputs, inp_lap, decoder_inputs, last_known_value_input],
                                [gc_1_attn, gc_2_attn],  # Include attention coefficients here
                            )

                        return model





                    if iter % model_weights_update_iter == 0 or run_first_time==True:
                        run_first_time=False
                        input_shape = X_train.shape[1:]
                        correlation_shape = correlation_train.shape[1:]

                        model = build_model(
                            input_shape,
                            correlation_shape,
                            use_graph_layer=use_graph_layer
                        )


                        model.compile(
                            loss="mse",
                            optimizer=tf.keras.optimizers.Adam(learning_rate=1e-4),
                            metrics=[tf.keras.metrics.MeanSquaredError()],
                        )


                        model.summary()

                        callbacks = [
                            keras.callbacks.EarlyStopping(patience=20, min_delta=0.001, monitor='val_mean_squared_error',
                                                          mode='auto',
                                                          restore_best_weights=True)]



                        # Define the size of the validation set
                        validation_ratio = 0.25  # 20% or 25% of the data for validation
                        total_data_size = len(X_train)
                        validation_size = int(total_data_size * validation_ratio)

                        # Generate random indices for the validation set
                        # np.random.seed(42)
                        # validation_indices = np.random.choice(total_data_size, validation_size, replace=False)

                        # Select the last validation_size indices for the validation set
                        validation_indices = np.arange(total_data_size - validation_size, total_data_size)


                        # Create masks for selecting training and validation data
                        validation_mask = np.zeros(total_data_size, dtype=bool)
                        validation_mask[validation_indices] = True
                        training_mask = ~validation_mask

                        # Split the data into training and validation sets using the masks
                        X_train_train = X_train[training_mask]
                        if correlation_train.ndim > 1:
                            correlation_train_train = np.squeeze(correlation_train,axis=0)
                        else:
                            correlation_train_train=[]
                        decoder_Y_train_train = decoder_Y_train[training_mask]
                        last_known_values_train_train= last_known_values_train[training_mask]
                        Y_train_train = Y_train[training_mask]

                        X_train_valid = X_train[validation_mask]
                        if correlation_train.ndim > 1:
                            correlation_train_valid = np.squeeze(correlation_train,axis=0)
                        else:
                            correlation_train_valid=[]
                        decoder_Y_train_valid = decoder_Y_train[validation_mask]
                        last_known_values_train_valid = last_known_values_train[validation_mask]
                        Y_train_valid = Y_train[validation_mask]



                        train_gen = data_generator(X_train_train, correlation_train_train, decoder_Y_train_train, Y_train_train, last_known_values_train_train, batch_size=batch_size, new_data_ratio=0)

                        # Create the validation generator
                        val_gen = data_generator(X_train_valid, correlation_train_train, decoder_Y_train_valid, Y_train_valid, last_known_values_train_valid,
                                                   batch_size=batch_size, new_data_ratio=0)

                        # Calculate the steps per epoch for training and validation
                        train_steps = len(X_train_train) // batch_size
                        val_steps = len(X_train_valid) // batch_size


                        if ignore_weights == True:
                            history = model.fit(
                                train_gen,
                                steps_per_epoch=train_steps,
                                validation_data=val_gen,
                                validation_steps=val_steps,
                                epochs=100,
                                callbacks=callbacks
                            )
                            ignore_weights = False


                        else:
                            model.load_weights('model_weights_{}.h5'.format(model_index))
                            # history = model.fit(
                            #     train_gen,
                            #     steps_per_epoch=train_steps,
                            #     validation_data=val_gen,
                            #     validation_steps=val_steps,
                            #     epochs=20,
                            #     callbacks=callbacks
                            # )

                        # Saving weights as per your existing code
                        model.save_weights('model_weights_{}.h5'.format(model_index))
                        
                        # Analyze feature importance if the model has a FeatureGate layer
                        if use_graph_layer:
                            feature_importance = analyze_feature_importance(model, max_lag=max_lag)
                            # Save feature importance for future analysis if needed
                            if feature_importance is not None:
                                np.save(f'feature_importance_{model_index}.npy', feature_importance)


                    # Predict from last sequence


                    actual, y_pred_rev, datetime_index, y_pred = final_prediction(x_seq=save_x_seq,
                                                                                  Y_scaler=Y_scaler,
                                                                                  scaler2= scaler2,
                                                                                  decoder_Y_seq=save_decoder_y_seq,
                                                                                  last_known_values=save_last_known_values,
                                                                                  correlation_seq=save_correlation_seq,
                                                                                  sequence_length=sequence_length,
                                                                                  horizon=horizon,
                                                                                  model= model,
                                                                                  data=data_norm,
                                                                                  stride=stride,
                                                                                  target_name=target_name,
                                                                                  differenced_target=differenced_target,
                                                                                  diff_order=diff_order,
                                                                                  save_dict=directory)


                    if single_step==True:

                        # Save prediction in each loop
                        forecast_list = [(np.expand_dims(y_pred_rev[-1, :][-1],axis=0), datetime_index[-1][-1])]
                        actual_list = [(actual[-1, :][-1], datetime_index[-1][-1])]


                    else:

                        #Multi-step Forecast

                        # Save prediction in each loop
                        forecast_list = [(y_pred_rev[-1, :], datetime_index[-1])]
                        actual_list = [(actual[-1, :], datetime_index[-1])]

                        # forecast_list.append((y_pred_rev[-1, :], datetime_index[-1]))
                        # actual_list.append((actual[-1, :], datetime_index[-1]))

                    # Flatten the 3D arrays to 1D arrays
                    flattened_actual = np.concatenate([t[0] for t in actual_list]).flatten()
                    flattened_forecast = np.concatenate([t[0] for t in forecast_list]).flatten()

                    Test_CORR = CORR(flattened_actual, flattened_forecast)
                    Test_RSE = RSE(flattened_actual, flattened_forecast)
                    Test_MAPE = mean_absolute_percentage_error(flattened_actual, flattened_forecast)
                    Test_MAE = mean_absolute_error(flattened_actual, flattened_forecast)
                    Test_MSE = mean_squared_error(flattened_actual, flattened_forecast)

                    if reverse_normalization==True and normalized_data==True:
                      plot_prediction_graph(original_y=data.loc[end_index-horizon:end_index][target_name_original], forecast_list=forecast_list,
                                            test_boundary=test_start_index, horizon=horizon,
                                            Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    elif normalized_data== False:
                      transformed_data = scaler2.transform(data.loc[end_index-horizon:end_index][target_name_original])
                      # Convert the transformed data back into a pandas DataFrame or Series with the original index
                      original_y = pd.Series(transformed_data.squeeze(), index=data.loc[end_index-horizon:end_index].index)

                      plot_prediction_graph(original_y=original_y, forecast_list=forecast_list,
                                            test_boundary=test_start_index, horizon=horizon,
                                            Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    else:
                      # For ETD datasets we used (data_norm[target_name] instead of data[target_name_original]) It seems papers used normalized data for prediction results
                      plot_prediction_graph(original_y=data_norm[target_name], forecast_list=forecast_list,
                                          test_boundary=test_start_index,horizon=horizon,
                                          Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    '''##################### Exporting and Saving  ########################'''


                    if single_step==True and Y_sequence==True:
                        #Single-step Forecast
                        new_row = [iter,
                                   test_range[iter],
                                   str(actual[-1][-1]),
                                   str(y_pred_rev[-1][-1]),
                                   Test_MAPE,
                                   Test_MAE,
                                   Test_MSE,
                                   Test_CORR,
                                   Test_RSE]
                    else:
                        new_row = [iter,
                                   test_range[iter],
                                   str(actual[-1]),
                                   str(y_pred_rev[-1]),
                                   Test_MAPE,
                                   Test_MAE,
                                   Test_MSE,
                                   Test_CORR,
                                   Test_RSE]

                    buffered_rows.append(new_row)


                    if iter % save_to_excel_iter == 0:
                        wb = openpyxl.load_workbook(filename=output_filename)
                        sheet = wb.active
                        # Append the buffered rows to the sheet
                        for row in buffered_rows:
                            sheet.append(row)

                        # Save the Excel file
                        wb.save(output_filename)

                        # Clear the buffer
                        buffered_rows = []

                    iter = iter + 1

                    # Clear the session and invoke garbage collection
                    clear_session()
                    gc.collect()

                # After the loop, append any remaining buffered rows
                if buffered_rows:
                    wb = openpyxl.load_workbook(filename=output_filename)
                    sheet = wb.active
                    for row in buffered_rows:
                        sheet.append(row)
                    wb.save(output_filename)

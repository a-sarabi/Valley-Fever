import seaborn as sns
import os
GPU_ID="0"
os.environ["CUDA_VISIBLE_DEVICES"] = GPU_ID
import gc
from tensorflow.keras.backend import clear_session

print(os.environ['PATH'])
#
# from dlib import cuda
# #cuda.set_device(1)
from tensorflow.python.client import device_lib
print(device_lib.list_local_devices())
import os
import openpyxl
import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.model_selection import train_test_split
import keras.backend as K
import tensorflow as tf
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_percentage_error, mean_squared_error, mean_absolute_error
from spektral.layers import GCNConv, SAGPool, MinCutPool, GlobalAttentionPool, JustBalancePool, DiffPool
from spektral.utils import normalized_adjacency
from sklearn.metrics.pairwise import cosine_similarity
from gatv2_conv import GATv2Conv  # Import the GATv2Conv layer


import random
import datetime
import math
#from tensorflow.keras.layers import LSTM, Attention, GlobalAveragePooling1D, Bidirectional
from joblib import Parallel, delayed
from keras.layers import Dense, Dropout, Embedding, LSTM, Bidirectional
from keras.layers import BatchNormalization
from tensorflow.keras.layers import Dropout, Input
from spektral.layers import GATConv
from keras.layers import Add, Concatenate
from tensorflow.keras.layers import Lambda, MultiHeadAttention, LayerNormalization
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras.regularizers import l2

# Add Positional Embedding Layer
class PositionalEmbedding(layers.Layer):
    """
    Adds positional embeddings to the input embeddings.
    """
    def __init__(self, max_sequence_length, d_model, **kwargs):
        super().__init__(**kwargs)
        self.token_embeddings = layers.Dense(d_model)  # Linear projection to d_model dimension
        self.pos_embeddings = layers.Embedding(input_dim=max_sequence_length, output_dim=d_model)
        self.max_sequence_length = max_sequence_length
        self.d_model = d_model

    def call(self, x):
        """
        x shape: (batch_size, sequence_length, input_dim)
        """
        seq_len = tf.shape(x)[1]
        # Ensure positions don't exceed max_sequence_length
        positions = tf.range(start=0, limit=seq_len, delta=1)
        positions = tf.minimum(positions, self.max_sequence_length - 1)
        
        embedded_tokens = self.token_embeddings(x)  # (batch_size, seq_len, d_model)
        embedded_positions = self.pos_embeddings(positions)  # (seq_len, d_model)
        
        # Add positional embeddings to token embeddings
        return embedded_tokens + embedded_positions

# Create a look-ahead mask for decoder self-attention
def create_look_ahead_mask(size):
    """
    Creates a causal/look-ahead mask to prevent decoder from attending to future positions.
    
    Args:
        size: The sequence length for both query and key dimensions
    
    Returns:
        A mask of shape (1, 1, size, size) for use with MultiHeadAttention
    """
    # Create a 2D upper triangular matrix with ones (will be masked positions)
    mask_2d = 1 - tf.linalg.band_part(tf.ones((size, size)), -1, 0)
    
    # Expand dimensions to (1, 1, size, size) for compatibility with MultiHeadAttention
    # This shape represents (batch_size, num_heads, query_seq_len, key_seq_len)
    mask_4d = mask_2d[tf.newaxis, tf.newaxis, :, :]
    
    return mask_4d  # Shape: (1, 1, size, size)

# Add Transformer Encoder Block
class TransformerEncoder(layers.Layer):
    """
    Transformer Encoder block with multi-head attention and feed-forward network.
    """
    def __init__(self, d_model, num_heads, ff_dim, dropout=0.1, **kwargs):
        super().__init__(**kwargs)
        # Self-attention
        self.mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        self.ffn = keras.Sequential([
            layers.Dense(ff_dim, activation='relu'),
            layers.Dense(d_model),
        ])

        self.layernorm1 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm2 = layers.LayerNormalization(epsilon=1e-6)
        self.dropout1 = layers.Dropout(dropout)
        self.dropout2 = layers.Dropout(dropout)

    def call(self, x, training=False, mask=None):
        """
        Process input through the transformer encoder block.
        
        Args:
            x: Input tensor of shape (batch_size, seq_len, d_model)
            training: Whether in training mode
            mask: Optional attention mask of shape (batch_size, 1, 1, seq_len)
        
        Returns:
            Processed tensor of shape (batch_size, seq_len, d_model)
        """
        # Self Attention
        attn_output, _ = self.mha(
            query=x, value=x, key=x, 
            attention_mask=mask,
            return_attention_scores=True
        )
        attn_output = self.dropout1(attn_output, training=training)
        out1 = self.layernorm1(x + attn_output)

        # Feed-Forward
        ffn_output = self.ffn(out1)
        ffn_output = self.dropout2(ffn_output, training=training)
        out2 = self.layernorm2(out1 + ffn_output)

        return out2

# Add Transformer Decoder Block
class TransformerDecoder(layers.Layer):
    """
    Transformer Decoder block with masked self-attention, cross-attention and feed-forward network.
    """
    def __init__(self, d_model, num_heads, ff_dim, dropout=0.1, **kwargs):
        super().__init__(**kwargs)
        # Self-attention (decoder-side)
        self.self_mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        # Cross-attention (encoder-to-decoder)
        self.cross_mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        self.ffn = keras.Sequential([
            layers.Dense(ff_dim, activation='relu'),
            layers.Dense(d_model),
        ])

        self.layernorm1 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm2 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm3 = layers.LayerNormalization(epsilon=1e-6)
        self.dropout1 = layers.Dropout(dropout)
        self.dropout2 = layers.Dropout(dropout)
        self.dropout3 = layers.Dropout(dropout)

    def call(self, x, enc_output, training=False,
             look_ahead_mask=None, padding_mask=None):
        """
        x shape:          (batch_size, target_seq_len, d_model)  # decoder input
        enc_output shape: (batch_size, input_seq_len, d_model)
        look_ahead_mask:  (1, 1, target_seq_len, target_seq_len) or None
        padding_mask:     (batch_size, 1, 1, input_seq_len) or None
        """
        # Apply the look_ahead_mask if not provided but needed
        target_seq_len = tf.shape(x)[1]
        if look_ahead_mask is None:
            look_ahead_mask = create_look_ahead_mask(target_seq_len)
            
        # 1) Masked self-attention
        # The mask shape should be (1, 1, target_seq_len, target_seq_len)
        attn1, attn_weights_1 = self.self_mha(
            query=x, value=x, key=x, 
            attention_mask=look_ahead_mask,
            return_attention_scores=True
        )
        attn1 = self.dropout1(attn1, training=training)
        out1 = self.layernorm1(x + attn1)

        # 2) Cross-attention with encoder output
        # For cross-attention, padding_mask shape should be (batch_size, 1, target_seq_len, input_seq_len)
        attn2, attn_weights_2 = self.cross_mha(
            query=out1, value=enc_output, key=enc_output,
            attention_mask=padding_mask,
            return_attention_scores=True
        )
        attn2 = self.dropout2(attn2, training=training)
        out2 = self.layernorm2(out1 + attn2)

        # 3) Feed-forward
        ffn_output = self.ffn(out2)
        ffn_output = self.dropout3(ffn_output, training=training)
        out3 = self.layernorm3(out2 + ffn_output)

        return out3

# IMPROVED FEATURE GATE WITH BETTER GRADIENT FLOW
class ImprovedFeatureGate(layers.Layer):
    def __init__(self, num_features, k_percent=0.5, temperature=1.0, **kwargs):
        super().__init__(**kwargs)
        self.num_features = num_features
        self.k_percent = k_percent
        self.temperature = temperature  # For controlling sharpness of selection

    def build(self, input_shape):
        # Trainable importance scores
        self.logits = self.add_weight(
            shape=(self.num_features,),
            initializer="random_normal",
            trainable=True,
            name="feature_importance_logits"
        )

    def call(self, inputs, training=None):
        """
        Two approaches for better gradient flow:
        1. Soft gating (training) - allows gradients to all features
        2. Hard gating (inference) - actual top-k selection
        """
        # Convert logits to importance probabilities
        importance_scores = tf.nn.sigmoid(self.logits)
        
        if training:
            # SOFT GATING during training - better gradient flow
            # Use temperature-scaled sigmoid for soft selection
            soft_gates = tf.nn.sigmoid(self.logits / self.temperature)
            
            # Optional: Apply additional sharpening based on percentile
            threshold = tf.nn.top_k(soft_gates, k=int(self.num_features * self.k_percent))[0][-1]
            soft_gates = tf.nn.sigmoid((soft_gates - threshold) / self.temperature)
            
            # Reshape for broadcasting: (1, num_features, 1)
            gates_reshaped = tf.reshape(soft_gates, (1, -1, 1))
            
        else:
            # HARD GATING during inference - actual top-k selection
            k = tf.cast(tf.math.ceil(self.k_percent * tf.cast(self.num_features, tf.float32)), tf.int32)
            
            # Get top-k indices
            _, top_k_indices = tf.nn.top_k(importance_scores, k=k)
            
            # Create hard mask
            mask = tf.zeros_like(importance_scores)
            mask = tf.tensor_scatter_nd_update(
                mask,
                tf.expand_dims(top_k_indices, axis=1),
                tf.ones((k,), dtype=tf.float32)
            )
            
            # Apply hard mask
            hard_gates = importance_scores * mask
            gates_reshaped = tf.reshape(hard_gates, (1, -1, 1))
        
        # Apply gating to inputs
        return inputs * gates_reshaped
    
    def get_feature_importance(self):
        """Get current feature importance scores"""
        return tf.nn.sigmoid(self.logits).numpy()
    
    def get_selected_features(self):
        """Get indices of top-k features"""
        importance = tf.nn.sigmoid(self.logits)
        k = tf.cast(tf.math.ceil(self.k_percent * tf.cast(self.num_features, tf.float32)), tf.int32)
        _, top_k_indices = tf.nn.top_k(importance, k=k)
        return top_k_indices.numpy()

# Alternative simpler approach - Learnable feature weights without hard selection
class SoftFeatureGate(layers.Layer):
    def __init__(self, num_features, **kwargs):
        super().__init__(**kwargs)
        self.num_features = num_features

    def build(self, input_shape):
        # Learnable importance weights
        self.feature_weights = self.add_weight(
            shape=(self.num_features,),
            initializer="ones",  # Start with equal importance
            trainable=True,
            name="feature_weights"
        )

    def call(self, inputs):
        # Apply sigmoid to ensure weights are in [0,1]
        normalized_weights = tf.nn.sigmoid(self.feature_weights)
        
        # Reshape for broadcasting: (1, num_features, 1)
        weights_reshaped = tf.reshape(normalized_weights, (1, -1, 1))
        
        # Apply learnable weights to features
        return inputs * weights_reshaped
    
    def get_feature_importance(self):
        """Get current feature importance scores"""
        return tf.nn.sigmoid(self.feature_weights).numpy()

def data_preprocess(df, diff_order, start_index, end_index,moving_average,MA_window_size):
    '''
    This function removes columns with more than 10% missing values and fills in the NaNs in the remaining columns.
    '''
    # start_date = pd.to_datetime(start_date)
    # end_date = pd.to_datetime(end_date)

    df = pd.DataFrame(df)
    df = df.loc[start_index:end_index]

    # Apply differencing according to the specified diff_order
    df_diff = df.copy()
    for _ in range(diff_order):

        if moving_average==True:
            # Assuming df_diff is your DataFrame

            # Function to pad the beginning of the DataFrame
            def pad_head(df, pad_width):
                # head = pd.DataFrame([df.iloc[0]] * pad_width)
                head = pd.DataFrame([df.iloc[0].values] * pad_width, columns=df.columns)
                return pd.concat([head, df]).reset_index(drop=True)

            # Apply padding only to the head
            pad_width = MA_window_size - 1
            df_padded = pad_head(df_diff, pad_width)

            # Compute the moving average on the padded DataFrame
            df_ma_padded = df_padded.rolling(window=MA_window_size, center=False).mean()

            # Now, remove the padding by trimming the extra rows from the start of the moving average DataFrame
            df_ma = df_ma_padded.iloc[pad_width:].reset_index(drop=True)

            # Shift the moving average DataFrame by one to align it for differencing
            df_ma_shifted = df_ma.shift(1)

            # Ensure df_diff is aligned in size with df_ma_shifted
            df_diff_aligned = df_diff.iloc[:df_ma_shifted.shape[0]]

            # Calculate the difference between the actual value at t+1 and the moving average at t
            df_diff = df_diff_aligned - df_ma_shifted
        else:
            df_diff = df_diff.diff(periods=1)

    # # Select the related range of data
    # df2 = df2.loc[start_index:end_index]
    # Adjust the index offset for the differencing order
    df_diff = df_diff.loc[start_index + diff_order:end_index]

    return df, df_diff

def preprocessing(data_, data_diff_, diff_order, sequence_length, horizon, stride, use_graph_layer,
                  save_instance,ignore_first_instance_stride, save_x_seq, save_y_seq, save_correlation_seq,
                  save_decoder_y_seq, save_last_known_values, target_name,
                  target_as_feature, target_comp_updated_list, differenced_target, differenced_X, moving_average,MA_window_size):
    '''

    This function takes in a time series data, normalizes it, generates sequences, and splits the data into
    training and testing sets.

    data: Time series data
    sequence_length: Length of the sequence used in the model
    horizon: The number of steps to predict in the future
    stride: The step size used to generate sequences
    test_size: The proportion of the data used for testing
    scaler: Scaler used to normalize the data
    '''
    if differenced_X == True:
        X = data_diff_
    else:
        # first row of X is dropped because of differencing
        X = data_[diff_order:]

    X_corr = data_diff_

    if differenced_target == True:
        Y = data_diff_[target_name].values
    else:
        Y = data_[diff_order:][target_name].values

    Y_original = data_[diff_order:][target_name].values

    if target_as_feature == False:
        # Drop columns from X based on target list
        X = X.drop(target_comp_updated_list, axis=1)
        X_corr = X_corr.drop(target_comp_updated_list, axis=1)
    X = X.values

    # Sequence Generation Part
    x_seq = []
    y_seq = []
    correlation_seq = []
    decoder_y_seq = []
    n = np.shape(X)[1]
    # Create ones matrix of size (n, n)
    corr = np.ones((n, n))

    # def calculate_corr(j, k, df_window):
    #     return j, k, chatterjee(df_window[:, j], df_window[:, k])

    if use_graph_layer == True:
        if cal_all_adj == False:
            # restart save_correlation
            save_correlation_seq = []

            # calculate starting correlation based on all available data
            # results = Parallel(n_jobs=-1)(delayed(calculate_corr)(j, k, X[:len(X) - horizon]) for j in range(n) for k in range(n))

            # # Populate the correlation matrix
            # for j, k, result in results:
            #     corr[j][k] = result

            # we put absolute value because GCN does not accept negative adjacency
            corr = np.corrcoef(X_corr[:len(X_corr) - horizon], rowvar=False)
            # Apply the filter to keep only correlations with absolute values greater than 0.05
            corr = np.where(np.abs(corr) > 0.05, corr, 0)

            # Normalize the corr to be in the range [0, 1]
            # normalized_corr = (1 + corr) / 2
            correlation_seq.append(corr)

    for instance in range(save_instance, len(X) - sequence_length - horizon +1, stride):
        if ignore_first_instance_stride:
            ignore_first_instance_stride = False
            continue
        x_seq.append(X[instance: instance + sequence_length])

        if Y_sequence == True:
            y_seq.append(Y_original[instance + sequence_length:instance + sequence_length + horizon])
        else:
            y_seq.append(np.expand_dims(np.sum(Y_original[instance + sequence_length:instance + sequence_length + horizon], axis=0), axis=0))

        decoder_y_seq.append(Y[instance + sequence_length - horizon:instance + sequence_length])
        df_window = X[instance: instance + sequence_length]

        if use_graph_layer == True:
            if cal_all_adj == True:
                # Parallelize the calculation of correlations
                results = Parallel(n_jobs=-1)(
                    delayed(calculate_corr)(j, k, df_window) for j in range(n) for k in range(n))

                # Populate the correlation matrix
                for j, k, result in results:
                    corr[j][k] = result

                correlation_seq.append(corr)

        # # Immediately after sequence generation, capture the last known value
        # if differenced_target == True:

        if moving_average==True:
            ## Apply moving average on last_value_before_sequence
            last_value_before_sequence = data_[target_name].iloc[max(0,
                                                                     instance + sequence_length - diff_order + 2 - MA_window_size):instance + sequence_length - diff_order + 2].mean()
        else:
            # Adjust the index to capture the last known value correctly
            last_value_before_sequence = data_[target_name].iloc[instance + sequence_length - diff_order + 1]

        last_value_before_sequence = np.expand_dims(last_value_before_sequence, axis=0)
        save_last_known_values.append(last_value_before_sequence)
        # else:
        #     save_last_known_values.append(1)

    del X
    save_instance = instance
    save_x_seq.extend(x_seq)
    save_y_seq.extend(y_seq)
    save_decoder_y_seq.extend(decoder_y_seq)
    save_correlation_seq.extend(correlation_seq)

    # last_known_values_input = [item[0] for item in last_known_values]  # Reshape as necessary
    return save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values

def plot_prediction_graph(original_y, forecast_list, test_boundary,
                          horizon, Test_KPI, iter, save_dict):
    plt.figure(figsize=(15, 10))
    # plt.axvline(x=test_boundary, ls=':', color='black')
    # plt.plot(test_index, actual_list[:len(forecast_list)], 'r--',marker='o')
    plt.plot(original_y.index[-horizon:], original_y[-horizon:], 'r--', marker='o')
    plt.plot(forecast_list[-1][1], forecast_list[-1][0], 'b', linewidth=0.6, marker='o')

    # Create a legend with custom labels
    actual_labels = ['actual_{}'.format(i) for i in range(len(target_name))]
    forecast_labels = ['forecast_{}'.format(i) for i in range(len(target_name))]
    legend_labels = actual_labels + forecast_labels
    plt.legend(legend_labels)

    # plt.legend(['test boundary', 'actual', 'prediction'])
    plt.xlabel('Date')
    plt.ylabel('Value')
    plt.text(0.02, 0.8, 'KPI for Test Range = {:.4f}'.format(Test_KPI), bbox=dict(facecolor='red', alpha=0.2),
             transform=plt.gca().transAxes, va="top", ha="left")
    plt.title('{}- Ahead Model_it={}'.format(horizon, iter + 1), loc='center')

    plt.savefig("{}\prediction_{}_it={}.png".format(save_dict, horizon, iter + 1), dpi=300)
    plt.close()

def plot_norm_diff_graph(Y_seq, y_pred, datetime_index, test_boundary,
                    horizon, Test_KPI, iter, save_dict):
    plt.figure(figsize=(15, 10))
    plt.axvline(x=test_boundary, ls=':', color='black')

    for i in range(-1, 0):
        plt.plot(datetime_index[i], Y_seq[i], 'r--', marker='o')
        plt.plot(datetime_index[i], y_pred[i], 'b', linewidth=0.6, marker='o')

    # Create a legend with custom labels
    actual_labels = ['actual_{}'.format(i) for i in range(len(target_name))]
    forecast_labels = ['forecast_{}'.format(i) for i in range(len(target_name))]
    legend_labels = ['test boundary'] + actual_labels + forecast_labels
    plt.legend(legend_labels)

    plt.xlabel('Date')
    plt.ylabel('Value')
    plt.text(0.02, 0.8, 'KPI for Test Range = {:.4f}'.format(Test_KPI), bbox=dict(facecolor='red', alpha=0.2),
             transform=plt.gca().transAxes, va="top", ha="left")
    plt.title('norm_diff_{}- Ahead Model_it={}'.format(horizon, iter + 1), loc='center')

    plt.savefig("{}\dnorm_diff_{}_it={}.png".format(save_dict, horizon, iter + 1), dpi=300)
    plt.close()

class ReverseDifferencingLayer(tf.keras.layers.Layer):
    def __init__(self, **kwargs):
        super(ReverseDifferencingLayer, self).__init__(**kwargs)

    def call(self, inputs):
        """
        Perform the reversal of differencing operation.

        Args:
        inputs: A list of two tensors, where
                inputs[0] is the tensor of predicted differenced values of shape (batch_size, sequence_length, len(target)),
                and inputs[1] is the tensor of the last known value of the series of shape (batch_size, 1, len(target)).

        Returns:
        A tensor of shape (batch_size, sequence_length, len(target)) containing the predictions in the original scale.
        """
        preds, last_known_value = inputs
        # Use tf.cumsum to reverse the differencing, adding the last known value as the base.
        # TensorFlow handles the broadcasting automatically.
        reversed_preds = tf.cumsum(preds, axis=1) + last_known_value
        return reversed_preds

def final_prediction(x_seq,Y_scaler, scaler2, decoder_Y_seq,last_known_values, correlation_seq, sequence_length, horizon, model, data, stride,
                     target_name, differenced_target, diff_order, save_dict):

    x_seq = np.expand_dims(x_seq[-1], axis=0)
    decoder_Y_seq = np.expand_dims(decoder_Y_seq[-1], axis=0)
    last_known_values = np.expand_dims(last_known_values[-1], axis=0)

    if len(correlation_seq) > 0:
        correlation_seq = np.expand_dims(correlation_seq[-1], axis=0)

    # Prediction
    y_pred = model.predict([x_seq, correlation_seq, decoder_Y_seq, last_known_values])

    if use_graph_layer:
        # Get attention coefficients
        gc_1_attn, gc_2_attn = model.attention_submodel.predict([x_seq, correlation_seq, decoder_Y_seq, last_known_values])
        # Visualize attention

        # Assuming gc_1_attn and gc_2_attn have shape (batch_size, N, attn_heads, N)
        # Step 1: Merge the heads by averaging
        gc_1_attn_merged = np.mean(gc_1_attn, axis=2)  # Shape: (batch_size, N, N)
        gc_2_attn_merged = np.mean(gc_2_attn, axis=2)  # Shape: (batch_size, N, N)

        # Step 2: Visualize attention coefficients for the first batch
        attn_coef_mean_1 = gc_1_attn_merged[0]  # Select the first graph in the batch
        plt.figure(figsize=(6, 4))
        sns.heatmap(attn_coef_mean_1, cmap="viridis")
        plt.title("Attention Coefficients - GAT Layer 1 (Merged Heads)")
        plt.xlabel("Node j")
        plt.ylabel("Node i")
        plt.savefig("gat_layer_1.png")
        plt.savefig(f"{save_dict}/gat_layer_1_it={iter + 1}.png", dpi=300)

        attn_coef_mean_2 = gc_2_attn_merged[0]  # Select the first graph in the batch
        plt.figure(figsize=(6, 4))
        sns.heatmap(attn_coef_mean_2, cmap="viridis")
        plt.title("Attention Coefficients - GAT Layer 2 (Merged Heads)")
        plt.xlabel("Node j")
        plt.ylabel("Node i")
        plt.savefig(f"{save_dict}/gat_layer_2_it={iter + 1}.png", dpi=300)

    y_pred_rev = y_pred

    if normalized_data==True:

        if reverse_normalization==True:
            y_pred_rev = np.expand_dims(Y_scaler.inverse_transform(np.squeeze(y_pred_rev, axis=0)),
                                     axis=0)  # reverse scaler to plot later in the original scale
        else:
            # It seems papers used normalized data for prediction results
            y_pred_rev = np.expand_dims(np.squeeze(y_pred_rev, axis=0),
                                        axis=0)
    else:
        # y_pred_rev = scaler2.transform(np.squeeze(y_pred_rev, axis=0))
        # y_pred_rev = np.expand_dims(y_pred_rev,
        #                         axis=0) # reverse scaler to plot later in the original scale

        y_pred_rev = np.expand_dims(np.squeeze(y_pred_rev, axis=0),
                                    axis=0)

    n_rows = data.shape[0] - sequence_length - horizon
    indices = np.arange(1 + sequence_length, 1 + n_rows + sequence_length, stride)

    if Y_sequence == True:

        if normalized_data == True:
            if reverse_normalization == True:
                actual_array = np.expand_dims(Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + np.arange(horizon)]), axis=0)
            else:
                # It seems papers used normalized data for prediction results
                actual_array = np.expand_dims(
                    data[target_name].values[indices[-1, None] + np.arange(horizon)], axis=0)
        else:

            actual_array = np.expand_dims(
                data[target_name].values[indices[-1, None] + np.arange(horizon)], axis=0)

        datetime_index = np.expand_dims(data.index.to_numpy()[indices[-1, None] + np.arange(horizon)], axis=0)
    else:
        if normalized_data == True:
            actual_array = Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + horizon - 1])
        else:
            actual_array = data[target_name].values[indices[-1, None] + horizon - 1]
        datetime_index = data.index.to_numpy()[indices[-1, None] + horizon - 1]

    return actual_array, y_pred_rev, datetime_index, y_pred

def data_set_generation(data, data_diff, max_lag, target_as_feature, target_name):
    # update the input_components list
    input_components = list(data.columns)

    # remove the target column before lagging the variables
    if not target_as_feature:
        for target in target_name:
            if target in input_components:
                input_components.remove(target)

    # List of updated input components.
    target_comp_updated_list = []

    # Initialize forecast dataset.
    new_data = pd.DataFrame()
    new_data_diff = pd.DataFrame()

    # Iterate through input components and their lags.
    for col in input_components:
        for lag in range(0, max_lag + 1):
            # Generate new component name.
            comp_new = f"{col} ({-lag})"
            if col in target_name:
                target_comp_updated_list.append(comp_new)
            # Get start index based on optimal time lag.
            new_data_time_series = data[col].shift(lag).rename(comp_new)
            new_data_diff_time_series = data_diff[col].shift(lag).rename(comp_new)

            # Add to dataset and forecast dataset
            new_data = pd.concat([new_data, new_data_time_series], axis=1)
            new_data_diff = pd.concat([new_data_diff, new_data_diff_time_series], axis=1)

    # Drop the first few rows to remove Nans.

    new_data = new_data.iloc[max_lag:]
    new_data_diff = new_data_diff.iloc[max_lag:]

    update_target_name = []
    if target_as_feature == True:
        for i, col in enumerate(target_name):
            update_target_name.append(f"{col} (0)")
        target_name = update_target_name
    else:
        # append the removed target column to datasets
        new_data[target_name] = data[target_name]
        new_data_diff[target_name] = data_diff[target_name]
        target_comp_updated_list.extend(target_name)

    # new_data = new_data.reset_index(drop=True)
    # new_data_diff = new_data_diff.reset_index(drop=True)

    return new_data, new_data_diff, target_comp_updated_list, target_name

def CORR(flattened_actual, flattened_forecast):
    """
    Calculate the Empirical Correlation Coefficient between two flattened arrays.

    Parameters:
    flattened_actual (array-like): Flattened array of actual values.
    flattened_forecast (array-like): Flattened array of forecasted values.

    Returns:
    float: Empirical Correlation Coefficient.
    """

    # Ensuring the inputs are numpy arrays
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    # Calculate the standard deviation and mean of the inputs
    sigma_p = flattened_forecast.std()
    sigma_g = flattened_actual.std()
    mean_p = flattened_forecast.mean()
    mean_g = flattened_actual.mean()

    # Calculate the correlation only if sigma_g and sigma_p are not equal to zero
    valid_indices = (sigma_g != 0) and (sigma_p != 0)
    if np.any(valid_indices):
        correlation = np.mean(
            ((flattened_forecast[valid_indices] - mean_p) * (flattened_actual[valid_indices] - mean_g)) / (sigma_p * sigma_g)
        )
    else:
        correlation = 0

    return correlation

def RSE(flattened_actual, flattened_forecast):
    """
    Calculate the Root Relative Squared Error between two flattened arrays or multi-dimensional arrays.

    Parameters:
    flattened_actual (array-like): Flattened array or multi-dimensional array of actual values.
    flattened_forecast (array-like): Flattened array or multi-dimensional array of forecasted values.

    Returns:
    float: Root Relative Squared Error.
    """

    # Ensuring the inputs are numpy arrays
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    # Calculate the mean of the actual values
    mean_actual = np.mean(flattened_actual)

    # Calculate the numerator (sum of squared differences between actual and forecast values)
    numerator = np.sum((flattened_actual - flattened_forecast) ** 2)

    # Calculate the denominator (sum of squared differences between actual values and mean of actual values)
    denominator = np.sum((flattened_actual - mean_actual) ** 2)

    # Calculate RSE
    if denominator != 0:
        rse = np.sqrt(numerator / denominator)
    else:
        rse = np.nan

    return rse

def mean_absolute_scaled_error(y_true, y_pred):
    naive_forecast = tf.roll(y_true, shift=1, axis=0)  # Shift the tensor by 1 position
    errors = tf.abs(y_true - y_pred)
    naive_errors = tf.abs(y_true - naive_forecast)
    epsilon = 1e-8  # Small constant to prevent division by zero
    mase = tf.reduce_mean(errors / (naive_errors + epsilon))
    return mase

def symmetric_mean_absolute_percentage_error(y_true, y_pred):
    y_true = K.maximum(K.epsilon(), y_true)  # Avoid division by zero
    return K.mean(2 * K.abs(y_pred - y_true) / (y_pred + y_true), axis=-1)

def symmetric_mean_absolute_percentage_error_loss(y_true, y_pred):
    epsilon = 1e-7  # A small constant to avoid division by zero
    y_true = K.maximum(epsilon, y_true)
    smape_loss = K.mean(2 * K.abs(y_pred - y_true) / (y_pred + y_true), axis=-1)
    return smape_loss

# =============================================================================
# STABILITY TESTING FUNCTIONS
# =============================================================================

def test_feature_importance_stability(seeds_to_test, model_training_function, lagged_feature_names):
    """
    Test how stable feature importance is across different random seeds
    
    Args:
        seeds_to_test: List of random seeds to test
        model_training_function: Function that trains and returns a model given a seed
        lagged_feature_names: List of feature names for visualization
    
    Returns:
        DataFrame with feature importance for each seed
    """
    
    importance_results = {}
    
    for seed in seeds_to_test:
        print(f"🔄 Training stability test with seed {seed}...")
        
        # Set all random seeds
        tf.random.set_seed(seed)
        os.environ['PYTHONHASHSEED'] = str(seed)
        np.random.seed(seed)
        random.seed(seed)
        
        # Train model
        model = model_training_function(seed)
        
        # Extract feature importance
        feature_importance = get_feature_importance_from_model(model)
        
        if feature_importance is not None:
            importance_results[f'seed_{seed}'] = feature_importance
        
        # Clear session for next iteration
        clear_session()
        gc.collect()
    
    # Convert to DataFrame for analysis
    importance_df = pd.DataFrame(importance_results)
    
    return importance_df

def get_feature_importance_from_model(model):
    """
    Extract feature importance from trained model
    """
    # Find the FeatureGate layer in the model
    feature_gate_layer = None
    for layer in model.layers:
        if isinstance(layer, (ImprovedFeatureGate, SoftFeatureGate)):
            feature_gate_layer = layer
            break
            
    # If nested, search through sub-models
    if feature_gate_layer is None:
        for layer in model.layers:
            if hasattr(layer, 'layers'):
                for sublayer in layer.layers:
                    if isinstance(sublayer, (ImprovedFeatureGate, SoftFeatureGate)):
                        feature_gate_layer = sublayer
                        break
    
    if feature_gate_layer is None:
        print("⚠️  Could not find FeatureGate layer in the model!")
        return None
        
    # Get feature importance values
    importance = feature_gate_layer.get_feature_importance()
    return importance

def analyze_stability(importance_df, lagged_feature_names=None):
    """
    Analyze stability of feature importance across different seeds
    """
    
    # Calculate statistics across seeds
    mean_importance = importance_df.mean(axis=1)
    std_importance = importance_df.std(axis=1)
    cv_importance = std_importance / (mean_importance + 1e-8)  # Coefficient of variation with small epsilon
    
    # Create results dataframe
    stability_results = pd.DataFrame({
        'feature_index': range(len(mean_importance)),
        'mean_importance': mean_importance,
        'std_importance': std_importance, 
        'coefficient_of_variation': cv_importance,
        'min_importance': importance_df.min(axis=1),
        'max_importance': importance_df.max(axis=1)
    })
    
    if lagged_feature_names and len(lagged_feature_names) == len(mean_importance):
        stability_results['feature_name'] = lagged_feature_names
    
    # Sort by mean importance
    stability_results = stability_results.sort_values('mean_importance', ascending=False)
    
    return stability_results

def plot_stability_analysis(importance_df, stability_results, save_path, lagged_feature_names=None):
    """
    Create comprehensive visualizations for feature importance stability
    """
    
    # Helper function to get feature name with truncation for display
    def get_display_name(idx, max_length=25):
        if lagged_feature_names and idx < len(lagged_feature_names):
            name = lagged_feature_names[idx]
            if len(name) > max_length:
                return name[:max_length-3] + "..."
            return name
        return f"Feature {idx}"
    
    # Helper function to get full feature name
    def get_full_name(idx):
        if lagged_feature_names and idx < len(lagged_feature_names):
            return lagged_feature_names[idx]
        return f"Feature {idx}"
    
    fig, axes = plt.subplots(2, 2, figsize=(22, 18))
    
    # 1. Heatmap of feature importance across seeds (top 20 features for better readability)
    top_n = min(20, len(importance_df))
    top_features = stability_results.head(top_n)['feature_index'].values
    
    heatmap_data = importance_df.loc[top_features]
    
    # Create feature labels for heatmap (truncated for space)
    feature_labels = [get_display_name(i, max_length=30) for i in top_features]
    
    sns.heatmap(heatmap_data, annot=True, fmt='.3f', ax=axes[0,0], cmap='viridis',
                yticklabels=feature_labels, cbar_kws={'label': 'Importance Score'})
    axes[0,0].set_title(f'Feature Importance Across Different Seeds\n(Top {top_n} Features)', fontsize=14, fontweight='bold')
    axes[0,0].set_xlabel('Random Seed', fontsize=12)
    axes[0,0].set_ylabel('Features', fontsize=12)
    
    # 2. Mean importance vs standard deviation (stability plot)
    scatter = axes[0,1].scatter(stability_results['mean_importance'], 
                               stability_results['std_importance'], 
                               c=stability_results['coefficient_of_variation'],
                               cmap='RdYlBu_r', alpha=0.7, s=80)
    axes[0,1].set_xlabel('Mean Importance', fontsize=12)
    axes[0,1].set_ylabel('Standard Deviation', fontsize=12)
    axes[0,1].set_title('Feature Importance Stability\n(Color = Coefficient of Variation)', fontsize=14, fontweight='bold')
    
    # Add colorbar
    cbar = plt.colorbar(scatter, ax=axes[0,1])
    cbar.set_label('Coefficient of Variation\n(Lower = More Stable)', fontsize=11)
    
    # Annotate most/least stable high-importance features with actual names
    high_importance = stability_results[stability_results['mean_importance'] > 0.3]
    if len(high_importance) > 0:
        most_stable = high_importance.nsmallest(2, 'coefficient_of_variation')  # Reduced to 2 to avoid clutter
        least_stable = high_importance.nlargest(2, 'coefficient_of_variation')
        
        for idx, row in most_stable.iterrows():
            feature_name = get_display_name(row["feature_index"], max_length=20)
            axes[0,1].annotate(f'Stable: {feature_name}', 
                              (row['mean_importance'], row['std_importance']),
                              xytext=(5, 5), textcoords='offset points', 
                              fontsize=9, color='darkgreen', fontweight='bold',
                              bbox=dict(boxstyle="round,pad=0.3", facecolor='lightgreen', alpha=0.7))
        
        for idx, row in least_stable.iterrows():
            feature_name = get_display_name(row["feature_index"], max_length=20)
            axes[0,1].annotate(f'Unstable: {feature_name}', 
                              (row['mean_importance'], row['std_importance']),
                              xytext=(5, -15), textcoords='offset points', 
                              fontsize=9, color='darkred', fontweight='bold',
                              bbox=dict(boxstyle="round,pad=0.3", facecolor='lightcoral', alpha=0.7))
    
    # 3. Box plot of top 10 most important features (reduced from 15 for better readability)
    top_n_box = min(10, len(stability_results))
    top_features_box = stability_results.head(top_n_box)['feature_index'].values
    top_data_box = importance_df.loc[top_features_box].T

    box_data = [top_data_box.iloc[:, i] for i in range(len(top_features_box))]

    # Create labels with actual feature names (print completely)
    box_labels = [get_full_name(i) for i in top_features_box]

    bp = axes[1,0].boxplot(box_data, labels=box_labels, patch_artist=True)
    axes[1,0].set_xticklabels(box_labels, rotation=45, ha='right', fontsize=10, wrap=True)

    # Color boxes based on stability
    cv_values = [stability_results[stability_results['feature_index'] == f]['coefficient_of_variation'].iloc[0] 
                 for f in top_features_box]

    # Normalize CV values for color mapping
    if len(cv_values) > 0 and max(cv_values) > 0:
        colors = plt.cm.RdYlGn_r([cv/max(cv_values) for cv in cv_values])
        for patch, color in zip(bp['boxes'], colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.8)

    axes[1,0].set_title(f'Distribution of Importance for Top {top_n_box} Features\n(Red = Less Stable, Green = More Stable)', 
                       fontsize=14, fontweight='bold')
    axes[1,0].set_xlabel('Features', fontsize=12)
    axes[1,0].set_ylabel('Importance Score', fontsize=12)
    axes[1,0].tick_params(axis='x', rotation=45, labelsize=10)

    # Add a text box with full feature names for the box plot
    full_names_text = "Feature Names:\n" + "\n".join([f"{i+1}. {get_full_name(top_features_box[i])}" 
                                                      for i in range(min(5, len(top_features_box)))])
    if len(top_features_box) > 5:
        full_names_text += f"\n... and {len(top_features_box)-5} more"

    axes[1,0].text(0.02, 0.98, full_names_text, transform=axes[1,0].transAxes, 
                   verticalalignment='top', fontsize=8, 
                   bbox=dict(boxstyle="round,pad=0.5", facecolor='lightyellow', alpha=0.8))
    
    # 4. Stability distribution histogram
    axes[1,1].hist(stability_results['coefficient_of_variation'], bins=30, alpha=0.7, 
                   color='skyblue', edgecolor='black')
    axes[1,1].set_xlabel('Coefficient of Variation (std/mean)', fontsize=12)
    axes[1,1].set_ylabel('Number of Features', fontsize=12)
    axes[1,1].set_title('Distribution of Feature Stability', fontsize=14, fontweight='bold')
    
    median_cv = stability_results['coefficient_of_variation'].median()
    axes[1,1].axvline(median_cv, color='red', linestyle='--', linewidth=2,
                     label=f'Median CV: {median_cv:.3f}')
    
    # Add stability thresholds
    axes[1,1].axvline(0.3, color='green', linestyle=':', alpha=0.7, label='Stable (CV < 0.3)')
    axes[1,1].axvline(0.7, color='orange', linestyle=':', alpha=0.7, label='Unstable (CV > 0.7)')
    axes[1,1].legend(fontsize=10)
    
    # Add grid for better readability
    for ax in axes.flat:
        ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    # === NEW: Violinplot for feature importance distribution (top 10 features) ===
    # Prepare long-form DataFrame for violinplot
    if lagged_feature_names and 'feature' not in importance_df.columns:
        importance_df = importance_df.copy()
        importance_df['feature'] = lagged_feature_names
    elif 'feature' not in importance_df.columns:
        importance_df = importance_df.copy()
        importance_df['feature'] = [f'Feature {i}' for i in range(len(importance_df))]

    # Melt to long-form
    long_df = importance_df.melt(id_vars='feature', var_name='seed', value_name='importance')
    # Remove non-numeric seeds (e.g., 'feature' column)
    long_df = long_df[long_df['seed'].str.startswith('seed_')]

    # Select top 10 features by mean importance
    if lagged_feature_names:
        mean_importance = importance_df.drop(columns=['feature']).mean(axis=1)
    else:
        mean_importance = importance_df.drop(columns=['feature']).mean(axis=1)
    top_n_violin = 10
    top_features_violin = mean_importance.sort_values(ascending=False).head(top_n_violin).index
    top_feature_names = [importance_df.loc[i, 'feature'] for i in top_features_violin]
    violin_df = long_df[long_df['feature'].isin(top_feature_names)]

    # Get CV for each top feature for coloring
    cv_map = dict(zip(stability_results['feature_name'] if 'feature_name' in stability_results.columns else [f'Feature {i}' for i in stability_results['feature_index']],
                      stability_results['coefficient_of_variation']))
    # Normalize CVs for color mapping (red=less stable, green=more stable)
    cv_for_palette = [cv_map[f] for f in top_feature_names]
    max_cv = max(cv_for_palette) if len(cv_for_palette) > 0 else 1
    min_cv = min(cv_for_palette) if len(cv_for_palette) > 0 else 0
    norm_cvs = [(cv-min_cv)/(max_cv-min_cv+1e-8) for cv in cv_for_palette]
    violin_colors = [plt.cm.RdYlGn_r(norm) for norm in norm_cvs]

    # Create a palette dict for seaborn
    palette_dict = dict(zip(top_feature_names, violin_colors))

    plt.figure(figsize=(14, 7))
    ax = sns.violinplot(
        data=violin_df,
        x='feature',
        y='importance',
        inner='box',
        fill=True,
        cut=0,
        bw_adjust=0.8,
        palette=palette_dict
    )
    plt.xticks(rotation=45, ha='right')
    plt.title('Feature Importance Distribution Across Seeds (Top 10 Features)\n(Red = Less Stable, Green = More Stable)')
    plt.xlabel('Feature')
    plt.ylabel('Importance')
    plt.tight_layout()
    violin_save_path = save_path.replace('.png', '_violin.png')
    plt.savefig(violin_save_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return fig

def interpret_stability_results(stability_results, lagged_feature_names=None):
    """
    Provide comprehensive interpretation of stability results
    """
    print("\n" + "="*90)
    print("🔍 FEATURE IMPORTANCE STABILITY ANALYSIS")
    print("="*90)
    
    # Helper function to get feature name
    def get_feature_name(idx):
        if lagged_feature_names and idx < len(lagged_feature_names):
            return lagged_feature_names[idx]
        return f"Feature {idx}"
    
    print("\n🟢 MOST RELIABLE FEATURES (High Importance + Low Variability):")
    print("-" * 90)
    stable_important = stability_results[
        (stability_results['mean_importance'] > 0.4) & 
        (stability_results['coefficient_of_variation'] < 0.3)
    ].head(15)
    
    if len(stable_important) > 0:
        print(f"{'Rank':<4} {'Feature Name':<50} {'Mean Imp.':<10} {'CV':<8} {'Status'}")
        print("-" * 90)
        for rank, (idx, row) in enumerate(stable_important.iterrows(), 1):
            feature_name = get_feature_name(row['feature_index'])
            print(f"{rank:<4} {feature_name[:48]:<50} {row['mean_importance']:.3f}     {row['coefficient_of_variation']:.3f}   ✅ Reliable")
    else:
        print("  ⚠️  No features meet the stable + important criteria")
    
    print(f"\n🟡 HIGH IMPACT BUT UNSTABLE FEATURES (Potential Overfitting):")
    print("-" * 90)
    unstable_important = stability_results[
        (stability_results['mean_importance'] > 0.4) & 
        (stability_results['coefficient_of_variation'] > 0.5)
    ].head(10)
    
    if len(unstable_important) > 0:
        print(f"{'Rank':<4} {'Feature Name':<50} {'Mean Imp.':<10} {'CV':<8} {'Status'}")
        print("-" * 90)
        for rank, (idx, row) in enumerate(unstable_important.iterrows(), 1):
            feature_name = get_feature_name(row['feature_index'])
            print(f"{rank:<4} {feature_name[:48]:<50} {row['mean_importance']:.3f}     {row['coefficient_of_variation']:.3f}   ⚠️  Unstable")
    else:
        print("  ✅ No high-importance features are highly unstable")
    
    print(f"\n🔴 CONSISTENTLY LOW IMPORTANCE FEATURES (Safe to Remove):")
    print("-" * 90)
    consistently_low = stability_results[
        stability_results['max_importance'] < 0.15
    ].head(15)
    
    if len(consistently_low) > 0:
        print(f"{'Rank':<4} {'Feature Name':<50} {'Max Imp.':<10} {'Mean Imp.':<10} {'Status'}")
        print("-" * 90)
        for rank, (idx, row) in enumerate(consistently_low.iterrows(), 1):
            feature_name = get_feature_name(row['feature_index'])
            print(f"{rank:<4} {feature_name[:48]:<50} {row['max_importance']:.3f}     {row['mean_importance']:.3f}    🗑️  Removable")
    else:
        print("  📈 All features show some importance across seeds")
    
    # Feature type analysis
    print(f"\n📊 FEATURE TYPE ANALYSIS:")
    print("-" * 90)
    
    if lagged_feature_names:
        # Group by base feature (before lag suffix)
        feature_types = {}
        for idx, row in stability_results.iterrows():
            full_name = get_feature_name(row['feature_index'])
            # Extract base name (before the lag part like "(-0)", "(-1)", etc.)
            if '(' in full_name and ')' in full_name:
                base_name = full_name.split('(')[0].strip()
                lag_part = full_name.split('(')[1].split(')')[0]
                
                if base_name not in feature_types:
                    feature_types[base_name] = []
                feature_types[base_name].append({
                    'lag': lag_part,
                    'importance': row['mean_importance'],
                    'stability': row['coefficient_of_variation'],
                    'full_name': full_name
                })
        
        # Show top base features
        base_importance = {}
        for base_name, lags in feature_types.items():
            avg_importance = np.mean([lag['importance'] for lag in lags])
            max_importance = np.max([lag['importance'] for lag in lags])
            base_importance[base_name] = {
                'avg_importance': avg_importance,
                'max_importance': max_importance,
                'lag_count': len(lags),
                'best_lag': max(lags, key=lambda x: x['importance'])
            }
        
        print("Top Base Features (by average importance across all lags):")
        sorted_base = sorted(base_importance.items(), key=lambda x: x[1]['avg_importance'], reverse=True)
        
        for rank, (base_name, info) in enumerate(sorted_base[:10], 1):
            best_lag_info = info['best_lag']
            print(f"{rank:2d}. {base_name:<25} | Avg: {info['avg_importance']:.3f} | Best lag: {best_lag_info['lag']:<4} (imp: {best_lag_info['importance']:.3f})")
    
    # Overall statistics
    stable_count = len(stability_results[stability_results['coefficient_of_variation'] < 0.3])
    unstable_count = len(stability_results[stability_results['coefficient_of_variation'] > 0.7])
    high_importance_count = len(stability_results[stability_results['mean_importance'] > 0.4])
    
    print(f"\n📊 OVERALL SUMMARY:")
    print("-" * 90)
    print(f"  • Total features analyzed: {len(stability_results)}")
    print(f"  • Stable features (CV < 0.3): {stable_count} ({stable_count/len(stability_results)*100:.1f}%)")
    print(f"  • Unstable features (CV > 0.7): {unstable_count} ({unstable_count/len(stability_results)*100:.1f}%)")
    print(f"  • High importance features (Mean > 0.4): {high_importance_count} ({high_importance_count/len(stability_results)*100:.1f}%)")
    print(f"  • Median stability (CV): {stability_results['coefficient_of_variation'].median():.3f}")
    
    # Show top features by name
    top_stable = stability_results.nsmallest(1, 'coefficient_of_variation')['feature_index'].iloc[0]
    top_important = stability_results.nlargest(1, 'mean_importance')['feature_index'].iloc[0]
    
    print(f"  • Most stable feature: {get_feature_name(top_stable)}")
    print(f"  • Most important feature: {get_feature_name(top_important)}")
    
    print(f"\n💡 RECOMMENDATIONS:")
    print("-" * 90)
    if len(stable_important) >= 10:
        print(f"  ✅ You have {len(stable_important)} reliable features - good foundation for modeling")
    else:
        print(f"  ⚠️  Only {len(stable_important)} truly reliable features - consider feature engineering")
    
    if len(unstable_important) > 0:
        print(f"  🔍 {len(unstable_important)} features show high but unstable importance - investigate for overfitting")
        print(f"      Consider: regularization, cross-validation, or removing these features")
    
    if len(consistently_low) > 10:
        print(f"  🗑️  Consider removing {len(consistently_low)} consistently unimportant features")
        print(f"      This could improve model efficiency and reduce overfitting")
    
    # Calculate recommended k_percent
    recommended_features = len(stable_important) + len(stability_results[
        (stability_results['mean_importance'] > 0.25) & 
        (stability_results['coefficient_of_variation'] < 0.5)
    ])
    
    if recommended_features > 0:
        recommended_percent = min(1.0, recommended_features / len(stability_results))
        print(f"  🎯 Recommended feature_gate_k_percent: {recommended_percent:.2f}")
        print(f"      (Keep top {recommended_features} features out of {len(stability_results)} total)")
    
    print(f"\n🎯 TOP 5 FEATURES TO FOCUS ON:")
    print("-" * 90)
    top_overall = stability_results.head(5)
    for rank, (idx, row) in enumerate(top_overall.iterrows(), 1):
        feature_name = get_feature_name(row['feature_index'])
        stability_status = "Stable" if row['coefficient_of_variation'] < 0.3 else "Unstable" if row['coefficient_of_variation'] > 0.7 else "Moderate"
        print(f"{rank}. {feature_name[:60]:<60} (Imp: {row['mean_importance']:.3f}, {stability_status})")
    
    print("="*90)

def run_stability_analysis(model_training_function, seeds_to_test, lagged_feature_names, save_directory):
    """
    Complete stability analysis workflow
    """
    print("🚀 Starting Feature Importance Stability Analysis...")
    print(f"Testing with seeds: {seeds_to_test}")
    
    # Run stability test
    importance_df = test_feature_importance_stability(seeds_to_test, model_training_function, lagged_feature_names)
    
    if importance_df.empty:
        print("❌ No feature importance data collected - check your model implementation")
        return None, None
    
    # Analyze results
    stability_results = analyze_stability(importance_df, lagged_feature_names)
    
    # Create visualizations
    plot_path = f"{save_directory}/feature_stability_analysis.png"
    plot_stability_analysis(importance_df, stability_results, plot_path, lagged_feature_names)
    
    # Print interpretation
    interpret_stability_results(stability_results, lagged_feature_names)
    
    # Save results to files
    importance_df.to_csv(f"{save_directory}/feature_importance_across_seeds.csv")
    stability_results.to_csv(f"{save_directory}/stability_analysis_results.csv", index=False)
    
    print(f"\n💾 Results saved to:")
    print(f"  - {save_directory}/feature_importance_across_seeds.csv")
    print(f"  - {save_directory}/stability_analysis_results.csv")
    print(f"  - {plot_path}")
    
    return importance_df, stability_results

# =============================================================================
# IMPROVED MODEL BUILDING FUNCTIONS
# =============================================================================

def data_generator(X, correlation, decoder_Y, Y, last_known_values, batch_size, new_data_ratio=0.1):
    while True:
        # Number of new data samples to include in each batch
        new_data_count = int(batch_size * new_data_ratio)

        # Indices of the new data
        new_data_indices = np.arange(len(X) - new_data_count, len(X))

        for i in range(0, len(X), batch_size):
            # Select random indices from the entire dataset
            random_indices = np.random.permutation(len(X) - new_data_count)

            # Combine new data indices with random indices and shuffle
            combined_indices = np.concatenate([new_data_indices, random_indices[:batch_size - new_data_count]])
            np.random.shuffle(combined_indices)

            batch_X = X[combined_indices]
            batch_decoder_Y = decoder_Y[combined_indices]
            batch_last_known_values = last_known_values[combined_indices]
            batch_Y = Y[combined_indices]

            if  len(correlation) > 0:
                if repeat_corr == True:
                    batch_correlation = np.repeat(correlation[np.newaxis, ...], len(batch_X), axis=0)
                else:
                    batch_correlation = correlation[combined_indices]
            else:
                batch_correlation=[]

            yield [batch_X, batch_correlation, batch_decoder_Y, batch_last_known_values], batch_Y

def graph_processing_block(inputs, inp_lap, head_size, num_heads, dropout, horizon):
    """
    Process inputs through graph convolutional layers with improved feature gating.
    """
    l2_reg = 2.5e-4  # L2 regularization rate
    
    # Save original sequence length for reshaping later
    batch_size, original_seq_len, n_features = tf.shape(inputs)[0], tf.shape(inputs)[1], tf.shape(inputs)[2]
    
    # Transpose from (batch, time, features) to (batch, features, time)
    x = tf.transpose(inputs, perm=[0, 2, 1])
    
    # Use improved feature gate based on k_percent
    if feature_gate_k_percent < 1.0:
        # Hard selection with improved gradient flow
        x = ImprovedFeatureGate(
            num_features=int(x.shape[1]), 
            k_percent=feature_gate_k_percent, 
            temperature=0.1,  # Lower = sharper selection
            name='improved_feature_gate'
        )(x)
    else:
        # Soft weighting without hard selection
        x = SoftFeatureGate(
            num_features=int(x.shape[1]),
            name='soft_feature_gate'
        )(x)
    
    do_1 = Dropout(dropout)(x)
    
    # GAT operates on the node/feature dimension
    gc_1, gc_1_attn = GATv2Conv(
        int(math.ceil(x.shape[2])),
        attn_heads=num_heads,
        concat_heads=False,
        dropout_rate=dropout,
        activation="relu",
        kernel_regularizer=l2(l2_reg),
        attn_kernel_regularizer=l2(l2_reg),
        bias_regularizer=l2(l2_reg),
        return_attn_coef=True,
    )([do_1, inp_lap])
    
    # Transpose back to (batch, time, features)
    graph_output = tf.transpose(gc_1, perm=[0, 2, 1])
    
    # For compatibility with the attention visualizations
    gc_2_attn = gc_1_attn
    
    return graph_output, gc_1_attn, gc_2_attn

def transformer_encoder_decoder_block(graph_output, decoder_inputs, d_model, num_heads, ff_dim, horizon, target_name, dropout):
    """
    Transformer-based encoder-decoder block for time series forecasting.
    """
    l2_reg = 2.5e-4  # L2 regularization rate
    
    # Get the actual sequence length from the input for positional encoding
    sequence_length = tf.shape(graph_output)[1]
    max_encoder_length = 200  # Maximum possible sequence length, should be larger than any expected sequence
    
    # Add positional embeddings to encoder input
    encoder_embedding = PositionalEmbedding(max_sequence_length=max_encoder_length, d_model=d_model)
    enc_emb = encoder_embedding(graph_output)
    
    # Encoder stacks (2 layers)
    encoder_output = enc_emb
    for i in range(2):  # Number of encoder layers
        encoder_block = TransformerEncoder(d_model, num_heads, ff_dim, dropout=dropout)
        encoder_output = encoder_block(encoder_output)
    
    # Create look-ahead mask for decoder to ensure causal attention
    look_ahead_mask = create_look_ahead_mask(horizon)  # Now returns shape (1, 1, horizon, horizon)
    
    # Add positional embeddings to decoder input
    decoder_embedding = PositionalEmbedding(max_sequence_length=horizon, d_model=d_model)
    dec_emb = decoder_embedding(decoder_inputs)
    
    # Decoder stacks (2 layers)
    decoder_output = dec_emb
    for i in range(2):  # Number of decoder layers
        decoder_block = TransformerDecoder(d_model, num_heads, ff_dim, dropout=dropout)
        decoder_output = decoder_block(
            decoder_output, 
            encoder_output,
            look_ahead_mask=look_ahead_mask
        )
    
    # Final output layer
    if Y_sequence == True:
        # Output sequence with TimeDistributed dense layer
        outputs = keras.layers.TimeDistributed(keras.layers.Dense(len(target_name)))(decoder_output)
    else:
        # Single output
        outputs = keras.layers.Dense(len(target_name))(decoder_output[:, -1:, :])
    
    return outputs

def build_model(input_shape, correlation_shape, use_graph_layer):
    inputs = keras.Input(shape=input_shape)
    if use_graph_layer == True:
        inp_lap = keras.Input(shape=correlation_shape)
    else:
        inp_lap = []

    if use_graph_layer:
        graph_output, gc_1_attn, gc_2_attn = graph_processing_block(inputs, inp_lap, head_size,
                                                                    num_heads, dropout,
                                                                    horizon)
    else:
        graph_output = inputs

    # Project to consistent dimension for transformer if needed
    d_model = 256  # Transformer embedding dimension
    graph_output = layers.Dense(d_model)(graph_output)
    
    # Create decoder inputs
    decoder_inputs = tf.keras.Input(
        shape=(horizon, len(target_name)), name='decoder_inputs')

    # Use transformer encoder-decoder instead of LSTM
    outputs = transformer_encoder_decoder_block(
        graph_output=graph_output,
        decoder_inputs=decoder_inputs,
        d_model=d_model,
        num_heads=num_heads,
        ff_dim=ff_dim,
        horizon=horizon,
        target_name=target_name,
        dropout=dropout
    )

    # Assume 'last_known_value_input' is an additional input layer for the last known value before prediction starts
    last_known_value_input = tf.keras.Input(shape=(1, len(target_name)), name='last_known_values')
    
    if differenced_target == True:
        # Integrate the ReverseDifferencingLayer
        outputs = ReverseDifferencingLayer()([outputs, last_known_value_input])

    # Return a model with outputs for training and a submodel for attention
    model = keras.Model(
        [inputs, inp_lap, decoder_inputs, last_known_value_input],
        outputs,  # Use only `outputs` for training
    )
    if use_graph_layer:
        # Store attention as a submodel
        model.attention_submodel = keras.Model(
            [inputs, inp_lap, decoder_inputs, last_known_value_input],
            [gc_1_attn, gc_2_attn],  # Include attention coefficients here
        )

    return model

def analyze_feature_importance(model, max_lag=6, return_values_only=False):
    """
    Extracts and analyzes feature importance from the FeatureGate layer.
    """
    # Find the FeatureGate layer in the model
    feature_gate_layer = None
    for layer in model.layers:
        if isinstance(layer, (ImprovedFeatureGate, SoftFeatureGate)):
            feature_gate_layer = layer
            break
            
    # If nested, search through sub-models
    if feature_gate_layer is None:
        for layer in model.layers:
            if hasattr(layer, 'layers'):
                for sublayer in layer.layers:
                    if isinstance(sublayer, (ImprovedFeatureGate, SoftFeatureGate)):
                        feature_gate_layer = sublayer
                        break
    
    if feature_gate_layer is None:
        if not return_values_only:
            print("Could not find FeatureGate layer in the model!")
        return None
        
    # Get feature importance values
    importance = feature_gate_layer.get_feature_importance()
    
    if return_values_only:
        return importance
    
    # Get the indices of selected features if applicable
    selected_features = None
    k_percent = getattr(feature_gate_layer, 'k_percent', 1.0)
    
    if hasattr(feature_gate_layer, 'get_selected_features'):
        selected_features = feature_gate_layer.get_selected_features()
    
    # Get column names from the original data (before lagging)
    original_columns = list(data.columns)
    
    # Generate lagged feature names
    lagged_feature_names = []
    for col in original_columns:
        for lag in range(0, max_lag+1):
            lagged_feature_names.append(f"{col} ({-lag})")
            
    # Print the top features
    feature_indices = np.argsort(importance)[::-1]  # Descending order
    
    if not return_values_only:
        print("Feature Importance Analysis:")
        print("-" * 60)
        print(f"{'Feature Index':<10} {'Feature Name':<30} {'Importance':<10} {'Selected':<10}")
        print("-" * 60)
        
        for idx in feature_indices[:20]:  # Show top 20
            is_selected = "Yes" if selected_features is not None and idx in selected_features else "N/A"
            feature_name = lagged_feature_names[idx] if idx < len(lagged_feature_names) else f"Feature {idx}"
            print(f"{idx:<10} {feature_name:<30} {importance[idx]:.6f} {is_selected:<10}")
            
        if selected_features is not None:
            print(f"\nKeeping top {k_percent*100:.1f}% of features ({len(selected_features)}/{len(importance)})")
    
    return importance

'''
##################################################
# Main codes
##################################################
'''

if __name__ == '__main__':

    # Define the base path and file pattern
    base_path = [
                  # 'C:/Users/sarab/Desktop/Valley Fever/Data/Aggregated Data/Maricopa/Maricopa - For Analysis.csv'
                "D:/Shared/desktop 3/Valley Fever/Weather Data MARICOPA (MMWR Weeks)/Clean Data(Maricopa)/Clean Processed Data2.csv"
                ]

    # Loop over the numbers you're interested in
    for model_index, file_path in enumerate(base_path):  # Adjust the range as needed

        # Extract the model name from the file path
        model_name = file_path.split('/')[-1].split('.')[0]

        # for ETD
        data = pd.read_csv(file_path)
        data.drop('Date', axis=1, inplace=True)  # axis=1 specifies that you are dropping a column

        # for horizon in [96, 192, 336, 720]:
        for horizon in [8]:
                model_index = model_name + '_'+ str(horizon)

                # Data preprocessing parameters
                model_weights_update_iter = int(round(0.2 * len(data)))+1 # I did not want to have dynamic retraining for simplicity
                sequence_length = 3*horizon
                stride = 1
                dynamic_test_size = math.ceil(horizon / stride)
                use_graph_layer = True
                cal_all_adj = False
                target_as_feature = True
                differenced_target = True
                differenced_X = True
                diff_order=1
                normalized_data = True
                reverse_normalization=True
                initial_training_normalization= True
                repeat_corr = True
                Y_sequence = True
                single_step = False
                moving_average=False
                MA_window_size=12
                batch_size = 10
                feature_gate_k_percent = 0.3  # Keep only top 30% of features (70% reduction - more aggressive than 50%)
                max_lag = 6  # Reduced from 12 to 6 for fewer lag features

                # Set Target Name
                target_name_original = ['MARICOPA']  # Example: change to your actual column name
                target_name = target_name_original

                new_row = []
                # Initialize a list to store new rows
                buffered_rows = []
                save_to_excel_iter = 10

                head_size = 16
                num_heads = 8
                ff_dim = 256
                mlp_units = [256]
                dropout = 0.05
                mlp_dropout = 0.05

                '''#################### create an output file in the current directory ########################'''
                # Generate a string representing the current date and time
                now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                directory = "results_{}".format(now)
                if not os.path.exists(directory):
                    os.makedirs(directory)

                # Create the output file name with the current date and time
                output_filename = directory + '/dynamic_results_' + now +str(model_index) + '.xlsx'

                if os.path.exists(output_filename) == False:
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    # make Header
                    new_row = ['iteration',
                               'prediction_date',
                               'actual_value',
                               'forecast_value',
                               'Test_MAPE',
                               'Test_MAE',
                               'Test_MSE',
                               'Test_CORR',
                               'Test_RSE']

                    sheet.append(new_row)
                    wb.save(output_filename)

                start_index = 0
                train_end = 850
                test_start_index = 900
                test_end_index = 1002

                test_range = range(test_start_index, test_end_index)
                initial_train_data = data.loc[start_index:train_end]

                df_norm=data.copy()
                Y_scaler=[]

                if initial_training_normalization ==True and  normalized_data == True:
                    # apply normalization on initial data
                    scaler = StandardScaler()  # scaler applied to the predictors
                    # normalize data based on initial training set
                    scaler.fit(initial_train_data)
                    # keep the scaler attributes to reverse the transform later
                    scaler2 = StandardScaler()
                    Y_scaler = scaler2.fit(initial_train_data[target_name])

                    df_norm = scaler.transform(data)

                    # Wrap back in a DataFrame
                    df_norm = pd.DataFrame(
                        df_norm,
                        columns=data.columns,
                        index=data.index
                    )

                if normalized_data == False:
                    scaler2 = StandardScaler()
                    Y_scaler = scaler2.fit(initial_train_data[target_name])

                forecast_list = []
                actual_list = []

                iter = 0
                save_instance = 0
                ignore_first_instance_stride = False
                ignore_weights = True
                save_x_seq = []
                save_y_seq = []
                save_correlation_seq = []
                save_decoder_y_seq = []
                save_last_known_values = []
                run_first_time = True
                iter = horizon - 1

                if ignore_weights==False:
                    iter = 2619

                # ========================================================================================
                # STABILITY ANALYSIS SECTION
                # ========================================================================================
                
                print("\n🚀 Starting Feature Importance Stability Analysis...")
                
                # Define function for model training (used in stability testing)
                def train_model_for_stability(seed):
                    """
                    Train a model for stability testing using the SAME normalization approach as main code
                    Uses a SUBSTANTIAL training dataset, not just a tiny test window
                    """
                    # Set the seed
                    tf.random.set_seed(seed)
                    os.environ['PYTHONHASHSEED'] = str(seed)
                    np.random.seed(seed)
                    random.seed(seed)
                    
                    # IMPORTANT: Use a substantial training dataset for stability testing
                    # Use data from start to test_start_index (same as initial training range)
                    # This gives us ~900 data points instead of just 8!
                    end_index_stability = test_start_index - 1  # Use substantial training data
                    
                    print(f"    Training stability model with seed {seed} on data range [{start_index}:{end_index_stability}] ({end_index_stability-start_index+1} data points)")
                    
                    # Use the SAME normalization approach as main code
                    # Since initial_training_normalization=True, we use the pre-computed df_norm
                    df_norm_local = df_norm.loc[start_index:end_index_stability]
                    Y_scaler_local = Y_scaler
                    scaler2_local = scaler2

                    data_norm_local, data_diff_local = data_preprocess(df_norm_local, diff_order, start_index, end_index_stability, moving_average, MA_window_size)
                    
                    # Generate lagged variables
                    data_norm_local, data_diff_local, target_comp_updated_list_local, target_name_local = data_set_generation(
                        data_norm_local, data_diff_local, max_lag=max_lag, target_as_feature=target_as_feature, target_name=target_name_original)

                    # Preprocessing for stability test
                    save_instance_local = 0
                    save_x_seq_local = []
                    save_y_seq_local = []
                    save_correlation_seq_local = []
                    save_decoder_y_seq_local = []
                    save_last_known_values_local = []
                    
                    save_instance_local, save_x_seq_local, save_y_seq_local, save_decoder_y_seq_local, save_correlation_seq_local, save_last_known_values_local = preprocessing(
                        data_=data_norm_local, data_diff_=data_diff_local, diff_order=diff_order, sequence_length=sequence_length,
                        horizon=horizon, stride=stride, use_graph_layer=use_graph_layer, save_instance=save_instance_local,
                        ignore_first_instance_stride=False, save_x_seq=save_x_seq_local, save_y_seq=save_y_seq_local,
                        save_correlation_seq=save_correlation_seq_local, save_decoder_y_seq=save_decoder_y_seq_local,
                        save_last_known_values=save_last_known_values_local, target_name=target_name_local,
                        target_as_feature=target_as_feature, target_comp_updated_list=target_comp_updated_list_local,
                        differenced_target=differenced_target, differenced_X=differenced_X, moving_average=moving_average, MA_window_size=MA_window_size
                    )

                    print(f"    Generated {len(save_x_seq_local)} sequences for stability training")

                    # For stability testing, use ALL sequences for training (no test split)
                    # We're not evaluating model performance, just learning feature importance
                    X_train_local = np.asarray(save_x_seq_local)
                    Y_train_local = np.asarray(save_y_seq_local)
                    decoder_Y_train_local = np.asarray(save_decoder_y_seq_local)
                    last_known_values_train_local = np.asarray(save_last_known_values_local)
                    
                    if repeat_corr == True:
                        correlation_train_local = np.array(save_correlation_seq_local)
                    else:
                        correlation_train_local = np.asarray(save_correlation_seq_local)

                    # Build model
                    input_shape_local = X_train_local.shape[1:]
                    correlation_shape_local = correlation_train_local.shape[1:]
                    
                    model_local = build_model(input_shape_local, correlation_shape_local, use_graph_layer=use_graph_layer)
                    model_local.compile(loss="mse", optimizer=tf.keras.optimizers.Adam(learning_rate=1e-4), metrics=[tf.keras.metrics.MeanSquaredError()])

                    # Prepare training data with validation split
                    validation_ratio = 0.2  # Use smaller validation for stability test
                    total_data_size = len(X_train_local)
                    validation_size = int(total_data_size * validation_ratio)
                    validation_indices = np.arange(total_data_size - validation_size, total_data_size)
                    validation_mask = np.zeros(total_data_size, dtype=bool)
                    validation_mask[validation_indices] = True
                    training_mask = ~validation_mask

                    X_train_train_local = X_train_local[training_mask]
                    correlation_train_train_local = np.squeeze(correlation_train_local, axis=0) if correlation_train_local.ndim > 1 else []
                    decoder_Y_train_train_local = decoder_Y_train_local[training_mask]
                    last_known_values_train_train_local = last_known_values_train_local[training_mask]
                    Y_train_train_local = Y_train_local[training_mask]

                    X_train_valid_local = X_train_local[validation_mask]
                    decoder_Y_train_valid_local = decoder_Y_train_local[validation_mask]
                    last_known_values_train_valid_local = last_known_values_train_local[validation_mask]
                    Y_train_valid_local = Y_train_local[validation_mask]

                    train_gen_local = data_generator(X_train_train_local, correlation_train_train_local, decoder_Y_train_train_local, 
                                                   Y_train_train_local, last_known_values_train_train_local, batch_size=batch_size, new_data_ratio=0)
                    val_gen_local = data_generator(X_train_valid_local, correlation_train_train_local, decoder_Y_train_valid_local, 
                                                 Y_train_valid_local, last_known_values_train_valid_local, batch_size=batch_size, new_data_ratio=0)

                    train_steps_local = max(1, len(X_train_train_local) // batch_size)
                    val_steps_local = max(1, len(X_train_valid_local) // batch_size)

                    # Train model with fewer epochs for stability test (focus on feature learning)
                    callbacks_local = [keras.callbacks.EarlyStopping(patience=10, min_delta=0.001, monitor='val_mean_squared_error', mode='auto', restore_best_weights=True)]
                    
                    print(f"    Training with {len(X_train_train_local)} training samples, {len(X_train_valid_local)} validation samples")
                    
                    history_local = model_local.fit(train_gen_local, steps_per_epoch=train_steps_local, validation_data=val_gen_local, 
                                                  validation_steps=val_steps_local, epochs=30, callbacks=callbacks_local, verbose=0)  # Reduced epochs for faster testing

                    print(f"    ✅ Completed training for seed {seed}")
                    return model_local

                # Generate lagged feature names for visualization
                original_columns = list(data.columns)
                lagged_feature_names = []
                for col in original_columns:
                    for lag in range(0, max_lag+1):
                        lagged_feature_names.append(f"{col} ({-lag})")

                # Run stability analysis with multiple seeds
                seeds_to_test = [33, 42, 123, 456, 789]
                
                try:
                    importance_df, stability_results = run_stability_analysis(
                        model_training_function=train_model_for_stability,
                        seeds_to_test=seeds_to_test,
                        lagged_feature_names=lagged_feature_names,
                        save_directory=directory
                    )
                    
                    print("✅ Feature importance stability analysis completed!")
                    
                except Exception as e:
                    print(f"❌ Error during stability analysis: {str(e)}")
                    print("Continuing with main forecasting loop...")

                # ========================================================================================
                # MAIN FORECASTING LOOP (YOUR ORIGINAL CODE)
                # ========================================================================================

                print(f"\n🔄 Starting main forecasting loop...")

                for end_index in test_range[iter:]:

                    print("##################### run # {} / end index: {} ########################### ".format(iter, end_index))
                    SEED = 33
                    tf.random.set_seed(SEED)
                    os.environ['PYTHONHASHSEED'] = str(SEED)
                    np.random.seed(SEED)
                    random.seed(SEED)

                    session_conf = tf.compat.v1.ConfigProto(
                        intra_op_parallelism_threads=1,
                        inter_op_parallelism_threads=1
                    )
                    sess = tf.compat.v1.Session(
                        graph=tf.compat.v1.get_default_graph(),
                        config=session_conf
                    )
                    tf.compat.v1.keras.backend.set_session(sess)

                    # Data preprocessing
                    if initial_training_normalization==False and  normalized_data == True:
                        df = data.loc[start_index:end_index]
                        # apply normalization on initial data
                        scaler = StandardScaler()
                        # normalize data based on dynamic training set
                        scaler.fit(df[:-horizon])
                        df_norm_array = scaler.transform(df)

                        # Wrap back in a DataFrame
                        df_norm = pd.DataFrame(
                            df_norm_array,
                            columns=df.columns,
                            index=df.index
                        )

                        # keep the scaler attributes to reverse the transform later
                        scaler2 = StandardScaler()
                        Y_scaler = scaler2.fit(df[target_name_original][:-horizon])

                    data_norm, data_diff= data_preprocess(df_norm,diff_order, start_index, end_index,moving_average,MA_window_size)

                    # generate lagged variables
                    data_norm, data_diff, target_comp_updated_list, target_name = data_set_generation(data_norm,
                                                                                                 data_diff,
                                                                                                 max_lag=max_lag,
                                                                                                 target_as_feature=target_as_feature,
                                                                                                 target_name=target_name_original)

                    save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values = preprocessing(
                        data_=data_norm,
                        data_diff_=data_diff,
                        diff_order=diff_order,
                        sequence_length=sequence_length,
                        horizon=horizon, stride=stride,
                        use_graph_layer=use_graph_layer,
                        save_instance=save_instance,
                        ignore_first_instance_stride=ignore_first_instance_stride,
                        save_x_seq=save_x_seq,
                        save_y_seq=save_y_seq,
                        save_correlation_seq=save_correlation_seq,
                        save_decoder_y_seq=save_decoder_y_seq,
                        save_last_known_values=save_last_known_values,
                        target_name=target_name,
                        target_as_feature=target_as_feature,
                        target_comp_updated_list=target_comp_updated_list,
                        differenced_target=differenced_target,
                        differenced_X=differenced_X,
                        moving_average=moving_average,
                        MA_window_size=MA_window_size
                    )

                    del data_diff
                    # Test data split
                    if repeat_corr == True:
                        X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test,last_known_values_train,last_known_values_test = \
                            train_test_split(np.asarray(save_x_seq),
                                             np.asarray(save_y_seq),
                                             np.asarray(save_decoder_y_seq),
                                             np.asarray(save_last_known_values),
                                             test_size=dynamic_test_size,
                                             shuffle=False,
                                             random_state=1004)  # random_state ignored because shuffle = False

                        correlation_train = np.array(save_correlation_seq)
                        correlation_test = correlation_train

                    else:
                        X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test, last_known_values_train,last_known_values_test, correlation_train, correlation_test = \
                            train_test_split(np.asarray(save_x_seq),
                                             np.asarray(save_y_seq),
                                             np.asarray(save_decoder_y_seq),
                                             np.asarray(save_last_known_values),
                                             np.asarray(save_correlation_seq),
                                             test_size=dynamic_test_size,
                                             shuffle=False,
                                             random_state=1004)  # random_state ignored because shuffle = False

                    ignore_first_instance_stride = True

                    if iter % model_weights_update_iter == 0 or run_first_time==True:
                        run_first_time=False
                        input_shape = X_train.shape[1:]
                        correlation_shape = correlation_train.shape[1:]

                        model = build_model(
                            input_shape,
                            correlation_shape,
                            use_graph_layer=use_graph_layer
                        )

                        model.compile(
                            loss="mse",
                            optimizer=tf.keras.optimizers.Adam(learning_rate=1e-4),
                            metrics=[tf.keras.metrics.MeanSquaredError()],
                        )

                        model.summary()

                        callbacks = [
                            keras.callbacks.EarlyStopping(patience=20, min_delta=0.001, monitor='val_mean_squared_error',
                                                          mode='auto',
                                                          restore_best_weights=True)]

                        # Define the size of the validation set
                        validation_ratio = 0.25  # 20% or 25% of the data for validation
                        total_data_size = len(X_train)
                        validation_size = int(total_data_size * validation_ratio)

                        # Select the last validation_size indices for the validation set
                        validation_indices = np.arange(total_data_size - validation_size, total_data_size)

                        # Create masks for selecting training and validation data
                        validation_mask = np.zeros(total_data_size, dtype=bool)
                        validation_mask[validation_indices] = True
                        training_mask = ~validation_mask

                        # Split the data into training and validation sets using the masks
                        X_train_train = X_train[training_mask]
                        if correlation_train.ndim > 1:
                            correlation_train_train = np.squeeze(correlation_train,axis=0)
                        else:
                            correlation_train_train=[]
                        decoder_Y_train_train = decoder_Y_train[training_mask]
                        last_known_values_train_train= last_known_values_train[training_mask]
                        Y_train_train = Y_train[training_mask]

                        X_train_valid = X_train[validation_mask]
                        if correlation_train.ndim > 1:
                            correlation_train_valid = np.squeeze(correlation_train,axis=0)
                        else:
                            correlation_train_valid=[]
                        decoder_Y_train_valid = decoder_Y_train[validation_mask]
                        last_known_values_train_valid = last_known_values_train[validation_mask]
                        Y_train_valid = Y_train[validation_mask]

                        train_gen = data_generator(X_train_train, correlation_train_train, decoder_Y_train_train, Y_train_train, last_known_values_train_train, batch_size=batch_size, new_data_ratio=0)

                        # Create the validation generator
                        val_gen = data_generator(X_train_valid, correlation_train_train, decoder_Y_train_valid, Y_train_valid, last_known_values_train_valid,
                                                   batch_size=batch_size, new_data_ratio=0)

                        # Calculate the steps per epoch for training and validation
                        train_steps = len(X_train_train) // batch_size
                        val_steps = len(X_train_valid) // batch_size

                        if ignore_weights == True:
                            history = model.fit(
                                train_gen,
                                steps_per_epoch=train_steps,
                                validation_data=val_gen,
                                validation_steps=val_steps,
                                epochs=100,
                                callbacks=callbacks
                            )
                            ignore_weights = False

                        else:
                            model.load_weights('model_weights_{}.h5'.format(model_index))

                        # Saving weights as per your existing code
                        model.save_weights('model_weights_{}.h5'.format(model_index))
                        
                        # Analyze feature importance after training
                        if use_graph_layer:
                            feature_importance = analyze_feature_importance(model, max_lag=max_lag)
                            # Save feature importance for future analysis if needed
                            if feature_importance is not None:
                                np.save(f'{directory}/feature_importance_{model_index}_iter_{iter}.npy', feature_importance)

                    # Predict from last sequence
                    actual, y_pred_rev, datetime_index, y_pred = final_prediction(x_seq=save_x_seq,
                                                                                  Y_scaler=Y_scaler,
                                                                                  scaler2= scaler2,
                                                                                  decoder_Y_seq=save_decoder_y_seq,
                                                                                  last_known_values=save_last_known_values,
                                                                                  correlation_seq=save_correlation_seq,
                                                                                  sequence_length=sequence_length,
                                                                                  horizon=horizon,
                                                                                  model= model,
                                                                                  data=data_norm,
                                                                                  stride=stride,
                                                                                  target_name=target_name,
                                                                                  differenced_target=differenced_target,
                                                                                  diff_order=diff_order,
                                                                                  save_dict=directory)

                    if single_step==True:
                        # Save prediction in each loop
                        forecast_list = [(np.expand_dims(y_pred_rev[-1, :][-1],axis=0), datetime_index[-1][-1])]
                        actual_list = [(actual[-1, :][-1], datetime_index[-1][-1])]

                    else:
                        #Multi-step Forecast
                        # Save prediction in each loop
                        forecast_list = [(y_pred_rev[-1, :], datetime_index[-1])]
                        actual_list = [(actual[-1, :], datetime_index[-1])]

                    # Flatten the 3D arrays to 1D arrays
                    flattened_actual = np.concatenate([t[0] for t in actual_list]).flatten()
                    flattened_forecast = np.concatenate([t[0] for t in forecast_list]).flatten()

                    Test_CORR = CORR(flattened_actual, flattened_forecast)
                    Test_RSE = RSE(flattened_actual, flattened_forecast)
                    Test_MAPE = mean_absolute_percentage_error(flattened_actual, flattened_forecast)
                    Test_MAE = mean_absolute_error(flattened_actual, flattened_forecast)
                    Test_MSE = mean_squared_error(flattened_actual, flattened_forecast)

                    if reverse_normalization==True and normalized_data==True:
                      plot_prediction_graph(original_y=data.loc[end_index-horizon:end_index][target_name_original], forecast_list=forecast_list,
                                            test_boundary=test_start_index, horizon=horizon,
                                            Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    elif normalized_data== False:
                      transformed_data = scaler2.transform(data.loc[end_index-horizon:end_index][target_name_original])
                      # Convert the transformed data back into a pandas DataFrame or Series with the original index
                      original_y = pd.Series(transformed_data.squeeze(), index=data.loc[end_index-horizon:end_index].index)

                      plot_prediction_graph(original_y=original_y, forecast_list=forecast_list,
                                            test_boundary=test_start_index, horizon=horizon,
                                            Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    else:
                      # For ETD datasets we used (data_norm[target_name] instead of data[target_name_original]) It seems papers used normalized data for prediction results
                      plot_prediction_graph(original_y=data_norm[target_name], forecast_list=forecast_list,
                                          test_boundary=test_start_index,horizon=horizon,
                                          Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    '''##################### Exporting and Saving  ########################'''

                    if single_step==True and Y_sequence==True:
                        #Single-step Forecast
                        new_row = [iter,
                                   test_range[iter],
                                   str(actual[-1][-1]),
                                   str(y_pred_rev[-1][-1]),
                                   Test_MAPE,
                                   Test_MAE,
                                   Test_MSE,
                                   Test_CORR,
                                   Test_RSE]
                    else:
                        new_row = [iter,
                                   test_range[iter],
                                   str(actual[-1]),
                                   str(y_pred_rev[-1]),
                                   Test_MAPE,
                                   Test_MAE,
                                   Test_MSE,
                                   Test_CORR,
                                   Test_RSE]

                    buffered_rows.append(new_row)

                    if iter % save_to_excel_iter == 0:
                        wb = openpyxl.load_workbook(filename=output_filename)
                        sheet = wb.active
                        # Append the buffered rows to the sheet
                        for row in buffered_rows:
                            sheet.append(row)

                        # Save the Excel file
                        wb.save(output_filename)

                        # Clear the buffer
                        buffered_rows = []

                    iter = iter + 1

                    # Clear the session and invoke garbage collection
                    clear_session()
                    gc.collect()

                # After the loop, append any remaining buffered rows
                if buffered_rows:
                    wb = openpyxl.load_workbook(filename=output_filename)
                    sheet = wb.active
                    for row in buffered_rows:
                        sheet.append(row)
                    wb.save(output_filename)

                print(f"\n✅ Forecasting completed for horizon {horizon}!")
                print(f"📊 Results saved to: {output_filename}")
                if 'stability_results' in locals():
                    print(f"🔍 Feature stability analysis saved to: {directory}/")

        print(f"\n🎉 All analysis completed for model: {model_name}")
        print(f"📂 Check results in directory: {directory}/")
        
        # Print final summary
        print("\n" + "="*80)
        print("📋 ANALYSIS SUMMARY")
        print("="*80)
        print(f"🎯 Target Variable: {target_name_original}")
        print(f"🔮 Forecast Horizon: {horizon} steps")
        print(f"⏰ Max Lag Features: {max_lag} time steps")
        print(f"🧠 Feature Selection: Top {feature_gate_k_percent*100:.0f}% of features ({100-feature_gate_k_percent*100:.0f}% reduction)")
        print(f"📈 Graph Neural Network: {'Enabled' if use_graph_layer else 'Disabled'}")
        print(f"🔄 Transformer Architecture: Enabled")
        print(f"📊 Initial Training Normalization: {'Enabled' if initial_training_normalization else 'Disabled'}")
        print(f"🎲 Stability Testing: {'Completed' if 'stability_results' in locals() else 'Skipped'}")
        print("="*80)
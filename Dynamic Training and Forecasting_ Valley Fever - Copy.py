#
# import pandas as pd
# from statsmodels.tsa.stattools import adfuller
# import matplotlib.pyplot as plt
#
# # Load the data
# file_path = "C:/Users/sarab/Desktop/Valley Fever/Data/Aggregated Data/Maricopa/Maricopa - For Analysis.csv"
# data = pd.read_csv(file_path)
#
# # Ensure that the 'Date' column is treated as a string
# data['Date'] = data['Date'].astype(str)
#
# # Extract the 'Valley Fever Case Data' column
# valley_fever_data = data['Valley Fever Case Data']
#
# # Perform the Augmented Dickey-Fuller test
# result = adfuller(valley_fever_data)
#
# print('ADF Statistic: %f' % result[0])
# print('p-value: %f' % result[1])
# print('Critical Values:')
# for key, value in result[4].items():
#     print('\t%s: %.3f' % (key, value))
#
# # Plot the time series data
# plt.figure(figsize=(10, 6))
# plt.plot(data['Date'], valley_fever_data)
# plt.xlabel('Date')
# plt.ylabel('Valley Fever Case Data')
# plt.title('Valley Fever Case Data over Time')
# plt.xticks(rotation=45)
# plt.grid(True)
# plt.show()
'''
import pandas as pd
from statsmodels.tsa.stattools import adfuller
import matplotlib.pyplot as plt

# Load the data
file_path = "C:/Users/sarab/Desktop/Valley Fever/Data/Aggregated Data/Maricopa/Maricopa - For Analysis.csv"
data = pd.read_csv(file_path)

# Ensure that the 'Date' column is treated as a string
data['Date'] = data['Date'].astype(str)

# Extract the 'Valley Fever Case Data' column
valley_fever_data = data['Valley Fever Case Data']

# Differencing the data
valley_fever_data_diff = valley_fever_data.diff().dropna()

# Perform the Augmented Dickey-Fuller test on the differenced data
result_diff = adfuller(valley_fever_data_diff)

print('ADF Statistic (Differenced): %f' % result_diff[0])
print('p-value (Differenced): %f' % result_diff[1])
print('Critical Values (Differenced):')
for key, value in result_diff[4].items():
    print('\t%s: %.3f' % (key, value))

# Plot the original and differenced time series data
plt.figure(figsize=(12, 8))

# Plot the original data
plt.subplot(2, 1, 1)
plt.plot(data['Date'], valley_fever_data, label='Original')
plt.xlabel('Date')
plt.ylabel('Valley Fever Case Data')
plt.title('Original Valley Fever Case Data over Time')
plt.xticks(rotation=45)
plt.grid(True)
plt.legend()

# Plot the differenced data
plt.subplot(2, 1, 2)
plt.plot(data['Date'].iloc[1:], valley_fever_data_diff, label='Differenced', color='orange')
plt.xlabel('Date')
plt.ylabel('Differenced Valley Fever Case Data')
plt.title('Differenced Valley Fever Case Data over Time')
plt.xticks(rotation=45)
plt.grid(True)
plt.legend()

plt.tight_layout()
plt.show()

'''







import os
GPU_ID="0"
os.environ["CUDA_VISIBLE_DEVICES"] = GPU_ID
import gc
from tensorflow.keras.backend import clear_session

print(os.environ['PATH'])
#
# from dlib import cuda
# #cuda.set_device(1)
from tensorflow.python.client import device_lib
print(device_lib.list_local_devices())
import os
import openpyxl
import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.model_selection import train_test_split
import keras.backend as K
import tensorflow as tf
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_percentage_error, mean_squared_error, mean_absolute_error
from spektral.layers import GCNConv, SAGPool, MinCutPool, GlobalAttentionPool, JustBalancePool, DiffPool
from spektral.utils import normalized_adjacency
from sklearn.metrics.pairwise import cosine_similarity


import random
import datetime
import math
#from tensorflow.keras.layers import LSTM, Attention, GlobalAveragePooling1D, Bidirectional
from joblib import Parallel, delayed
from keras.layers import Dense, Dropout, Embedding, LSTM, Bidirectional
from keras.layers import BatchNormalization
from tensorflow.keras.layers import Dropout, Input
from spektral.layers import GATConv
from keras.layers import Add, Concatenate
from tensorflow.keras.layers import Lambda
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras.regularizers import l2



def data_preprocess(df, diff_order, start_index, end_index,moving_average,MA_window_size):
    '''
    This function removes columns with more than 10% missing values and fills in the NaNs in the remaining columns.
    '''
    # start_date = pd.to_datetime(start_date)
    # end_date = pd.to_datetime(end_date)

    df = pd.DataFrame(df)
    df = df.loc[start_index:end_index]



    # Apply differencing according to the specified diff_order
    df_diff = df.copy()
    for _ in range(diff_order):


        if moving_average==True:
            # Assuming df_diff is your DataFrame


            # Function to pad the beginning of the DataFrame
            def pad_head(df, pad_width):
                head = pd.DataFrame([df.iloc[0]] * pad_width)
                return pd.concat([head, df]).reset_index(drop=True)

            # Apply padding only to the head
            pad_width = MA_window_size - 1
            df_padded = pad_head(df_diff, pad_width)

            # Compute the moving average on the padded DataFrame
            df_ma_padded = df_padded.rolling(window=MA_window_size, center=False).mean()

            # Now, remove the padding by trimming the extra rows from the start of the moving average DataFrame
            df_ma = df_ma_padded.iloc[pad_width:].reset_index(drop=True)

            # Shift the moving average DataFrame by one to align it for differencing
            df_ma_shifted = df_ma.shift(1)

            # Ensure df_diff is aligned in size with df_ma_shifted
            df_diff_aligned = df_diff.iloc[:df_ma_shifted.shape[0]]

            # Calculate the difference between the actual value at t+1 and the moving average at t
            df_diff = df_diff_aligned - df_ma_shifted
        else:
            df_diff = df_diff.diff(periods=1)


    # # Select the related range of data
    # df2 = df2.loc[start_index:end_index]
    # Adjust the index offset for the differencing order
    df_diff = df_diff.loc[start_index + diff_order:end_index]

    # # Find out the percentage of missing values in each column in the given dataset
    # percent_missing = df.isnull().mean()
    # missing_value_df = pd.DataFrame({'column_name': df.columns,
    #                                  'percent_missing': percent_missing})
    # df = df[df.columns[percent_missing < 0.05]]
    # df = df.fillna(method='ffill').fillna(method='bfill')
    #
    # # Find out the percentage of missing values in each column in the given dataset
    # percent_missing = df_diff.isnull().mean()
    # missing_value_df_diff = pd.DataFrame({'column_name': df_diff.columns,
    #                                       'percent_missing': percent_missing})
    # df_diff = df_diff[df_diff.columns[percent_missing < 0.05]]
    # df_diff = df_diff.fillna(method='ffill').fillna(method='bfill')

    return df, df_diff


def preprocessing(data_, data_diff_, diff_order, sequence_length, horizon, stride, use_graph_layer,
                  save_instance,ignore_first_instance_stride, save_x_seq, save_y_seq, save_correlation_seq,
                  save_decoder_y_seq, save_last_known_values, target_name,
                  target_as_feature, target_comp_updated_list, differenced_target, differenced_X, moving_average,MA_window_size):
    '''

    This function takes in a time series data, normalizes it, generates sequences, and splits the data into
    training and testing sets.

    data: Time series data
    sequence_length: Length of the sequence used in the model
    horizon: The number of steps to predict in the future
    stride: The step size used to generate sequences
    test_size: The proportion of the data used for testing
    scaler: Scaler used to normalize the data
    '''
    if differenced_X == True:
        X = data_diff_
    else:
        # first row of X is dropped because of differencing
        X = data_[diff_order:]

    X_corr = data_diff_

    if differenced_target == True:
        Y = data_diff_[target_name].values
    else:
        Y = data_[diff_order:][target_name].values

    Y_original = data_[diff_order:][target_name].values


    if target_as_feature == False:
        # Drop columns from X based on target list
        X = X.drop(target_comp_updated_list, axis=1)
        X_corr = X_corr.drop(target_comp_updated_list, axis=1)
    X = X.values

    # Sequence Generation Part
    x_seq = []
    y_seq = []
    correlation_seq = []
    decoder_y_seq = []
    n = np.shape(X)[1]
    # Create ones matrix of size (n, n)
    corr = np.ones((n, n))

    # def calculate_corr(j, k, df_window):
    #     return j, k, chatterjee(df_window[:, j], df_window[:, k])

    if use_graph_layer == True:
        if cal_all_adj == False:
            # restart save_correlation
            save_correlation_seq = []

            # calculate starting correlation based on all available data
            # results = Parallel(n_jobs=-1)(delayed(calculate_corr)(j, k, X[:len(X) - horizon]) for j in range(n) for k in range(n))

            # # Populate the correlation matrix
            # for j, k, result in results:
            #     corr[j][k] = result

            # we put absolute value because GCN does not accept negative adjacency
            corr = np.corrcoef(X_corr[:len(X_corr) - horizon], rowvar=False)
            # Apply the filter to keep only correlations with absolute values greater than 0.05
            # corr = np.where(np.abs(corr) > 0.05, corr, 0)
            # Normalize the corr to be in the range [0, 1]
            normalized_corr = (1 + corr) / 2
            correlation_seq.append(normalized_corr)

            # # Assuming X_corr is your data matrix and horizon is the number of steps ahead for forecasting
            # X_corr_subset = X_corr[:len(X_corr) - horizon]
            #
            # # Transpose the matrix to get (variables, instances)
            # X_corr_subset_transposed = X_corr_subset.T
            #
            # # Calculate cosine similarity between columns (variables)
            # cos_sim = cosine_similarity(X_corr_subset_transposed)
            #
            # # Normalize the cosine similarity to be in the range [0, 1]
            # normalized_cos_sim = (1 + cos_sim) / 2
            #
            #
            # # Append the result to correlation_seq
            # correlation_seq.append(normalized_cos_sim)



            # corr = np.corrcoef(X_corr[:len(X_corr) - horizon], rowvar=False)
            # corr = (corr + 1) / 2  # This shifts and scales the values to be between 0 and 1
            # correlation_seq.append(corr)

    for instance in range(save_instance, len(X) - sequence_length - horizon +1, stride):
        if ignore_first_instance_stride:
            ignore_first_instance_stride = False
            continue
        x_seq.append(X[instance: instance + sequence_length])

        if Y_sequence == True:
            y_seq.append(Y_original[instance + sequence_length:instance + sequence_length + horizon])
        else:
            y_seq.append(np.expand_dims(np.sum(Y_original[instance + sequence_length:instance + sequence_length + horizon], axis=0), axis=0))

        decoder_y_seq.append(Y[instance + sequence_length - horizon:instance + sequence_length])
        df_window = X[instance: instance + sequence_length]

        if use_graph_layer == True:
            if cal_all_adj == True:
                # Parallelize the calculation of correlations
                results = Parallel(n_jobs=-1)(
                    delayed(calculate_corr)(j, k, df_window) for j in range(n) for k in range(n))

                # Populate the correlation matrix
                for j, k, result in results:
                    corr[j][k] = result

                correlation_seq.append(corr)

        # Immediately after sequence generation, capture the last known value
        if differenced_target == True:

            if moving_average==True:
                ## Apply moving average on last_value_before_sequence
                last_value_before_sequence = data_[target_name].iloc[max(0,
                                                                         instance + sequence_length - diff_order + 2 - MA_window_size):instance + sequence_length - diff_order + 2].mean()
            else:
                # Adjust the index to capture the last known value correctly
                last_value_before_sequence = data_[target_name].iloc[instance + sequence_length - diff_order + 1]

            last_value_before_sequence = np.expand_dims(last_value_before_sequence, axis=0)
            save_last_known_values.append(last_value_before_sequence)
        else:
            save_last_known_values.append(1)

    del X
    save_instance = instance
    save_x_seq.extend(x_seq)
    save_y_seq.extend(y_seq)
    save_decoder_y_seq.extend(decoder_y_seq)
    save_correlation_seq.extend(correlation_seq)


    # last_known_values_input = [item[0] for item in last_known_values]  # Reshape as necessary
    return save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values


def plot_prediction_graph(original_y, forecast_list, test_boundary,
                          horizon, Test_KPI, iter, save_dict):
    plt.figure(figsize=(15, 10))
    # plt.axvline(x=test_boundary, ls=':', color='black')
    # plt.plot(test_index, actual_list[:len(forecast_list)], 'r--',marker='o')
    plt.plot(original_y.index[-horizon:], original_y[-horizon:], 'r--', marker='o')
    plt.plot(forecast_list[-1][1], forecast_list[-1][0], 'b', linewidth=0.6, marker='o')

    # Create a legend with custom labels
    actual_labels = ['actual_{}'.format(i) for i in range(len(target_name))]
    forecast_labels = ['forecast_{}'.format(i) for i in range(len(target_name))]
    legend_labels = ['test boundary'] + actual_labels + forecast_labels
    plt.legend(legend_labels)

    # plt.legend(['test boundary', 'actual', 'prediction'])
    plt.xlabel('Date')
    plt.ylabel('Value')
    plt.text(0.02, 0.8, 'KPI for Test Range = {:.4f}'.format(Test_KPI), bbox=dict(facecolor='red', alpha=0.2),
             transform=plt.gca().transAxes, va="top", ha="left")
    plt.title('{}- Ahead Model_it={}'.format(horizon, iter + 1), loc='center')

    plt.savefig("{}\prediction_{}_it={}.png".format(save_dict, horizon, iter + 1), dpi=300)
    plt.close()


def plot_norm_diff_graph(Y_seq, y_pred, datetime_index, test_boundary,
                    horizon, Test_KPI, iter, save_dict):
    plt.figure(figsize=(15, 10))
    plt.axvline(x=test_boundary, ls=':', color='black')

    for i in range(-1, 0):
        plt.plot(datetime_index[i], Y_seq[i], 'r--', marker='o')
        plt.plot(datetime_index[i], y_pred[i], 'b', linewidth=0.6, marker='o')

    # Create a legend with custom labels
    actual_labels = ['actual_{}'.format(i) for i in range(len(target_name))]
    forecast_labels = ['forecast_{}'.format(i) for i in range(len(target_name))]
    legend_labels = ['test boundary'] + actual_labels + forecast_labels
    plt.legend(legend_labels)

    plt.xlabel('Date')
    plt.ylabel('Value')
    plt.text(0.02, 0.8, 'KPI for Test Range = {:.4f}'.format(Test_KPI), bbox=dict(facecolor='red', alpha=0.2),
             transform=plt.gca().transAxes, va="top", ha="left")
    plt.title('norm_diff_{}- Ahead Model_it={}'.format(horizon, iter + 1), loc='center')

    plt.savefig("{}\dnorm_diff_{}_it={}.png".format(save_dict, horizon, iter + 1), dpi=300)
    plt.close()



class ReverseDifferencingLayer(tf.keras.layers.Layer):
    def __init__(self, **kwargs):
        super(ReverseDifferencingLayer, self).__init__(**kwargs)

    def call(self, inputs):
        """
        Perform the reversal of differencing operation.

        Args:
        inputs: A list of two tensors, where
                inputs[0] is the tensor of predicted differenced values of shape (batch_size, sequence_length, len(target)),
                and inputs[1] is the tensor of the last known value of the series of shape (batch_size, 1, len(target)).

        Returns:
        A tensor of shape (batch_size, sequence_length, len(target)) containing the predictions in the original scale.
        """
        preds, last_known_value = inputs
        # Use tf.cumsum to reverse the differencing, adding the last known value as the base.
        # TensorFlow handles the broadcasting automatically.
        reversed_preds = tf.cumsum(preds, axis=1) + last_known_value
        return reversed_preds


def final_prediction(x_seq,Y_scaler, scaler2, decoder_Y_seq,last_known_values, correlation_seq, sequence_length, horizon, model, data, stride,
                     target_name, differenced_target, diff_order):

    x_seq = np.expand_dims(x_seq[-1], axis=0)
    decoder_Y_seq = np.expand_dims(decoder_Y_seq[-1], axis=0)
    last_known_values = np.expand_dims(last_known_values[-1], axis=0)

    if len(correlation_seq) > 0:
        correlation_seq = np.expand_dims(correlation_seq[-1], axis=0)


    # Prediction
    y_pred = model.predict([x_seq, correlation_seq, decoder_Y_seq, last_known_values])

    y_pred_rev = y_pred


    '''
    # Note: I removed this reversal operation and relocated it to the final layer within the model.
    # This adjustment was made to gain more control over the model's performance.


    # if differenced_target and diff_order>=1:
    #     period_sum=y_pred.copy()
    #     # Reverse differencing based on the diff_order
    #     df_diff = data.copy()
    #     save = []
    #     save.append(data)
    #     for i in range(1, diff_order):
    #         df_diff = df_diff.diff(periods=1)
    #         # Add the differenced data to the save list
    #         save.append(df_diff)
    #     for i in range(diff_order):
    #         # Compute the cumulative sum for each row
    #         period_sum = period_sum.cumsum(axis=1)
    #
    #     # Create a boolean mask to select the columns that are in target_name
    #         column_mask = np.isin(data.columns, target_name)
    #         new_value = save[diff_order-1-i].values[1 + sequence_length - diff_order:-horizon:stride, column_mask][-1, None]
    #         period_sum = period_sum + new_value
    #         period_sum = np.insert(period_sum, 0, new_value, axis=1)
    #
    #     y_pred_rev=period_sum[:,diff_order:,:]
    # else:
    #     y_pred_rev = y_pred
    '''


    if normalized_data==True:

        if reverse_normalization==True:
            y_pred_rev = np.expand_dims(Y_scaler.inverse_transform(np.squeeze(y_pred_rev, axis=0)),
                                     axis=0)  # reverse scaler to plot later in the original scale
        else:
            # It seems papers used normalized data for prediction results
            y_pred_rev = np.expand_dims(np.squeeze(y_pred_rev, axis=0),
                                        axis=0)
    else:
        y_pred_rev = scaler2.transform(np.squeeze(y_pred_rev, axis=0))
        y_pred_rev = np.expand_dims(y_pred_rev,
                                axis=0) # reverse scaler to plot later in the original scale


    n_rows = data.shape[0] - sequence_length - horizon
    indices = np.arange(1 + sequence_length, 1 + n_rows + sequence_length, stride)


    if Y_sequence == True:

        if normalized_data == True:
            if reverse_normalization == True:
                actual_array = np.expand_dims(Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + np.arange(horizon)]), axis=0)
            else:
                # It seems papers used normalized data for prediction results
                actual_array = np.expand_dims(
                    data[target_name].values[indices[-1, None] + np.arange(horizon)], axis=0)
        else:
            actual_array = scaler2.transform(data[target_name].values[indices[-1, None] + np.arange(horizon)])
            actual_array = np.expand_dims(actual_array
                , axis=0)

        datetime_index = np.expand_dims(data.index.to_numpy()[indices[-1, None] + np.arange(horizon)], axis=0)
    else:
        if normalized_data == True:
            actual_array = Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + horizon - 1])
        else:
            actual_array = scaler2.transform(data[target_name].values[indices[-1, None] + horizon - 1])
        datetime_index = data.index.to_numpy()[indices[-1, None] + horizon - 1]


    return actual_array, y_pred_rev, datetime_index, y_pred


'''
# Other correlation coefficients that I did not use. I used simple correlation
def chatterjee(x, y, tie_method='average'):
    """
    Calculate the Chatterjee correlation coefficient between two arrays.

    Parameters:
        x (array-like): The first array of data.
        y (array-like): The second array of data.
        tie_method (str, optional): Method to handle ties when ranking data. Default is 'average'.

    Returns:
        float: The Chatterjee correlation coefficient.
    """
    df = pd.DataFrame({"X": x, "Y": y})
    n = len(df)

    df['rank_x'] = df['X'].rank(method=tie_method)
    df['rank_y'] = df['Y'].rank(method=tie_method)

    df.sort_values(by='rank_x', inplace=True)

    sum_term_num = df['rank_y'].sub(df['rank_y'].shift()).abs().sum()

    l_i = n - df['rank_y']
    sum_den = (l_i * (n - l_i)).sum()

    coefficient = (1 - n * sum_term_num / (2 * sum_den))

    return np.maximum(coefficient, 0)
'''


def data_set_generation(data, data_diff, max_lag, target_as_feature, target_name):
    # update the input_components list
    input_components = list(data.columns)

    # remove the target column before lagging the variables
    if not target_as_feature:
        for target in target_name:
            if target in input_components:
                input_components.remove(target)

    # List of updated input components.
    target_comp_updated_list = []

    # Initialize forecast dataset.
    new_data = pd.DataFrame()
    new_data_diff = pd.DataFrame()

    # Iterate through input components and their lags.
    for col in input_components:
        for lag in range(0, max_lag + 1):
            # Generate new component name.
            comp_new = f"{col} ({-lag})"
            if col in target_name:
                target_comp_updated_list.append(comp_new)
            # Get start index based on optimal time lag.
            new_data_time_series = data[col].shift(lag).rename(comp_new)
            new_data_diff_time_series = data_diff[col].shift(lag).rename(comp_new)

            # Add to dataset and forecast dataset
            new_data = pd.concat([new_data, new_data_time_series], axis=1)
            new_data_diff = pd.concat([new_data_diff, new_data_diff_time_series], axis=1)

    # Drop the first few rows to remove Nans.

    new_data = new_data.dropna()
    new_data_diff = new_data_diff.dropna()

    update_target_name = []
    if target_as_feature == True:
        for i, col in enumerate(target_name):
            update_target_name.append(f"{col} (0)")
        target_name = update_target_name
    else:
        # append the removed target column to datasets
        new_data[target_name] = data[target_name]
        new_data_diff[target_name] = data_diff[target_name]
        target_comp_updated_list.extend(target_name)

    # new_data = new_data.reset_index(drop=True)
    # new_data_diff = new_data_diff.reset_index(drop=True)

    return new_data, new_data_diff, target_comp_updated_list, target_name



def CORR(flattened_actual, flattened_forecast):
    """
    Calculate the Empirical Correlation Coefficient between two flattened arrays.

    Parameters:
    flattened_actual (array-like): Flattened array of actual values.
    flattened_forecast (array-like): Flattened array of forecasted values.

    Returns:
    float: Empirical Correlation Coefficient.
    """

    # Ensuring the inputs are numpy arrays
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    # Calculate the standard deviation and mean of the inputs
    sigma_p = flattened_forecast.std()
    sigma_g = flattened_actual.std()
    mean_p = flattened_forecast.mean()
    mean_g = flattened_actual.mean()

    # Calculate the correlation only if sigma_g and sigma_p are not equal to zero
    valid_indices = (sigma_g != 0) and (sigma_p != 0)
    if np.any(valid_indices):
        correlation = np.mean(
            ((flattened_forecast[valid_indices] - mean_p) * (flattened_actual[valid_indices] - mean_g)) / (sigma_p * sigma_g)
        )
    else:
        correlation = 0

    return correlation


def RSE(flattened_actual, flattened_forecast):
    """
    Calculate the Root Relative Squared Error between two flattened arrays or multi-dimensional arrays.

    Parameters:
    flattened_actual (array-like): Flattened array or multi-dimensional array of actual values.
    flattened_forecast (array-like): Flattened array or multi-dimensional array of forecasted values.

    Returns:
    float: Root Relative Squared Error.
    """

    # Ensuring the inputs are numpy arrays
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    # Calculate the mean of the actual values
    mean_actual = np.mean(flattened_actual)

    # Calculate the numerator (sum of squared differences between actual and forecast values)
    numerator = np.sum((flattened_actual - flattened_forecast) ** 2)

    # Calculate the denominator (sum of squared differences between actual values and mean of actual values)
    denominator = np.sum((flattened_actual - mean_actual) ** 2)

    # Calculate RSE
    if denominator != 0:
        rse = np.sqrt(numerator / denominator)
    else:
        rse = np.nan

    return rse

def mean_absolute_scaled_error(y_true, y_pred):
    naive_forecast = tf.roll(y_true, shift=1, axis=0)  # Shift the tensor by 1 position
    errors = tf.abs(y_true - y_pred)
    naive_errors = tf.abs(y_true - naive_forecast)
    epsilon = 1e-8  # Small constant to prevent division by zero
    mase = tf.reduce_mean(errors / (naive_errors + epsilon))
    return mase


def symmetric_mean_absolute_percentage_error(y_true, y_pred):
    y_true = K.maximum(K.epsilon(), y_true)  # Avoid division by zero
    return K.mean(2 * K.abs(y_pred - y_true) / (y_pred + y_true), axis=-1)


def symmetric_mean_absolute_percentage_error_loss(y_true, y_pred):
    epsilon = 1e-7  # A small constant to avoid division by zero
    y_true = K.maximum(epsilon, y_true)
    smape_loss = K.mean(2 * K.abs(y_pred - y_true) / (y_pred + y_true), axis=-1)
    return smape_loss








'''
##################################################
# Main codes
##################################################
'''

if __name__ == '__main__':

    # Define the base path and file pattern
    base_path = [
                  # 'C:/Users/sarab/Desktop/Valley Fever/Data/Aggregated Data/Maricopa/Maricopa - For Analysis.csv'
                "C:/Users/sarab/Desktop/desktop/Valley Fever/Data/Aggregated Case Data/Final Data/Number of Cases_Weekly_data.csv"
                ]

    # Loop over the numbers you're interested in
    for model_index, file_path in enumerate(base_path):  # Adjust the range as needed

        # Extract the model name from the file path
        model_name = file_path.split('/')[-1].split('.')[0]

        # for ETD
        data = pd.read_csv(file_path)
        data.drop('Date', axis=1, inplace=True)  # axis=1 specifies that you are dropping a column


        # for horizon in [96, 192, 336, 720]:
        for horizon in [24]:
                model_index = model_name + '_'+ str(horizon)

                # Data preprocessing parameters

                model_weights_update_iter = int(round(0.2 * len(data)))+1 # I did not want to have dynamic retraining for simplicity
                sequence_length = 2*horizon
                stride = 1
                dynamic_test_size = math.ceil(horizon / stride)
                use_graph_layer = True
                cal_all_adj = False
                target_as_feature = True
                differenced_target = True
                differenced_X = True
                diff_order=1
                normalized_data = True
                reverse_normalization=True
                initial_training_normalization= True
                repeat_corr = True
                Y_sequence = True
                single_step = False
                moving_average=False
                MA_window_size=12
                batch_size = 5



                new_column_names = [str(i) for i in range(len(data.columns))]
                data.columns = new_column_names

                # Set Target Name
                # target_name_original = list(new_column_names)
                target_name_original = ['0']
                target_name = target_name_original

                new_row = []
                # Initialize a list to store new rows
                buffered_rows = []
                save_to_excel_iter = 10


                head_size = 16
                num_heads = 8
                ff_dim = 128
                mlp_units = [128]
                dropout = 0.05
                mlp_dropout = 0.05

                '''#################### create an output file in the current directory ########################'''
                # Generate a string representing the current date and time
                now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                directory = "results_{}".format(now)
                if not os.path.exists(directory):
                    os.makedirs(directory)

                # Create the output file name with the current date and time
                output_filename = directory + '/dynamic_results_' + now +str(model_index) + '.xlsx'

                if os.path.exists(output_filename) == False:
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    # make Header
                    new_row = ['iteration',
                               'prediction_date',
                               'actual_value',
                               'forecast_value',
                               'Test_MAPE',
                               'Test_MAE',
                               'Test_MSE',
                               'Test_CORR',
                               'Test_RSE']

                    sheet.append(new_row)
                    wb.save(output_filename)



                start_index = 0
                train_end = 1380
                # val_end = 12 * 30 * 24 + 4 * 30 * 24
                # test_end = 12 * 30 * 24 + 4 * 30 * 24 + 4 * 30 * 24
                test_start_index = 1730 # val_end
                test_end_index = 1764



                # test_range = pd.date_range(start=test_start_date, end=test_end_date, freq='H')
                test_range = range(test_start_index, test_end_index)
                initial_train_data = data.loc[start_index:train_end]

                df_norm=data.copy()
                Y_scaler=[]

                if initial_training_normalization ==True and  normalized_data == True:

                    # apply normalization on initial data
                    scaler = StandardScaler()  # scaler applied to the predictors
                    # normalize data based on initial training set
                    scaler.fit(initial_train_data)
                    # keep the scaler attributes to reverse the transform later
                    scaler2 = StandardScaler()
                    Y_scaler = scaler2.fit(initial_train_data[target_name])

                    df_norm = scaler.transform(data)
                if normalized_data == False:
                    scaler2 = StandardScaler()
                    Y_scaler = scaler2.fit(initial_train_data[target_name])



                forecast_list = []

                actual_list = []

                iter = 0
                save_instance = 0
                ignore_first_instance_stride = False
                ignore_weights = True
                save_x_seq = []
                save_y_seq = []
                save_correlation_seq = []
                save_decoder_y_seq = []
                save_last_known_values = []
                run_first_time = True
                iter = horizon - 1


                if ignore_weights==False:
                    iter = 2619


                for end_index in test_range[iter:]:

                    print("##################### run # {} / end index: {} ########################### ".format(iter, end_index))
                    SEED = 33
                    tf.random.set_seed(SEED)
                    os.environ['PYTHONHASHSEED'] = str(SEED)
                    np.random.seed(SEED)
                    random.seed(SEED)

                    session_conf = tf.compat.v1.ConfigProto(
                        intra_op_parallelism_threads=1,
                        inter_op_parallelism_threads=1
                    )
                    sess = tf.compat.v1.Session(
                        graph=tf.compat.v1.get_default_graph(),
                        config=session_conf
                    )
                    tf.compat.v1.keras.backend.set_session(sess)




                    # Data preprocessing

                    if initial_training_normalization==False and  normalized_data == True:

                        df = data.loc[start_index:end_index]
                        # apply normalization on initial data
                        scaler = StandardScaler()
                        # normalize data based on dynamic training set
                        scaler.fit(df[:-horizon])
                        df_norm = scaler.transform(df)  # keep the normalized data to predict later

                        # keep the scaler attributes to reverse the transform later
                        scaler2 = StandardScaler()
                        Y_scaler = scaler2.fit(df[target_name_original][:-horizon])



                    data_norm, data_diff= data_preprocess(df_norm,diff_order, start_index, end_index,moving_average,MA_window_size)

                    # generate lagged variables
                    data_norm, data_diff, target_comp_updated_list, target_name = data_set_generation(data_norm,
                                                                                                 data_diff,
                                                                                                 max_lag=12,
                                                                                                 target_as_feature=target_as_feature,
                                                                                                 target_name=target_name_original)

                    save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values = preprocessing(
                        data_=data_norm,
                        data_diff_=data_diff,
                        diff_order=diff_order,
                        sequence_length=sequence_length,
                        horizon=horizon, stride=stride,
                        use_graph_layer=use_graph_layer,
                        save_instance=save_instance,
                        ignore_first_instance_stride=ignore_first_instance_stride,
                        save_x_seq=save_x_seq,
                        save_y_seq=save_y_seq,
                        save_correlation_seq=save_correlation_seq,
                        save_decoder_y_seq=save_decoder_y_seq,
                        save_last_known_values=save_last_known_values,
                        target_name=target_name,
                        target_as_feature=target_as_feature,
                        target_comp_updated_list=target_comp_updated_list,
                        differenced_target=differenced_target,
                        differenced_X=differenced_X,
                        moving_average=moving_average,
                        MA_window_size=MA_window_size

                    )




                    del data_diff
                    # Test data split
                    if repeat_corr == True:
                        X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test,last_known_values_train,last_known_values_test = \
                            train_test_split(np.asarray(save_x_seq),
                                             np.asarray(save_y_seq),
                                             np.asarray(save_decoder_y_seq),
                                             np.asarray(save_last_known_values),
                                             test_size=dynamic_test_size,
                                             shuffle=False,
                                             random_state=1004)  # random_state ignored because shuffle = False

                        correlation_train = np.array(save_correlation_seq)
                        correlation_test = correlation_train


                    else:
                        X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test, last_known_values_train,last_known_values_test, correlation_train, correlation_test = \
                            train_test_split(np.asarray(save_x_seq),
                                             np.asarray(save_y_seq),
                                             np.asarray(save_decoder_y_seq),
                                             np.asarray(save_last_known_values),
                                             np.asarray(save_correlation_seq),
                                             test_size=dynamic_test_size,
                                             shuffle=False,
                                             random_state=1004)  # random_state ignored because shuffle = False


                    ignore_first_instance_stride = True


                    def data_generator(X, correlation, decoder_Y, Y, last_known_values, batch_size, new_data_ratio=0.1):
                        while True:
                            # Number of new data samples to include in each batch
                            new_data_count = int(batch_size * new_data_ratio)

                            # Indices of the new data
                            new_data_indices = np.arange(len(X) - new_data_count, len(X))

                            for i in range(0, len(X), batch_size):
                                # Select random indices from the entire dataset
                                random_indices = np.random.permutation(len(X) - new_data_count)

                                # Combine new data indices with random indices and shuffle
                                combined_indices = np.concatenate([new_data_indices, random_indices[:batch_size - new_data_count]])
                                np.random.shuffle(combined_indices)

                                batch_X = X[combined_indices]
                                batch_decoder_Y = decoder_Y[combined_indices]
                                batch_last_known_values = last_known_values[combined_indices]
                                batch_Y = Y[combined_indices]

                                if  len(correlation) > 0:
                                    if repeat_corr == True:
                                        batch_correlation = np.repeat(correlation[np.newaxis, ...], len(batch_X), axis=0)
                                    else:
                                        batch_correlation = correlation[combined_indices]
                                else:
                                    batch_correlation=[]

                                yield [batch_X, batch_correlation, batch_decoder_Y, batch_last_known_values], batch_Y

                    def graph_processing_block(inputs, inp_lap, head_size, num_heads, dropout, horizon):

                        l2_reg = 2.5e-4  # L2 regularization rate
                        x = tf.transpose(inputs, perm=[0, 2, 1])

                        x_out2, lap_out = MinCutPool(int(x.shape[1] // 2), return_selection=False,
                                                     name='mincut_pool_1')([x, inp_lap])

                        do_1 = Dropout(dropout)(x_out2)
                        gc_1 = GATConv(
                            int(horizon // 2),
                            attn_heads=num_heads,
                            concat_heads=False,
                            dropout_rate=dropout,
                            activation="elu",
                            kernel_regularizer=l2(l2_reg),
                            attn_kernel_regularizer=l2(l2_reg),
                            bias_regularizer=l2(l2_reg),
                        )([do_1, lap_out])

                        do_2 = Dropout(dropout)(gc_1)
                        gc_2 = GATConv(
                            int(horizon // 2),
                            attn_heads=num_heads,
                            concat_heads=False,
                            dropout_rate=dropout,
                            activation="softmax",
                            kernel_regularizer=l2(l2_reg),
                            attn_kernel_regularizer=l2(l2_reg),
                            bias_regularizer=l2(l2_reg),
                        )([do_2, lap_out])

                        graph_output = tf.transpose(gc_2, perm=[0, 2, 1])

                        return graph_output


                    def lstm_encoder_decoder_block(graph_output, decoder_inputs, latent_dim, horizon, target_name, dropout):
                        # One-directional LSTM

                        l2_reg = 2.5e-4  # L2 regularization rate

                        encoder_lstm_1 = LSTM(latent_dim, kernel_regularizer=tf.keras.regularizers.l2(l2_reg), return_sequences=True)(graph_output)
                        encoder_drop_1 = Dropout(0.1)(encoder_lstm_1)

                                                # Final LSTM layer of the encoder
                        encoder_outputs, forward_h, forward_c = LSTM(int(latent_dim // 2), kernel_regularizer=tf.keras.regularizers.l2(l2_reg), return_state=True)(
                            encoder_drop_1)

                        initial_state = [forward_h, forward_c]


                        # Decoder
                        decoder_lstm_1 = LSTM(int(latent_dim // 2), kernel_regularizer=tf.keras.regularizers.l2(l2_reg), return_sequences=True)(decoder_inputs,
                                                                                           initial_state=initial_state)
                        decoder_drop_1 = Dropout(dropout)(decoder_lstm_1)
                        # Add additional LSTM layers for stacking in the decoder
                        decoder_lstm_2 = LSTM(latent_dim, kernel_regularizer=tf.keras.regularizers.l2(l2_reg), return_sequences=True)(decoder_drop_1)
                        decoder_drop_2 = Dropout(dropout)(decoder_lstm_2)

                        if Y_sequence == True:


                            outputs = LSTM(horizon, kernel_regularizer=tf.keras.regularizers.l2(l2_reg), return_sequences=False)(decoder_drop_2)
                            outputs = tf.expand_dims(outputs, axis=2)
                            outputs = keras.layers.Dense(len(target_name))(outputs)
                            # print("done")
                        else:
                            decoder_dense_input = LSTM(len(target_name), kernel_regularizer=tf.keras.regularizers.l2(l2_reg), return_sequences=False)(decoder_drop_2)
                            outputs = tf.expand_dims(decoder_dense_input, axis=1)

                        return outputs



                    #with strategy.scope():
                    def build_model(input_shape, correlation_shape, use_graph_layer):

                        inputs = keras.Input(shape=input_shape)
                        if use_graph_layer == True:
                            inp_lap = keras.Input(shape=correlation_shape)
                        else:
                            inp_lap = []

                        if use_graph_layer:
                            graph_output = graph_processing_block(inputs, inp_lap, head_size, num_heads, dropout,
                                                                  horizon)
                        else:
                            graph_output = inputs

                        latent_dim = 64
                        decoder_inputs = tf.keras.Input(
                            shape=(horizon, len(target_name)), name='decoder_inputs')

                        outputs = lstm_encoder_decoder_block(graph_output, decoder_inputs, latent_dim, horizon, target_name, dropout)

                        # Assume 'last_known_value_input' is an additional input layer for the last known value before prediction starts
                        last_known_value_input = tf.keras.Input(shape=(1, len(target_name)), name='last_known_values')
                        #
                        # Integrate the ReverseDifferencingLayer
                        outputs = ReverseDifferencingLayer()([outputs, last_known_value_input])


                        return keras.Model([inputs, inp_lap, decoder_inputs,last_known_value_input], outputs)




                    if iter % model_weights_update_iter == 0 or run_first_time==True:
                        run_first_time=False
                        input_shape = X_train.shape[1:]
                        correlation_shape = correlation_train.shape[1:]

                        model = build_model(
                            input_shape,
                            correlation_shape,
                            use_graph_layer=use_graph_layer
                        )


                        model.compile(
                            loss="mse",
                            optimizer=tf.keras.optimizers.Adam(learning_rate=1e-4),
                            metrics=[tf.keras.metrics.MeanSquaredError()],
                        )


                        model.summary()

                        callbacks = [
                            keras.callbacks.EarlyStopping(patience=20, min_delta=0.001, monitor='val_mean_squared_error',
                                                          mode='auto',
                                                          restore_best_weights=True)]



                        # Define the size of the validation set
                        validation_ratio = 0.25  # 20% or 25% of the data for validation
                        total_data_size = len(X_train)
                        validation_size = int(total_data_size * validation_ratio)

                        # Generate random indices for the validation set
                        # np.random.seed(42)
                        # validation_indices = np.random.choice(total_data_size, validation_size, replace=False)

                        # Select the last validation_size indices for the validation set
                        validation_indices = np.arange(total_data_size - validation_size, total_data_size)


                        # Create masks for selecting training and validation data
                        validation_mask = np.zeros(total_data_size, dtype=bool)
                        validation_mask[validation_indices] = True
                        training_mask = ~validation_mask

                        # Split the data into training and validation sets using the masks
                        X_train_train = X_train[training_mask]
                        if correlation_train.ndim > 1:
                            correlation_train_train = np.squeeze(correlation_train,axis=0)
                        else:
                            correlation_train_train=[]
                        decoder_Y_train_train = decoder_Y_train[training_mask]
                        last_known_values_train_train= last_known_values_train[training_mask]
                        Y_train_train = Y_train[training_mask]

                        X_train_valid = X_train[validation_mask]
                        if correlation_train.ndim > 1:
                            correlation_train_valid = np.squeeze(correlation_train,axis=0)
                        else:
                            correlation_train_valid=[]
                        decoder_Y_train_valid = decoder_Y_train[validation_mask]
                        last_known_values_train_valid = last_known_values_train[validation_mask]
                        Y_train_valid = Y_train[validation_mask]



                        train_gen = data_generator(X_train_train, correlation_train_train, decoder_Y_train_train, Y_train_train, last_known_values_train_train, batch_size=batch_size, new_data_ratio=0)

                        # Create the validation generator
                        val_gen = data_generator(X_train_valid, correlation_train_train, decoder_Y_train_valid, Y_train_valid, last_known_values_train_valid,
                                                   batch_size=batch_size, new_data_ratio=0)

                        # Calculate the steps per epoch for training and validation
                        train_steps = len(X_train_train) // batch_size
                        val_steps = len(X_train_valid) // batch_size


                        if ignore_weights == True:
                            history = model.fit(
                                train_gen,
                                steps_per_epoch=train_steps,
                                validation_data=val_gen,
                                validation_steps=val_steps,
                                epochs=100,
                                callbacks=callbacks
                            )
                            ignore_weights = False


                        else:
                            model.load_weights('model_weights_{}.h5'.format(model_index))
                            # history = model.fit(
                            #     train_gen,
                            #     steps_per_epoch=train_steps,
                            #     validation_data=val_gen,
                            #     validation_steps=val_steps,
                            #     epochs=20,
                            #     callbacks=callbacks
                            # )

                        # Saving weights as per your existing code
                        model.save_weights('model_weights_{}.h5'.format(model_index))

                        # Plotting the training and validation loss
                        # plt.figure(figsize=(10, 6))
                        # plt.plot(history.history['loss'], label='Train Loss')
                        # plt.plot(history.history['val_loss'], label='Validation Loss')
                        # plt.title('Model Loss Over Epochs')
                        # plt.ylabel('Loss')
                        # plt.xlabel('Epoch')
                        # plt.legend(loc='upper right')
                        # plt.show()



                    # Predict from last sequence


                    actual, y_pred_rev, datetime_index, y_pred = final_prediction(x_seq=save_x_seq,
                                                                                  Y_scaler=Y_scaler,
                                                                                  scaler2=scaler2,
                                                                                  decoder_Y_seq=save_decoder_y_seq,
                                                                                  last_known_values=save_last_known_values,
                                                                                  correlation_seq=save_correlation_seq,
                                                                                  sequence_length=sequence_length,
                                                                                  horizon=horizon,
                                                                                  model=model,
                                                                                  data=data_norm,
                                                                                  stride=stride,
                                                                                  target_name=target_name,
                                                                                  differenced_target=differenced_target,
                                                                                  diff_order=diff_order)


                    if single_step==True:

                        # Save prediction in each loop
                        forecast_list = [(np.expand_dims(y_pred_rev[-1, :][-1],axis=0), datetime_index[-1][-1])]
                        actual_list = [(actual[-1, :][-1], datetime_index[-1][-1])]

                        # #Single-step Forecast
                        # forecast_list.append((np.expand_dims(y_pred_rev[-1, :][-1],axis=0), datetime_index[-1][-1]))
                        # actual_list.append((actual[-1, :][-1], datetime_index[-1][-1]))
                    else:

                        #Multi-step Forecast

                        # Save prediction in each loop
                        forecast_list = [(y_pred_rev[-1, :], datetime_index[-1])]
                        actual_list = [(actual[-1, :], datetime_index[-1])]

                        # forecast_list.append((y_pred_rev[-1, :], datetime_index[-1]))
                        # actual_list.append((actual[-1, :], datetime_index[-1]))

                    # Flatten the 3D arrays to 1D arrays
                    flattened_actual = np.concatenate([t[0] for t in actual_list]).flatten()
                    flattened_forecast = np.concatenate([t[0] for t in forecast_list]).flatten()

                    Test_CORR = CORR(flattened_actual, flattened_forecast)
                    Test_RSE = RSE(flattened_actual, flattened_forecast)
                    Test_MAPE = mean_absolute_percentage_error(flattened_actual, flattened_forecast)
                    Test_MAE = mean_absolute_error(flattened_actual, flattened_forecast)
                    Test_MSE = mean_squared_error(flattened_actual, flattened_forecast)

                    # plot last 100 target values:

                    # if reverse_normalization==True and normalized_data==True:
                    #   plot_prediction_graph(original_y=data.loc[end_index-horizon:end_index][target_name_original], forecast_list=forecast_list,
                    #                         test_boundary=test_start_index, horizon=horizon,
                    #                         Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    # elif normalized_data== False:
                    #   transformed_data = scaler2.transform(data.loc[end_index-horizon:end_index][target_name_original])
                    #   # Convert the transformed data back into a pandas DataFrame or Series with the original index
                    #   original_y = pd.Series(transformed_data.squeeze(), index=data.loc[end_index-horizon:end_index].index)
                    #
                    #   plot_prediction_graph(original_y=original_y, forecast_list=forecast_list,
                    #                         test_boundary=test_start_index, horizon=horizon,
                    #                         Test_KPI=Test_MSE, iter=iter, save_dict=directory)
                    # else:
                    #   # For ETD datasets we used (data_norm[target_name] instead of data[target_name_original]) It seems papers used normalized data for prediction results
                    #   plot_prediction_graph(original_y=data_norm[target_name], forecast_list=forecast_list,
                    #                       test_boundary=test_start_index,horizon=horizon,
                    #                       Test_KPI=Test_MSE, iter=iter, save_dict=directory)


                    # plot_norm_diff_graph(Y_seq=np.asarray(save_y_seq), y_pred=y_pred,datetime_index=datetime_index,
                    #                    test_boundary=test_start_index,horizon=horizon,
                    #                    Test_KPI=Test_CORR, iter=iter, save_dict=directory)

                    '''##################### Exporting and Saving  ########################'''


                    if single_step==True and Y_sequence==True:
                        #Single-step Forecast
                        new_row = [iter,
                                   test_range[iter],
                                   str(actual[-1][-1]),
                                   str(y_pred_rev[-1][-1]),
                                   Test_MAPE,
                                   Test_MAE,
                                   Test_MSE,
                                   Test_CORR,
                                   Test_RSE]
                    else:
                        new_row = [iter,
                                   test_range[iter],
                                   str(actual[-1]),
                                   str(y_pred_rev[-1]),
                                   Test_MAPE,
                                   Test_MAE,
                                   Test_MSE,
                                   Test_CORR,
                                   Test_RSE]

                    buffered_rows.append(new_row)


                    if iter % save_to_excel_iter == 0:
                        wb = openpyxl.load_workbook(filename=output_filename)
                        sheet = wb.active
                        # Append the buffered rows to the sheet
                        for row in buffered_rows:
                            sheet.append(row)

                        # Save the Excel file
                        wb.save(output_filename)

                        # Clear the buffer
                        buffered_rows = []

                    iter = iter + 1

                    # Clear the session and invoke garbage collection
                    clear_session()
                    gc.collect()

                # After the loop, append any remaining buffered rows
                if buffered_rows:
                    wb = openpyxl.load_workbook(filename=output_filename)
                    sheet = wb.active
                    for row in buffered_rows:
                        sheet.append(row)
                    wb.save(output_filename)

# '''
# import os
# import pandas as pd
#
# # Directory containing the .dly files
# input_folder = r'C:\Users\sarab\Desktop\desktop\Valley Fever\Data\ghcnd_hcn\ghcnd_hcn'
#
# # Create a new folder for CSV files beside the input folder
# output_folder = os.path.join(os.path.dirname(input_folder), 'converted_csv_files')
#
# # Create the output folder if it doesn't exist
# if not os.path.exists(output_folder):
#     os.makedirs(output_folder)
#
# # Define column widths based on the provided guide
# colspecs = [
#     (0, 11),  # ID (1-11)
#     (11, 15),  # YEAR (12-15)
#     (15, 17),  # MONTH (16-17)
#     (17, 21),  # ELEMENT (18-21)
# ]
#
# # Add specs for the 31 days in the month (value, mflag, qflag, sflag for each day)
# for day in range(1, 32):
#     colspecs.extend([
#         (21 + (day - 1) * 8, 26 + (day - 1) * 8),  # VALUE
#         (26 + (day - 1) * 8, 27 + (day - 1) * 8),  # MFLAG
#         (27 + (day - 1) * 8, 28 + (day - 1) * 8),  # QFLAG
#         (28 + (day - 1) * 8, 29 + (day - 1) * 8),  # SFLAG
#     ])
#
# # Process each .dly file in the input folder
# for filename in os.listdir(input_folder):
#     if filename.endswith('.dly'):
#         file_path = os.path.join(input_folder, filename)
#
#         # Read the .dly file
#         df = pd.read_fwf(file_path, colspecs=colspecs, header=None)
#
#         # Assign names to the columns
#         columns = ['ID', 'YEAR', 'MONTH', 'ELEMENT']
#         for day in range(1, 32):
#             columns.extend([
#                 f'VALUE{day}', f'MFLAG{day}', f'QFLAG{day}', f'SFLAG{day}'
#             ])
#         df.columns = columns
#
#         # Reshape the DataFrame to have each day as a separate row
#         data_rows = []
#         for _, row in df.iterrows():
#             for day in range(1, 32):
#                 value_col = f'VALUE{day}'
#                 if pd.notnull(row[value_col]) and row[value_col] != -9999:  # Skip missing values (-9999)
#                     data_rows.append({
#                         'ID': row['ID'],
#                         'YEAR': row['YEAR'],
#                         'MONTH': row['MONTH'],
#                         'DAY': day,
#                         'ELEMENT': row['ELEMENT'],
#                         'VALUE': row[f'VALUE{day}'],
#                         'MFLAG': row[f'MFLAG{day}'],
#                         'QFLAG': row[f'QFLAG{day}'],
#                         'SFLAG': row[f'SFLAG{day}']
#                     })
#
#         # Convert the reshaped data into a DataFrame
#         df_cleaned = pd.DataFrame(data_rows)
#
#         # Save the reshaped data into a CSV file
#         output_file = os.path.join(output_folder, f'{os.path.splitext(filename)[0]}.csv')
#         df_cleaned.to_csv(output_file, index=False)
#
#         print(f'Converted {filename} to CSV format in {output_folder}.')
#
# '''
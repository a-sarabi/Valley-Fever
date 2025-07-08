import seaborn as sns
import os
GPU_ID="0"
os.environ["CUDA_VISIBLE_DEVICES"] = GPU_ID
import gc
from tensorflow.keras.backend import clear_session

print(os.environ['PATH'])
from tensorflow.python.client import device_lib
print(device_lib.list_local_devices())
import os
import openpyxl
import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.model_selection import train_test_split
import keras.backend as K
import tensorflow as tf
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_percentage_error, mean_squared_error, mean_absolute_error
from spektral.layers import GCNConv, SAGPool, MinCutPool, GlobalAttentionPool, JustBalancePool, DiffPool
from spektral.utils import normalized_adjacency
from sklearn.metrics.pairwise import cosine_similarity
from gatv2_conv import GATv2Conv  # Import the GATv2Conv layer

import random
import datetime
import math
from joblib import Parallel, delayed
from keras.layers import Dense, Dropout, Embedding, LSTM, Bidirectional
from keras.layers import BatchNormalization
from tensorflow.keras.layers import Dropout, Input
from spektral.layers import GATConv
from keras.layers import Add, Concatenate
from tensorflow.keras.layers import Lambda, MultiHeadAttention, LayerNormalization
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras.regularizers import l2
import json

# =====================================================================================
# SAME CLASSES AND FUNCTIONS AS ANALYSIS FILE (needed for model building)
# =====================================================================================

# Add Positional Embedding Layer
class PositionalEmbedding(layers.Layer):
    """
    Adds positional embeddings to the input embeddings.
    """
    def __init__(self, max_sequence_length, d_model, **kwargs):
        super().__init__(**kwargs)
        self.token_embeddings = layers.Dense(d_model)  # Linear projection to d_model dimension
        self.pos_embeddings = layers.Embedding(input_dim=max_sequence_length, output_dim=d_model)
        self.max_sequence_length = max_sequence_length
        self.d_model = d_model

    def call(self, x):
        """
        x shape: (batch_size, sequence_length, input_dim)
        """
        seq_len = tf.shape(x)[1]
        # Ensure positions don't exceed max_sequence_length
        positions = tf.range(start=0, limit=seq_len, delta=1)
        positions = tf.minimum(positions, self.max_sequence_length - 1)
        
        embedded_tokens = self.token_embeddings(x)  # (batch_size, seq_len, d_model)
        embedded_positions = self.pos_embeddings(positions)  # (seq_len, d_model)
        
        # Add positional embeddings to token embeddings
        return embedded_tokens + embedded_positions

# Create a look-ahead mask for decoder self-attention
def create_look_ahead_mask(size):
    """
    Creates a causal/look-ahead mask to prevent decoder from attending to future positions.
    """
    mask_2d = 1 - tf.linalg.band_part(tf.ones((size, size)), -1, 0)
    mask_4d = mask_2d[tf.newaxis, tf.newaxis, :, :]
    return mask_4d  # Shape: (1, 1, size, size)

# Add Transformer Encoder Block
class TransformerEncoder(layers.Layer):
    """
    Transformer Encoder block with multi-head attention and feed-forward network.
    """
    def __init__(self, d_model, num_heads, ff_dim, dropout=0.1, **kwargs):
        super().__init__(**kwargs)
        # Self-attention
        self.mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        self.ffn = keras.Sequential([
            layers.Dense(ff_dim, activation='relu'),
            layers.Dense(d_model),
        ])

        self.layernorm1 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm2 = layers.LayerNormalization(epsilon=1e-6)
        self.dropout1 = layers.Dropout(dropout)
        self.dropout2 = layers.Dropout(dropout)

    def call(self, x, training=False, mask=None):
        # Self Attention
        attn_output, _ = self.mha(
            query=x, value=x, key=x, 
            attention_mask=mask,
            return_attention_scores=True
        )
        attn_output = self.dropout1(attn_output, training=training)
        out1 = self.layernorm1(x + attn_output)

        # Feed-Forward
        ffn_output = self.ffn(out1)
        ffn_output = self.dropout2(ffn_output, training=training)
        out2 = self.layernorm2(out1 + ffn_output)

        return out2

# Add Transformer Decoder Block
class TransformerDecoder(layers.Layer):
    """
    Transformer Decoder block with masked self-attention, cross-attention and feed-forward network.
    """
    def __init__(self, d_model, num_heads, ff_dim, dropout=0.1, **kwargs):
        super().__init__(**kwargs)
        # Self-attention (decoder-side)
        self.self_mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        # Cross-attention (encoder-to-decoder)
        self.cross_mha = layers.MultiHeadAttention(num_heads=num_heads, key_dim=d_model, dropout=dropout)
        self.ffn = keras.Sequential([
            layers.Dense(ff_dim, activation='relu'),
            layers.Dense(d_model),
        ])

        self.layernorm1 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm2 = layers.LayerNormalization(epsilon=1e-6)
        self.layernorm3 = layers.LayerNormalization(epsilon=1e-6)
        self.dropout1 = layers.Dropout(dropout)
        self.dropout2 = layers.Dropout(dropout)
        self.dropout3 = layers.Dropout(dropout)

    def call(self, x, enc_output, training=False,
             look_ahead_mask=None, padding_mask=None):
        target_seq_len = tf.shape(x)[1]
        if look_ahead_mask is None:
            look_ahead_mask = create_look_ahead_mask(target_seq_len)
            
        # 1) Masked self-attention
        attn1, attn_weights_1 = self.self_mha(
            query=x, value=x, key=x, 
            attention_mask=look_ahead_mask,
            return_attention_scores=True
        )
        attn1 = self.dropout1(attn1, training=training)
        out1 = self.layernorm1(x + attn1)

        # 2) Cross-attention with encoder output
        attn2, attn_weights_2 = self.cross_mha(
            query=out1, value=enc_output, key=enc_output,
            attention_mask=padding_mask,
            return_attention_scores=True
        )
        attn2 = self.dropout2(attn2, training=training)
        out2 = self.layernorm2(out1 + attn2)

        # 3) Feed-forward
        ffn_output = self.ffn(out2)
        ffn_output = self.dropout3(ffn_output, training=training)
        out3 = self.layernorm3(out2 + ffn_output)

        return out3

# IMPROVED FEATURE GATE WITH BETTER GRADIENT FLOW
class ImprovedFeatureGate(layers.Layer):
    def __init__(self, num_features, k_percent=0.5, temperature=1.0, **kwargs):
        super().__init__(**kwargs)
        self.num_features = num_features
        self.k_percent = k_percent
        self.temperature = temperature  # For controlling sharpness of selection

    def build(self, input_shape):
        # Trainable importance scores
        self.logits = self.add_weight(
            shape=(self.num_features,),
            initializer="random_normal",
            trainable=True,
            name="feature_importance_logits"
        )

    def call(self, inputs, training=None):
        # Convert logits to importance probabilities
        importance_scores = tf.nn.sigmoid(self.logits)
        
        if training:
            # SOFT GATING during training - better gradient flow
            soft_gates = tf.nn.sigmoid(self.logits / self.temperature)
            threshold = tf.nn.top_k(soft_gates, k=int(self.num_features * self.k_percent))[0][-1]
            soft_gates = tf.nn.sigmoid((soft_gates - threshold) / self.temperature)
            gates_reshaped = tf.reshape(soft_gates, (1, -1, 1))
            
        else:
            # HARD GATING during inference - actual top-k selection
            k = tf.cast(tf.math.ceil(self.k_percent * tf.cast(self.num_features, tf.float32)), tf.int32)
            _, top_k_indices = tf.nn.top_k(importance_scores, k=k)
            mask = tf.zeros_like(importance_scores)
            mask = tf.tensor_scatter_nd_update(
                mask,
                tf.expand_dims(top_k_indices, axis=1),
                tf.ones((k,), dtype=tf.float32)
            )
            hard_gates = importance_scores * mask
            gates_reshaped = tf.reshape(hard_gates, (1, -1, 1))
        
        # Apply gating to inputs
        return inputs * gates_reshaped
    
    def get_feature_importance(self):
        """Get current feature importance scores"""
        return tf.nn.sigmoid(self.logits).numpy()
    
    def get_selected_features(self):
        """Get indices of top-k features"""
        importance = tf.nn.sigmoid(self.logits)
        k = tf.cast(tf.math.ceil(self.k_percent * tf.cast(self.num_features, tf.float32)), tf.int32)
        _, top_k_indices = tf.nn.top_k(importance, k=k)
        return top_k_indices.numpy()

# Alternative simpler approach - Learnable feature weights without hard selection
class SoftFeatureGate(layers.Layer):
    def __init__(self, num_features, **kwargs):
        super().__init__(**kwargs)
        self.num_features = num_features

    def build(self, input_shape):
        # Learnable importance weights
        self.feature_weights = self.add_weight(
            shape=(self.num_features,),
            initializer="ones",  # Start with equal importance
            trainable=True,
            name="feature_weights"
        )

    def call(self, inputs):
        # Apply sigmoid to ensure weights are in [0,1]
        normalized_weights = tf.nn.sigmoid(self.feature_weights)
        weights_reshaped = tf.reshape(normalized_weights, (1, -1, 1))
        return inputs * weights_reshaped
    
    def get_feature_importance(self):
        """Get current feature importance scores"""
        return tf.nn.sigmoid(self.feature_weights).numpy()

class ConsistentFeatureGate(layers.Layer):
    """
    Feature gate that uses pre-selected features instead of learning them dynamically.
    """
    def __init__(self, num_features, selected_feature_indices, **kwargs):
        super().__init__(**kwargs)
        self.num_features = num_features
        self.selected_feature_indices = selected_feature_indices
        
    def build(self, input_shape):
        # Create a fixed mask for selected features
        self.feature_mask = tf.zeros(self.num_features, dtype=tf.float32)
        self.feature_mask = tf.tensor_scatter_nd_update(
            self.feature_mask,
            tf.expand_dims(tf.constant(self.selected_feature_indices), axis=1),
            tf.ones(len(self.selected_feature_indices), dtype=tf.float32)
        )
        
    def call(self, inputs):
        # Apply the fixed mask
        mask_reshaped = tf.reshape(self.feature_mask, (1, -1, 1))
        return inputs * mask_reshaped
    
    def get_feature_importance(self):
        """Get the fixed feature importance mask"""
        return self.feature_mask.numpy()
    
    def get_selected_features(self):
        """Get the pre-selected feature indices"""
        return self.selected_feature_indices

class ReverseDifferencingLayer(tf.keras.layers.Layer):
    def __init__(self, **kwargs):
        super(ReverseDifferencingLayer, self).__init__(**kwargs)

    def call(self, inputs):
        """
        Perform the reversal of differencing operation.
        """
        preds, last_known_value = inputs
        # Use tf.cumsum to reverse the differencing, adding the last known value as the base.
        reversed_preds = tf.cumsum(preds, axis=1) + last_known_value
        return reversed_preds

# =====================================================================================
# DATA PROCESSING FUNCTIONS (same as analysis file)
# =====================================================================================

def data_preprocess(df, diff_order, start_index, end_index, moving_average, MA_window_size):
    """
    This function removes columns with more than 10% missing values and fills in the NaNs in the remaining columns.
    """
    df = pd.DataFrame(df)
    df = df.loc[start_index:end_index]

    # Apply differencing according to the specified diff_order
    df_diff = df.copy()
    for _ in range(diff_order):
        if moving_average==True:
            # Function to pad the beginning of the DataFrame
            def pad_head(df, pad_width):
                head = pd.DataFrame([df.iloc[0].values] * pad_width, columns=df.columns)
                return pd.concat([head, df]).reset_index(drop=True)

            # Apply padding only to the head
            pad_width = MA_window_size - 1
            df_padded = pad_head(df_diff, pad_width)

            # Compute the moving average on the padded DataFrame
            df_ma_padded = df_padded.rolling(window=MA_window_size, center=False).mean()

            # Now, remove the padding by trimming the extra rows from the start of the moving average DataFrame
            df_ma = df_ma_padded.iloc[pad_width:].reset_index(drop=True)

            # Shift the moving average DataFrame by one to align it for differencing
            df_ma_shifted = df_ma.shift(1)

            # Ensure df_diff is aligned in size with df_ma_shifted
            df_diff_aligned = df_diff.iloc[:df_ma_shifted.shape[0]]

            # Calculate the difference between the actual value at t+1 and the moving average at t
            df_diff = df_diff_aligned - df_ma_shifted
        else:
            df_diff = df_diff.diff(periods=1)

    # Adjust the index offset for the differencing order
    df_diff = df_diff.loc[start_index + diff_order:end_index]

    return df, df_diff

def data_set_generation(data, data_diff, max_lag, target_as_feature, target_name):
    # update the input_components list
    input_components = list(data.columns)

    # remove the target column before lagging the variables
    if not target_as_feature:
        for target in target_name:
            if target in input_components:
                input_components.remove(target)

    # List of updated input components.
    target_comp_updated_list = []

    # Initialize forecast dataset.
    new_data = pd.DataFrame()
    new_data_diff = pd.DataFrame()

    # Iterate through input components and their lags.
    for col in input_components:
        for lag in range(0, max_lag + 1):
            # Generate new component name.
            comp_new = f"{col} ({-lag})"
            if col in target_name:
                target_comp_updated_list.append(comp_new)
            # Get start index based on optimal time lag.
            new_data_time_series = data[col].shift(lag).rename(comp_new)
            new_data_diff_time_series = data_diff[col].shift(lag).rename(comp_new)

            # Add to dataset and forecast dataset
            new_data = pd.concat([new_data, new_data_time_series], axis=1)
            new_data_diff = pd.concat([new_data_diff, new_data_diff_time_series], axis=1)

    # Drop the first few rows to remove Nans.
    new_data = new_data.iloc[max_lag:]
    new_data_diff = new_data_diff.iloc[max_lag:]

    update_target_name = []
    if target_as_feature == True:
        for i, col in enumerate(target_name):
            update_target_name.append(f"{col} (0)")
        target_name = update_target_name
    else:
        # append the removed target column to datasets
        new_data[target_name] = data[target_name]
        new_data_diff[target_name] = data_diff[target_name]
        target_comp_updated_list.extend(target_name)

    return new_data, new_data_diff, target_comp_updated_list, target_name

def preprocessing(data_, data_diff_, diff_order, sequence_length, horizon, stride, use_graph_layer,
                  save_instance,ignore_first_instance_stride, save_x_seq, save_y_seq, save_correlation_seq,
                  save_decoder_y_seq, save_last_known_values, target_name,
                  target_as_feature, target_comp_updated_list, differenced_target, differenced_X, moving_average,MA_window_size):
    
    if differenced_X == True:
        X = data_diff_
    else:
        X = data_[diff_order:]

    X_corr = data_diff_

    if differenced_target == True:
        Y = data_diff_[target_name].values
    else:
        Y = data_[diff_order:][target_name].values

    Y_original = data_[diff_order:][target_name].values

    if target_as_feature == False:
        X = X.drop(target_comp_updated_list, axis=1)
        X_corr = X_corr.drop(target_comp_updated_list, axis=1)
    X = X.values

    # Sequence Generation Part
    x_seq = []
    y_seq = []
    correlation_seq = []
    decoder_y_seq = []
    n = np.shape(X)[1]
    corr = np.ones((n, n))

    if use_graph_layer == True:
        # calculate starting correlation based on all available data
        corr = np.corrcoef(X_corr[:len(X_corr) - horizon], rowvar=False)
        # Apply the filter to keep only correlations with absolute values greater than 0.05
        corr = np.where(np.abs(corr) > 0.05, corr, 0)
        correlation_seq.append(corr)

    for instance in range(save_instance, len(X) - sequence_length - horizon +1, stride):
        if ignore_first_instance_stride:
            ignore_first_instance_stride = False
            continue
        x_seq.append(X[instance: instance + sequence_length])

        Y_sequence = True  # Set this based on your original setting
        if Y_sequence == True:
            y_seq.append(Y_original[instance + sequence_length:instance + sequence_length + horizon])
        else:
            y_seq.append(np.expand_dims(np.sum(Y_original[instance + sequence_length:instance + sequence_length + horizon], axis=0), axis=0))

        decoder_y_seq.append(Y[instance + sequence_length - horizon:instance + sequence_length])

        if moving_average==True:
            last_value_before_sequence = data_[target_name].iloc[max(0,
                                                                     instance + sequence_length - diff_order + 2 - MA_window_size):instance + sequence_length - diff_order + 2].mean()
        else:
            last_value_before_sequence = data_[target_name].iloc[instance + sequence_length - diff_order + 1]

        last_value_before_sequence = np.expand_dims(last_value_before_sequence, axis=0)
        save_last_known_values.append(last_value_before_sequence)

    del X
    save_instance = instance
    save_x_seq.extend(x_seq)
    save_y_seq.extend(y_seq)
    save_decoder_y_seq.extend(decoder_y_seq)
    save_correlation_seq.extend(correlation_seq)

    return save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values

def data_generator(X, correlation, decoder_Y, Y, last_known_values, batch_size, new_data_ratio=0.1):
    while True:
        new_data_count = int(batch_size * new_data_ratio)
        new_data_indices = np.arange(len(X) - new_data_count, len(X))

        for i in range(0, len(X), batch_size):
            random_indices = np.random.permutation(len(X) - new_data_count)
            combined_indices = np.concatenate([new_data_indices, random_indices[:batch_size - new_data_count]])
            np.random.shuffle(combined_indices)

            batch_X = X[combined_indices]
            batch_decoder_Y = decoder_Y[combined_indices]
            batch_last_known_values = last_known_values[combined_indices]
            batch_Y = Y[combined_indices]

            if len(correlation) > 0:
                repeat_corr = True  # Set this based on your original setting
                if repeat_corr == True:
                    batch_correlation = np.repeat(correlation[np.newaxis, ...], len(batch_X), axis=0)
                else:
                    batch_correlation = correlation[combined_indices]
            else:
                batch_correlation=[]

            yield [batch_X, batch_correlation, batch_decoder_Y, batch_last_known_values], batch_Y

def graph_processing_block(inputs, inp_lap, head_size, num_heads, dropout, horizon, consistently_selected_features=None):
    """
    Process inputs through graph convolutional layers with improved feature gating.
    """
    l2_reg = 2.5e-4  # L2 regularization rate
    
    # Save original sequence length for reshaping later
    batch_size, original_seq_len, n_features = tf.shape(inputs)[0], tf.shape(inputs)[1], tf.shape(inputs)[2]
    
    # Transpose from (batch, time, features) to (batch, features, time)
    x = tf.transpose(inputs, perm=[0, 2, 1])
    
    # Use consistent feature selection if available, otherwise use dynamic selection
    if consistently_selected_features is not None:
        x = ConsistentFeatureGate(
            num_features=int(x.shape[1]),
            selected_feature_indices=consistently_selected_features,
            name='consistent_feature_gate'
        )(x)
    else:
        # Fallback to improved feature gate
        x = ImprovedFeatureGate(
            num_features=int(x.shape[1]), 
            k_percent=0.1, 
            temperature=0.1,
            name='improved_feature_gate'
        )(x)
    
    do_1 = Dropout(dropout)(x)
    
    # GAT operates on the node/feature dimension
    gc_1, gc_1_attn = GATv2Conv(
        int(math.ceil(x.shape[2])),
        attn_heads=num_heads,
        concat_heads=False,
        dropout_rate=dropout,
        activation="relu",
        kernel_regularizer=l2(l2_reg),
        attn_kernel_regularizer=l2(l2_reg),
        bias_regularizer=l2(l2_reg),
        return_attn_coef=True,
    )([do_1, inp_lap])
    
    # Transpose back to (batch, time, features)
    graph_output = tf.transpose(gc_1, perm=[0, 2, 1])
    
    gc_2_attn = gc_1_attn
    
    return graph_output, gc_1_attn, gc_2_attn

def transformer_encoder_decoder_block(graph_output, decoder_inputs, d_model, num_heads, ff_dim, horizon, target_name, dropout):
    """
    Transformer-based encoder-decoder block for time series forecasting.
    """
    l2_reg = 2.5e-4  # L2 regularization rate
    
    sequence_length = tf.shape(graph_output)[1]
    max_encoder_length = 200  # Maximum possible sequence length
    
    # Add positional embeddings to encoder input
    encoder_embedding = PositionalEmbedding(max_sequence_length=max_encoder_length, d_model=d_model)
    enc_emb = encoder_embedding(graph_output)
    
    # Encoder stacks (2 layers)
    encoder_output = enc_emb
    for i in range(2):
        encoder_block = TransformerEncoder(d_model, num_heads, ff_dim, dropout=dropout)
        encoder_output = encoder_block(encoder_output)
    
    # Create look-ahead mask for decoder
    look_ahead_mask = create_look_ahead_mask(horizon)
    
    # Add positional embeddings to decoder input
    decoder_embedding = PositionalEmbedding(max_sequence_length=horizon, d_model=d_model)
    dec_emb = decoder_embedding(decoder_inputs)
    
    # Decoder stacks (2 layers)
    decoder_output = dec_emb
    for i in range(2):
        decoder_block = TransformerDecoder(d_model, num_heads, ff_dim, dropout=dropout)
        decoder_output = decoder_block(
            decoder_output, 
            encoder_output,
            look_ahead_mask=look_ahead_mask
        )
    
    # Final output layer
    Y_sequence = True  # Set this based on your original setting
    if Y_sequence == True:
        outputs = keras.layers.TimeDistributed(keras.layers.Dense(len(target_name)))(decoder_output)
    else:
        outputs = keras.layers.Dense(len(target_name))(decoder_output[:, -1:, :])
    
    return outputs

def build_model(input_shape, correlation_shape, use_graph_layer, consistently_selected_features=None):
    inputs = keras.Input(shape=input_shape)
    if use_graph_layer == True:
        inp_lap = keras.Input(shape=correlation_shape)
    else:
        inp_lap = []

    # Set parameters
    head_size = 16
    num_heads = 8
    ff_dim = 256
    dropout = 0.05
    horizon = input_shape[0] // 3  # Derive horizon from sequence length
    target_name = ['MARICOPA']  # Set this based on your data

    if use_graph_layer:
        graph_output, gc_1_attn, gc_2_attn = graph_processing_block(inputs, inp_lap, head_size,
                                                                    num_heads, dropout,
                                                                    horizon, consistently_selected_features)
    else:
        graph_output = inputs

    # Project to consistent dimension for transformer if needed
    d_model = 256  # Transformer embedding dimension
    graph_output = layers.Dense(d_model)(graph_output)
    
    # Create decoder inputs
    decoder_inputs = tf.keras.Input(
        shape=(horizon, len(target_name)), name='decoder_inputs')

    # Use transformer encoder-decoder
    outputs = transformer_encoder_decoder_block(
        graph_output=graph_output,
        decoder_inputs=decoder_inputs,
        d_model=d_model,
        num_heads=num_heads,
        ff_dim=ff_dim,
        horizon=horizon,
        target_name=target_name,
        dropout=dropout
    )

    last_known_value_input = tf.keras.Input(shape=(1, len(target_name)), name='last_known_values')
    
    differenced_target = True  # Set this based on your original setting
    if differenced_target == True:
        outputs = ReverseDifferencingLayer()([outputs, last_known_value_input])

    # Return a model with outputs for training and a submodel for attention
    model = keras.Model(
        [inputs, inp_lap, decoder_inputs, last_known_value_input],
        outputs,
    )
    if use_graph_layer:
        model.attention_submodel = keras.Model(
            [inputs, inp_lap, decoder_inputs, last_known_value_input],
            [gc_1_attn, gc_2_attn],
        )

    return model

# =====================================================================================
# FORECASTING SPECIFIC FUNCTIONS
# =====================================================================================

def CORR(flattened_actual, flattened_forecast):
    """
    Calculate the Empirical Correlation Coefficient between two flattened arrays.
    """
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    sigma_p = flattened_forecast.std()
    sigma_g = flattened_actual.std()
    mean_p = flattened_forecast.mean()
    mean_g = flattened_actual.mean()

    valid_indices = (sigma_g != 0) and (sigma_p != 0)
    if np.any(valid_indices):
        correlation = np.mean(
            ((flattened_forecast[valid_indices] - mean_p) * (flattened_actual[valid_indices] - mean_g)) / (sigma_p * sigma_g)
        )
    else:
        correlation = 0

    return correlation

def RSE(flattened_actual, flattened_forecast):
    """
    Calculate the Root Relative Squared Error between two flattened arrays.
    """
    flattened_actual = np.array(flattened_actual)
    flattened_forecast = np.array(flattened_forecast)

    mean_actual = np.mean(flattened_actual)
    numerator = np.sum((flattened_actual - flattened_forecast) ** 2)
    denominator = np.sum((flattened_actual - mean_actual) ** 2)

    if denominator != 0:
        rse = np.sqrt(numerator / denominator)
    else:
        rse = np.nan

    return rse

def final_prediction(x_seq, Y_scaler, scaler2, decoder_Y_seq, last_known_values, correlation_seq, 
                    sequence_length, horizon, model, data, stride, target_name, differenced_target, 
                    diff_order, save_dict):

    x_seq = np.expand_dims(x_seq[-1], axis=0)
    decoder_Y_seq = np.expand_dims(decoder_Y_seq[-1], axis=0)
    last_known_values = np.expand_dims(last_known_values[-1], axis=0)

    if len(correlation_seq) > 0:
        correlation_seq = np.expand_dims(correlation_seq[-1], axis=0)

    # Prediction
    y_pred = model.predict([x_seq, correlation_seq, decoder_Y_seq, last_known_values])

    y_pred_rev = y_pred

    # Handle normalization reversal
    normalized_data = True
    reverse_normalization = True
    if normalized_data == True:
        if reverse_normalization == True:
            y_pred_rev = np.expand_dims(Y_scaler.inverse_transform(np.squeeze(y_pred_rev, axis=0)), axis=0)
        else:
            y_pred_rev = np.expand_dims(np.squeeze(y_pred_rev, axis=0), axis=0)
    else:
        y_pred_rev = np.expand_dims(np.squeeze(y_pred_rev, axis=0), axis=0)

    n_rows = data.shape[0] - sequence_length - horizon
    indices = np.arange(1 + sequence_length, 1 + n_rows + sequence_length, stride)

    Y_sequence = True
    if Y_sequence == True:
        if normalized_data == True:
            if reverse_normalization == True:
                actual_array = np.expand_dims(Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + np.arange(horizon)]), axis=0)
            else:
                actual_array = np.expand_dims(data[target_name].values[indices[-1, None] + np.arange(horizon)], axis=0)
        else:
            actual_array = np.expand_dims(data[target_name].values[indices[-1, None] + np.arange(horizon)], axis=0)

        datetime_index = np.expand_dims(data.index.to_numpy()[indices[-1, None] + np.arange(horizon)], axis=0)
    else:
        if normalized_data == True:
            actual_array = Y_scaler.inverse_transform(data[target_name].values[indices[-1, None] + horizon - 1])
        else:
            actual_array = data[target_name].values[indices[-1, None] + horizon - 1]
        datetime_index = data.index.to_numpy()[indices[-1, None] + horizon - 1]

    return actual_array, y_pred_rev, datetime_index, y_pred

def plot_prediction_graph(original_y, forecast_list, test_boundary, horizon, Test_KPI, iter, save_dict):
    plt.figure(figsize=(15, 10))
    plt.plot(original_y.index[-horizon:], original_y[-horizon:], 'r--', marker='o')
    plt.plot(forecast_list[-1][1], forecast_list[-1][0], 'b', linewidth=0.6, marker='o')

    # Create a legend with custom labels
    target_name = ['MARICOPA']  # Set based on your data
    actual_labels = ['actual_{}'.format(i) for i in range(len(target_name))]
    forecast_labels = ['forecast_{}'.format(i) for i in range(len(target_name))]
    legend_labels = actual_labels + forecast_labels
    plt.legend(legend_labels)

    plt.xlabel('Date')
    plt.ylabel('Value')
    plt.text(0.02, 0.8, 'KPI for Test Range = {:.4f}'.format(Test_KPI), bbox=dict(facecolor='red', alpha=0.2),
             transform=plt.gca().transAxes, va="top", ha="left")
    plt.title('{}- Week Ahead Model_it={}'.format(horizon, iter + 1), loc='center')

    plt.savefig("{}/prediction_{}_it={}.png".format(save_dict, horizon, iter + 1), dpi=300)
    plt.close()

# =====================================================================================
# FILE LOADING AND VALIDATION FUNCTIONS
# =====================================================================================

def load_analysis_results(results_directory):
    """
    Load the analysis results from the specified directory and validate required files
    
    Returns:
        dict: Dictionary containing loaded results for each horizon
    """
    
    print(f"🔍 Loading analysis results from: {results_directory}")
    
    if not os.path.exists(results_directory):
        print(f"❌ Error: Directory {results_directory} does not exist!")
        print("💡 Please run the stability analysis script first to generate the required files.")
        return None
    
    # Find all horizon directories
    horizon_dirs = []
    for item in os.listdir(results_directory):
        item_path = os.path.join(results_directory, item)
        if os.path.isdir(item_path) and item.startswith('horizon_'):
            try:
                horizon = int(item.split('_')[1])
                horizon_dirs.append((horizon, item_path))
            except (IndexError, ValueError):
                continue
    
    if not horizon_dirs:
        print(f"❌ Error: No horizon directories found in {results_directory}")
        print("💡 Expected directories: horizon_2, horizon_4, horizon_8, horizon_16")
        return None
    
    horizon_dirs.sort()  # Sort by horizon
    print(f"📂 Found {len(horizon_dirs)} horizon directories: {[f'horizon_{h}' for h, _ in horizon_dirs]}")
    
    # Load results for each horizon
    loaded_results = {}
    
    for horizon, horizon_dir in horizon_dirs:
        print(f"\n📖 Loading results for {horizon}-week horizon...")
        
        # Required files
        required_files = {
            'metadata': os.path.join(horizon_dir, 'forecasting_metadata.json'),
            'selected_features': os.path.join(horizon_dir, 'consistently_selected_features.csv'),
            'feature_info': os.path.join(horizon_dir, 'selected_features_info.csv'),
            'stability_results': os.path.join(horizon_dir, 'stability_analysis_results.csv')
        }
        
        # Check if all required files exist
        missing_files = []
        for file_type, file_path in required_files.items():
            if not os.path.exists(file_path):
                missing_files.append(f"{file_type}: {file_path}")
        
        if missing_files:
            print(f"⚠️  Warning: Missing files for {horizon}-week horizon:")
            for missing in missing_files:
                print(f"     - {missing}")
            print("   Skipping this horizon...")
            continue
        
        try:
            # Load metadata
            with open(required_files['metadata'], 'r') as f:
                metadata = json.load(f)
            
            # Load selected features
            selected_features_df = pd.read_csv(required_files['selected_features'])
            
            # Load feature info
            feature_info_df = pd.read_csv(required_files['feature_info'])
            
            # Load stability results
            stability_results_df = pd.read_csv(required_files['stability_results'])
            
            # Store loaded data
            loaded_results[horizon] = {
                'metadata': metadata,
                'selected_features_df': selected_features_df,
                'feature_info_df': feature_info_df,
                'stability_results_df': stability_results_df,
                'selected_feature_indices': metadata['selected_feature_indices'],
                'lagged_feature_names': metadata['lagged_feature_names'],
                'target_name_original': metadata['target_name_original'],
                'max_lag': metadata['max_lag'],
                'diff_order': metadata['diff_order']
            }
            
            print(f"   ✅ Successfully loaded {len(metadata['selected_feature_indices'])} selected features")
            print(f"   📊 Feature selection: {metadata['selected_features_count']}/{metadata['total_features']} features")
            
        except Exception as e:
            print(f"   ❌ Error loading files for {horizon}-week horizon: {str(e)}")
            continue
    
    if not loaded_results:
        print(f"\n❌ No valid horizon results could be loaded!")
        return None
    
    print(f"\n✅ Successfully loaded results for {len(loaded_results)} horizons: {list(loaded_results.keys())}")
    return loaded_results

def run_forecasting_for_horizon(horizon, horizon_data, data, df_norm, forecasting_dir):
    """
    Run the actual forecasting loop for a specific horizon using loaded feature selection results
    """
    
    print(f"\n🔮 Running forecasting for horizon: {horizon} weeks")
    
    # Extract data from loaded results
    metadata = horizon_data['metadata']
    consistently_selected_features = horizon_data['selected_feature_indices']
    target_name_original = metadata['target_name_original']
    max_lag = metadata['max_lag']
    diff_order = metadata['diff_order']
    
    print(f"🔒 Using {len(consistently_selected_features)} pre-selected features from stability analysis")
    
    # Set parameters based on horizon
    sequence_length = 3 * horizon
    stride = 1
    dynamic_test_size = math.ceil(horizon / stride)
    use_graph_layer = True
    target_as_feature = True
    differenced_target = True
    differenced_X = True
    normalized_data = True
    reverse_normalization = True
    initial_training_normalization = True
    repeat_corr = True
    Y_sequence = True
    moving_average = False
    MA_window_size = 12
    batch_size = 10
    
    # Model parameters
    head_size = 16
    num_heads = 8
    ff_dim = 256
    dropout = 0.05
    
    # Test parameters
    start_index = 0
    train_end = 850
    test_start_index = 900
    test_end_index = 1002
    test_range = range(test_start_index, test_end_index)
    
    # Create Excel file for results
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = f"{forecasting_dir}/forecasting_results_{horizon}week_{now}.xlsx"
    
    if not os.path.exists(output_filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        new_row = ['iteration', 'prediction_date', 'actual_value', 'forecast_value', 
                  'Test_MAPE', 'Test_MAE', 'Test_MSE', 'Test_CORR', 'Test_RSE']
        sheet.append(new_row)
        wb.save(output_filename)
    
    # Initialize variables
    forecast_list = []
    actual_list = []
    iter = 0
    save_instance = 0
    ignore_first_instance_stride = False
    ignore_weights = True
    save_x_seq = []
    save_y_seq = []
    save_correlation_seq = []
    save_decoder_y_seq = []
    save_last_known_values = []
    run_first_time = True
    model_weights_update_iter = int(round(0.2 * len(data))) + 1
    
    # Setup normalization
    initial_train_data = data.loc[start_index:train_end]
    if initial_training_normalization and normalized_data:
        scaler = StandardScaler()
        scaler.fit(initial_train_data)
        scaler2 = StandardScaler()
        Y_scaler = scaler2.fit(initial_train_data[target_name_original])
    else:
        scaler2 = StandardScaler()
        Y_scaler = scaler2.fit(initial_train_data[target_name_original])
    
    iter = horizon - 1
    if ignore_weights == False:
        iter = 2619
    
    # Start forecasting loop
    buffered_rows = []
    save_to_excel_iter = 10
    
    for end_index in test_range[iter:]:
        print("##################### run # {} / end index: {} ########################### ".format(iter, end_index))
        
        SEED = 33
        tf.random.set_seed(SEED)
        os.environ['PYTHONHASHSEED'] = str(SEED)
        np.random.seed(SEED)
        random.seed(SEED)

        session_conf = tf.compat.v1.ConfigProto(
            intra_op_parallelism_threads=1,
            inter_op_parallelism_threads=1
        )
        sess = tf.compat.v1.Session(
            graph=tf.compat.v1.get_default_graph(),
            config=session_conf
        )
        tf.compat.v1.keras.backend.set_session(sess)

        # Data preprocessing
        if initial_training_normalization == False and normalized_data == True:
            df = data.loc[start_index:end_index]
            scaler = StandardScaler()
            scaler.fit(df[:-horizon])
            df_norm_array = scaler.transform(df)
            df_norm_local = pd.DataFrame(df_norm_array, columns=df.columns, index=df.index)
            scaler2 = StandardScaler()
            Y_scaler = scaler2.fit(df[target_name_original][:-horizon])
        else:
            df_norm_local = df_norm.loc[start_index:end_index]

        data_norm, data_diff = data_preprocess(df_norm_local, diff_order, start_index, end_index, moving_average, MA_window_size)

        # Generate lagged variables
        data_norm, data_diff, target_comp_updated_list, target_name = data_set_generation(
            data_norm, data_diff, max_lag=max_lag, target_as_feature=target_as_feature, target_name=target_name_original)

        save_instance, save_x_seq, save_y_seq, save_decoder_y_seq, save_correlation_seq, save_last_known_values = preprocessing(
            data_=data_norm, data_diff_=data_diff, diff_order=diff_order, sequence_length=sequence_length,
            horizon=horizon, stride=stride, use_graph_layer=use_graph_layer, save_instance=save_instance,
            ignore_first_instance_stride=ignore_first_instance_stride, save_x_seq=save_x_seq, save_y_seq=save_y_seq,
            save_correlation_seq=save_correlation_seq, save_decoder_y_seq=save_decoder_y_seq,
            save_last_known_values=save_last_known_values, target_name=target_name,
            target_as_feature=target_as_feature, target_comp_updated_list=target_comp_updated_list,
            differenced_target=differenced_target, differenced_X=differenced_X, moving_average=moving_average, 
            MA_window_size=MA_window_size
        )

        del data_diff

        # Test data split
        if repeat_corr == True:
            X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test, last_known_values_train, last_known_values_test = \
                train_test_split(np.asarray(save_x_seq), np.asarray(save_y_seq), np.asarray(save_decoder_y_seq),
                               np.asarray(save_last_known_values), test_size=dynamic_test_size, shuffle=False, random_state=1004)

            correlation_train = np.array(save_correlation_seq)
            correlation_test = correlation_train
        else:
            X_train, X_test, Y_train, Y_test, decoder_Y_train, decoder_Y_test, last_known_values_train, last_known_values_test, correlation_train, correlation_test = \
                train_test_split(np.asarray(save_x_seq), np.asarray(save_y_seq), np.asarray(save_decoder_y_seq),
                               np.asarray(save_last_known_values), np.asarray(save_correlation_seq),
                               test_size=dynamic_test_size, shuffle=False, random_state=1004)

        ignore_first_instance_stride = True

        if iter % model_weights_update_iter == 0 or run_first_time == True:
            run_first_time = False
            input_shape = X_train.shape[1:]
            correlation_shape = correlation_train.shape[1:]

            # Use consistently selected features from loaded stability analysis
            model = build_model(input_shape, correlation_shape, use_graph_layer=use_graph_layer,
                              consistently_selected_features=consistently_selected_features)

            model.compile(loss="mse", optimizer=tf.keras.optimizers.Adam(learning_rate=1e-4),
                         metrics=[tf.keras.metrics.MeanSquaredError()])

            model.summary()

            callbacks = [keras.callbacks.EarlyStopping(patience=20, min_delta=0.001, monitor='val_mean_squared_error',
                                                      mode='auto', restore_best_weights=True)]

            # Define the size of the validation set
            validation_ratio = 0.25
            total_data_size = len(X_train)
            validation_size = int(total_data_size * validation_ratio)

            # Select the last validation_size indices for the validation set
            validation_indices = np.arange(total_data_size - validation_size, total_data_size)

            # Create masks for selecting training and validation data
            validation_mask = np.zeros(total_data_size, dtype=bool)
            validation_mask[validation_indices] = True
            training_mask = ~validation_mask

            # Split the data into training and validation sets using the masks
            X_train_train = X_train[training_mask]
            if correlation_train.ndim > 1:
                correlation_train_train = np.squeeze(correlation_train, axis=0)
            else:
                correlation_train_train = []
            decoder_Y_train_train = decoder_Y_train[training_mask]
            last_known_values_train_train = last_known_values_train[training_mask]
            Y_train_train = Y_train[training_mask]

            X_train_valid = X_train[validation_mask]
            decoder_Y_train_valid = decoder_Y_train[validation_mask]
            last_known_values_train_valid = last_known_values_train[validation_mask]
            Y_train_valid = Y_train[validation_mask]

            train_gen = data_generator(X_train_train, correlation_train_train, decoder_Y_train_train, 
                                     Y_train_train, last_known_values_train_train, batch_size=batch_size, new_data_ratio=0)

            val_gen = data_generator(X_train_valid, correlation_train_train, decoder_Y_train_valid, 
                                   Y_train_valid, last_known_values_train_valid, batch_size=batch_size, new_data_ratio=0)

            # Calculate the steps per epoch for training and validation
            train_steps = len(X_train_train) // batch_size
            val_steps = len(X_train_valid) // batch_size

            if ignore_weights == True:
                history = model.fit(train_gen, steps_per_epoch=train_steps, validation_data=val_gen,
                                  validation_steps=val_steps, epochs=100, callbacks=callbacks)
                ignore_weights = False
            else:
                model.load_weights(f'model_weights_{horizon}week.h5')

            # Saving weights
            model.save_weights(f'model_weights_{horizon}week.h5')

        # Predict from last sequence
        actual, y_pred_rev, datetime_index, y_pred = final_prediction(
            x_seq=save_x_seq, Y_scaler=Y_scaler, scaler2=scaler2, decoder_Y_seq=save_decoder_y_seq,
            last_known_values=save_last_known_values, correlation_seq=save_correlation_seq,
            sequence_length=sequence_length, horizon=horizon, model=model, data=data_norm,
            stride=stride, target_name=target_name, differenced_target=differenced_target,
            diff_order=diff_order, save_dict=forecasting_dir)

        # Save prediction in each loop
        forecast_list = [(y_pred_rev[-1, :], datetime_index[-1])]
        actual_list = [(actual[-1, :], datetime_index[-1])]

        # Flatten the 3D arrays to 1D arrays
        flattened_actual = np.concatenate([t[0] for t in actual_list]).flatten()
        flattened_forecast = np.concatenate([t[0] for t in forecast_list]).flatten()

        Test_CORR = CORR(flattened_actual, flattened_forecast)
        Test_RSE = RSE(flattened_actual, flattened_forecast)
        Test_MAPE = mean_absolute_percentage_error(flattened_actual, flattened_forecast)
        Test_MAE = mean_absolute_error(flattened_actual, flattened_forecast)
        Test_MSE = mean_squared_error(flattened_actual, flattened_forecast)

        if reverse_normalization == True and normalized_data == True:
            plot_prediction_graph(original_y=data.loc[end_index-horizon:end_index][target_name_original], 
                                 forecast_list=forecast_list, test_boundary=test_start_index, 
                                 horizon=horizon, Test_KPI=Test_MSE, iter=iter, save_dict=forecasting_dir)
        elif normalized_data == False:
            transformed_data = scaler2.transform(data.loc[end_index-horizon:end_index][target_name_original])
            original_y = pd.Series(transformed_data.squeeze(), index=data.loc[end_index-horizon:end_index].index)
            plot_prediction_graph(original_y=original_y, forecast_list=forecast_list,
                                 test_boundary=test_start_index, horizon=horizon,
                                 Test_KPI=Test_MSE, iter=iter, save_dict=forecasting_dir)
        else:
            plot_prediction_graph(original_y=data_norm[target_name], forecast_list=forecast_list,
                                 test_boundary=test_start_index, horizon=horizon,
                                 Test_KPI=Test_MSE, iter=iter, save_dict=forecasting_dir)

        # Export and save results
        new_row = [iter, test_range[iter], str(actual[-1]), str(y_pred_rev[-1]),
                  Test_MAPE, Test_MAE, Test_MSE, Test_CORR, Test_RSE]

        buffered_rows.append(new_row)

        if iter % save_to_excel_iter == 0:
            wb = openpyxl.load_workbook(filename=output_filename)
            sheet = wb.active
            for row in buffered_rows:
                sheet.append(row)
            wb.save(output_filename)
            buffered_rows = []

        iter = iter + 1

        # Clear the session and invoke garbage collection
        clear_session()
        gc.collect()

    # After the loop, append any remaining buffered rows
    if buffered_rows:
        wb = openpyxl.load_workbook(filename=output_filename)
        sheet = wb.active
        for row in buffered_rows:
            sheet.append(row)
        wb.save(output_filename)

    print(f"✅ Completed forecasting for {horizon}-week horizon!")
    print(f"📊 Results saved to: {output_filename}")
    
    return output_filename

# =====================================================================================
# MAIN EXECUTION CODE
# =====================================================================================

if __name__ == '__main__':

    print("🚀 Multi-Horizon Forecasting Script")
    print("="*60)
    print("This script loads pre-computed feature selection results")
    print("and runs forecasting using the selected features.")
    print("="*60)

    # ===== USER CONFIGURATION =====
    # MODIFY THESE PATHS TO MATCH YOUR SETUP
    
    # Path to your data file
    data_file_path = "D:/Shared/desktop 3/Valley Fever/Weather Data MARICOPA (MMWR Weeks)/Clean Data(Maricopa)/Clean Processed Data2.csv"
    
    # Path to the analysis results directory (from the analysis script)
    # Example: "multi_horizon_results_2024-12-19_14-30-45_Clean Processed Data2"
    analysis_results_directory = input("📂 Enter the path to your analysis results directory: ").strip()
    
    # Horizons to forecast (should match what was analyzed)
    horizons_to_forecast = [2, 4, 8, 16]
    
    # ===== LOAD DATA =====
    print(f"\n📖 Loading data from: {data_file_path}")
    
    if not os.path.exists(data_file_path):
        print(f"❌ Error: Data file not found at {data_file_path}")
        print("💡 Please update the data_file_path variable with the correct path to your data file.")
        exit(1)
    
    try:
        data = pd.read_csv(data_file_path)
        data.drop('Date', axis=1, inplace=True, errors='ignore')
        print(f"✅ Data loaded successfully: {data.shape[0]} rows, {data.shape[1]} columns")
    except Exception as e:
        print(f"❌ Error loading data: {str(e)}")
        exit(1)
    
    # ===== LOAD ANALYSIS RESULTS =====
    loaded_results = load_analysis_results(analysis_results_directory)
    
    if loaded_results is None:
        print("\n❌ Failed to load analysis results. Please check the directory path and ensure you've run the stability analysis script first.")
        exit(1)
    
    # ===== SETUP NORMALIZATION (same as analysis script) =====
    start_index = 0
    train_end = 850
    initial_train_data = data.loc[start_index:train_end]
    
    # Apply same normalization as analysis script
    scaler = StandardScaler()
    scaler.fit(initial_train_data)
    df_norm = scaler.transform(data)
    df_norm = pd.DataFrame(df_norm, columns=data.columns, index=data.index)
    
    print(f"✅ Data normalized using training data range: {start_index} to {train_end}")
    
    # ===== CREATE FORECASTING DIRECTORY =====
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    forecasting_main_dir = f"forecasting_results_{now}"
    os.makedirs(forecasting_main_dir, exist_ok=True)
    
    print(f"\n📁 Forecasting results will be saved to: {forecasting_main_dir}")
    
    # ===== RUN FORECASTING FOR EACH HORIZON =====
    forecasting_results = {}
    
    available_horizons = list(loaded_results.keys())
    horizons_to_run = [h for h in horizons_to_forecast if h in available_horizons]
    
    if not horizons_to_run:
        print(f"❌ No matching horizons found!")
        print(f"   Requested: {horizons_to_forecast}")
        print(f"   Available: {available_horizons}")
        exit(1)
    
    print(f"\n🔮 Will run forecasting for horizons: {horizons_to_run}")
    
    for horizon in horizons_to_run:
        print(f"\n{'='*60}")
        print(f"🚀 Starting forecasting for {horizon}-week horizon")
        print(f"{'='*60}")
        
        # Create horizon-specific forecasting directory
        horizon_forecasting_dir = os.path.join(forecasting_main_dir, f"horizon_{horizon}")
        os.makedirs(horizon_forecasting_dir, exist_ok=True)
        
        # Get the loaded results for this horizon
        horizon_data = loaded_results[horizon]
        
        # Print selected features info
        print(f"📊 Selected Features Summary:")
        print(f"   Total features: {horizon_data['metadata']['total_features']}")
        print(f"   Selected features: {horizon_data['metadata']['selected_features_count']}")
        print(f"   Feature selection rate: {horizon_data['metadata']['feature_selection_percent']*100:.1f}%")
        
        # Show top 5 selected features
        feature_info_df = horizon_data['feature_info_df']
        top_features = feature_info_df.nlargest(5, 'importance')
        print(f"\n🔝 Top 5 Selected Features:")
        for idx, (_, row) in enumerate(top_features.iterrows(), 1):
            print(f"   {idx}. {row['feature_name']:<40} (Importance: {row['importance']:.4f})")
        
        try:
            # Run forecasting for this horizon
            output_file = run_forecasting_for_horizon(
                horizon=horizon,
                horizon_data=horizon_data,
                data=data,
                df_norm=df_norm,
                forecasting_dir=horizon_forecasting_dir
            )
            
            forecasting_results[horizon] = {
                'output_file': output_file,
                'forecasting_dir': horizon_forecasting_dir,
                'selected_features_count': horizon_data['metadata']['selected_features_count']
            }
            
            print(f"✅ Completed forecasting for {horizon}-week horizon")
            
        except Exception as e:
            print(f"❌ Error during forecasting for {horizon}-week horizon: {str(e)}")
            continue
    
    # ===== FINAL SUMMARY =====
    print(f"\n" + "="*80)
    print("📋 FORECASTING COMPLETION SUMMARY")
    print("="*80)
    
    if forecasting_results:
        print(f"✅ Successfully completed forecasting for {len(forecasting_results)} horizons:")
        
        for horizon, results in forecasting_results.items():
            print(f"\n🔮 {horizon}-Week Horizon:")
            print(f"   📊 Used {results['selected_features_count']} selected features")
            print(f"   📈 Results: {results['output_file']}")
            print(f"   📁 Directory: {results['forecasting_dir']}")
        
        print(f"\n📂 All results saved in: {forecasting_main_dir}")
        print(f"\n📈 Files generated:")
        print(f"   - forecasting_results_[horizon]week_[timestamp].xlsx (performance metrics)")
        print(f"   - prediction_[horizon]_it=[iteration].png (prediction plots)")
        print(f"   - model_weights_[horizon]week.h5 (trained model weights)")
        
    else:
        print("❌ No forecasting completed successfully")
    
    print("="*80)
    print("🎉 Multi-horizon forecasting script completed!")
    print("="*80)
print(f"\n🎉 All analysis completed!")
print(f"📂 Check results in directory: {main_directory}/")